from datetime import datetime, timedelta
from time import time
import copy

from fast_bitrix24 import Bitrix
import openpyxl


b = Bitrix('https://vc4dk.bitrix24.ru/rest/311/wkq0a0mvsvfmoseo/')

service_deal_current_month = ['Контрагент', 'Линк', 'МДЛП', 'Старт ЭДО']
service_deal_values = {'Контрагент': 4800, 'Кабинет сотрудника': None, 'Линк': None, 'МДЛП': None, '1Спарк 3000': 3000, '1СпаркПЛЮС 22500': 22500,
                          'Старт ЭДО': 3000, 'Подпись': 0, 'Подпись 1000': 1000, 'РПД': None, 'ЭТП': None, '1СПАРК Риски': 3000, '1Спарк в договоре': 3000,
                      'Кабинет садовода': 1000, '1Спарк': 3000, 'ЭДО': None, 'Спарк сумма': None}
spark_names = ['1Спарк', '1Спарк в договоре', '1СПАРК Риски', '1СпаркПЛЮС 22500', '1Спарк 3000']
service_deal_types = list(service_deal_values.keys())
titles_for_sorting = []
month_names = ['Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
month_number = {
    'Январь': '01',
    'Февраль': '02',
    'Март': '03',
    'Апрель': '04',
    'Май': '05',
    'Июнь': '06',
    'Июль': '07',
    'Август': '08',
    'Сентябрь': '09',
    'Октябрь': '10',
    'Ноябрь': '11',
    'Декабрь': '12',
}
handled_data = {}


def read_deals_from_xlsx(filename: str) -> list:
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    max_rows = worksheet.max_row
    max_columns = worksheet.max_column
    data = []
    titles = {}
    for row in range(1, max_rows + 1):
        temp = {}
        for column in range(1, max_columns):
            cell_value = worksheet.cell(row, column).value
            if row == 1:
                titles.setdefault(column, cell_value)
            else:
                temp.setdefault(titles[column], cell_value)
        if temp:
            data.append(temp)
    return data


def get_service_deal_start_dates(month: str, deal_type: str, deal_date_end, deal_date_start):
    current_year = datetime.now().year
    if deal_type in service_deal_current_month:
        return f'{month_number[month]}.{current_year}'
    elif deal_type == 'Кабинет садовода':
        return {'Месяц': int(month_number[month]), 'Год': int(current_year)}
    elif deal_type == 'Подпись 1000' and f'{deal_date_start.day}.{deal_date_start.month}' == f'{deal_date_end.day}.{deal_date_end.month}':
        return f'{deal_date_start.month}.{deal_date_start.year}'
    else:
        new_deal_date_start_year = deal_date_end.year - 1
        new_deal_date_start = deal_date_end + timedelta(days=1)
        return f"{new_deal_date_start.strftime('%m')}.{new_deal_date_start_year}"


def get_deal_value(deal_value, deal_type, deal_id):
    if deal_value:
        return deal_value
    else:
        if deal_type in service_deal_values:
            if service_deal_values[deal_type] is None:
                print(f'Нет суммы для {deal_id}')
                return 0
            return service_deal_values[deal_type]
        else:
            print(f'Нет суммы для {deal_id}')
            return 0


def add_month_edo_value(edo_list_elements=None, month=None, users_info=None):
    service_deal_value_field = f'{month} Сервисы'
    for responsible in handled_data:
        user_name = responsible.split()[0]
        try:
            user_last_name = responsible.split()[1]
        except IndexError:
            user_last_name = ''

        user_info = list(filter(lambda x: x['NAME'] == user_name and x['LAST_NAME'] == user_last_name, users_info))
        if not user_info:
            continue
        user_id = str(user_info[0]['ID'])
        edo_elements = list(filter(lambda x: x['Месяц'] == month and x['Ответственный'] == user_id, edo_list_elements))
        if edo_elements:
            for element in edo_elements:
                handled_data[responsible][service_deal_value_field] += element['Сумма']
                handled_data[responsible]['Сервисы'][f"{month} {'ЭДО'}"] += element['Сумма']


def deal_info_handler(deals_info, users_info, month, edo_list_elements=None):
    its_deal_value_field = f'{month} ИТС'
    service_deal_value_field = f'{month} Сервисы'
    rpd_data = dict(zip(list(handled_data.keys()), [{} for _ in handled_data.keys()]))
    for deal_info in deals_info:

        if deal_info['Ответственный'] not in handled_data:
            continue

        deal_info['Сумма'] = int(float(deal_info['Сумма']))

        if deal_info['Тип'] == 'ЭДО':
            continue

        if deal_info['Группа'] == 'ИТС' and deal_info['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'] and 'ГРМ' not in deal_info['Тип']:
            handled_data[deal_info['Ответственный']][its_deal_value_field] += 1

        elif deal_info['Тип'] in service_deal_types and deal_info['Стадия сделки'] == 'Услуга активна':
            deal_start_date = get_service_deal_start_dates(month, deal_info['Тип'], deal_info['Предполагаемая дата закрытия'], deal_info['Дата начала'])
            if deal_info['Тип'] in service_deal_current_month:
                if deal_start_date in deal_info['Дата начала'].strftime('%d.%m.%Y'):
                    deal_value = get_deal_value(deal_info['Сумма'], deal_info['Тип'], deal_info['ID'])
                    handled_data[deal_info['Ответственный']][service_deal_value_field] += deal_value
                    handled_data[deal_info['Ответственный']]['Сервисы'][f"{month} {deal_info['Тип']}"] += deal_value

            elif deal_info['Тип'] == 'Кабинет садовода':
                deal_start_date = get_service_deal_start_dates(month, deal_info['Тип'],
                                                               deal_info['Предполагаемая дата закрытия'], deal_info['Дата начала'])
                deal_end_date_month = deal_info['Предполагаемая дата закрытия'].month
                deal_end_date_year = deal_info['Предполагаемая дата закрытия'].year
                if deal_start_date['Год'] == deal_end_date_year:
                    if deal_start_date['Месяц'] > deal_end_date_month:
                        continue
                elif deal_start_date['Год'] > deal_end_date_year:
                    continue
                deal_value = get_deal_value(deal_info['Сумма'], deal_info['Тип'], deal_info['ID'])
                handled_data[deal_info['Ответственный']][service_deal_value_field] += deal_value
                handled_data[deal_info['Ответственный']]['Сервисы'][f"{month} {deal_info['Тип']}"] += deal_value

            else:
                if deal_start_date == f'{month_number[month]}.2022':
                    if deal_info['Тип'] == 'Подпись 1000':
                        deal_value = 600
                    elif deal_info['Тип'] == 'Подпись':
                        deal_value = 0
                    elif deal_info['Тип'] == 'РПД':
                        rpd_data[deal_info['Ответственный']][deal_info['ID']] = get_deal_value(deal_info['Сумма'], deal_info['Тип'], deal_info['ID'])
                        continue
                    else:
                        deal_value = get_deal_value(deal_info['Сумма'], deal_info['Тип'], deal_info['ID'])
                    handled_data[deal_info['Ответственный']][service_deal_value_field] += deal_value
                    handled_data[deal_info['Ответственный']]['Сервисы'][f"{month} {deal_info['Тип']}"] += deal_value

    for responsible in rpd_data:
        rpd_values = list(rpd_data[responsible].values())
        handled_data[responsible]['Сервисы'][f'{month} РПД'] = sum(rpd_values)

    add_month_edo_value(edo_list_elements, month, users_info)


def get_services_summary():
    month_services_summary = {}
    for month in month_names:
        spark_names_with_months = list(map(lambda x: f'{month} {x}', spark_names))
        services_summary = dict(zip(service_deal_types, [0 for _  in service_deal_types]))
        for service_name in service_deal_types:
            if service_name == 'Спарк сумма':
                continue
            month_service_name = f'{month} {service_name}'
            for responsible in handled_data:
                for responsible_service_name in handled_data[responsible]['Сервисы']:
                    if month_service_name == responsible_service_name:
                        services_summary[service_name] += handled_data[responsible]['Сервисы'][responsible_service_name]
                        for spark_name in spark_names_with_months:
                            if spark_name == responsible_service_name:
                                services_summary['Спарк сумма'] += handled_data[responsible]['Сервисы'][responsible_service_name]

        month_services_summary.setdefault(month, services_summary)
    return month_services_summary


def write_data_to_xlsx(data, month_titles=None, service_titles=None, month_count=None):
    service_names = copy.deepcopy(service_deal_types)
    workbook = openpyxl.Workbook()
    filename = f'Отчет по сумме сервисов{time()}.xlsx'.replace(' ', '_')
    departments = ['ГО4', 'ГО3', 'ОВ', 'Прочие']

    worksheet = workbook.active
    worksheet.title = 'Сводные показатели'
    summary_titles = ['Сотрудник', 'Подразделение', f'ИТС сред({month_count})', f'Сумма сервисов сред({month_count})', 'Результат']
    employee_to_delete = []
    worksheet.append(summary_titles)
    for department in departments:
        for employee in data:
            row = []
            its_value = 0
            service_value = 0
            for field in data[employee]:
                if 'ИТС' in field:
                    its_value += data[employee][field]
                elif 'Сервисы' in field and field != 'Сервисы':
                    service_value += data[employee][field]

            try:
                summary_its = round(its_value / month_count, 2)
            except ZeroDivisionError:
                summary_its = 0
            if summary_its == 0:
                employee_to_delete.append(employee)
                continue
            try:
                summary_service = round(service_value / month_count, 2)
            except ZeroDivisionError:
                summary_service = 0
            try:
                result_value = round(service_value / its_value, 2)
            except ZeroDivisionError:
                result_value = 0

            if data[employee]['Подразделение'] == department:
                row = [employee, data[employee]['Подразделение'], summary_its, summary_service, result_value]
            elif department == 'Прочие' and data[employee]['Подразделение'] not in departments:
                row = [employee, data[employee]['Подразделение'], summary_its, summary_service, result_value]
            if row:
                worksheet.append(row)

    for employee in employee_to_delete:
        data.pop(employee, None)

    month_report = copy.deepcopy(data)
    service_report = copy.deepcopy(data)

    worksheet = workbook.create_sheet('Детализация по месяцам')
    if month_titles:
        worksheet.append(month_titles)
    for department in departments:
        for employee in month_report:
            month_report[employee].pop('Сервисы', None)
            row = []
            if month_report[employee]['Подразделение'] == department:
                row = [employee,] + list(month_report[employee].values())
            elif department == 'Прочие' and month_report[employee]['Подразделение'] not in departments:
                row = [employee, ] + list(month_report[employee].values())
            if row:
                worksheet.append(row)

    worksheet = workbook.create_sheet('Детализация по сервисам')
    worksheet.append(service_titles)
    services_summary = get_services_summary()
    first_employee_row = True
    service_title_count = 0
    for department in departments:
        worksheet.append([department, ])
        for employee in service_report:
            if service_report[employee]['Подразделение'] == department:
                if first_employee_row:
                    worksheet.append(['', employee, ] + ['' for _ in month_names] + ['', '', '', ''] + month_names)
                else:
                    worksheet.append(['', employee,])
                for service_name in service_names:
                    data_to_write = ['', '', service_name]
                    service_sum = 0
                    for service in service_report[employee]['Сервисы']:
                        if service_name == ' '.join(service.split()[1:]):
                            data_to_write.append(service_report[employee]['Сервисы'][service])
                            service_sum += service_report[employee]['Сервисы'][service]
                    if service_name == 'Спарк сумма':
                        data_to_write = ['', '', '', ''] + ['' for _ in month_names]
                    else:
                        data_to_write.append(service_sum)

                    if first_employee_row:
                        if service_title_count < len(service_deal_types):
                            data_to_write += ['', service_deal_types[service_title_count]]
                            for month_name in services_summary:
                                for summary_service_name in services_summary[month_name]:
                                    if summary_service_name == service_deal_types[service_title_count]:
                                        data_to_write.append(services_summary[month_name][summary_service_name])
                    service_title_count += 1

                    worksheet.append(data_to_write)
                first_employee_row = False

            elif department == 'Прочие' and service_report[employee]['Подразделение'] not in departments:
                worksheet.append(['', employee, ])
                for service_name in service_names:
                    data_to_write = ['', '', service_name]
                    service_sum = 0
                    for service in service_report[employee]['Сервисы']:
                        if service_name == ' '.join(service.split()[1:]):
                            data_to_write.append(service_report[employee]['Сервисы'][service])
                            service_sum += service_report[employee]['Сервисы'][service]
                    data_to_write.append(service_sum)
                    worksheet.append(data_to_write)
    workbook.save(filename)


def get_edo_list_elements():
    edo_list_month_codes = {
        '2371': 'Январь',
        '2373': 'Февраль',
        '2375': 'Март',
        '2377': 'Апрель',
        '2379': 'Май',
        '2381': 'Июнь',
        '2383': 'Июль',
        '2385': 'Август',
        '2387': 'Сентябрь',
        '2389': 'Октябрь',
        '2391': 'Ноябрь',
        '2393': 'Декабрь',
    }
    edo_list_year_codes = {
        '2395': '2022',
        '2397': '2023',
    }
    edo_list_elements = b.get_all('lists.element.get', {'IBLOCK_TYPE_ID': 'lists', 'IBLOCK_ID': '235'})
    edo_list_elements = list(map(lambda x: {
        'ID': x['ID'],
        'Месяц': edo_list_month_codes[list(x['PROPERTY_1567'].values())[0]],
        'Год': edo_list_year_codes[list(x['PROPERTY_1569'].values())[0]],
        'Ответственный': list(x['PROPERTY_1581'].values())[0] if 'PROPERTY_1581' in x else '',
        'Сумма': int(list(x['PROPERTY_1575'].values())[0]),
    }, edo_list_elements))

    return edo_list_elements


def find_all_responsibles(file_data, month, users_info):
    department_names = {29: 'ГО4', 27: 'ГО3', 5: 'ЦС', 231: 'ЛК', 225: 'ОВ', 1: '4DK', 233: 'Служебные'}
    ignore_names = ['Иван Иванов', 'Максим Карпов', 'Борис Ишкин', 'Отчет Сервисный выезд', 'Робот Задач']
    its_deal_value_field = f'{month} ИТС'
    service_deal_value_field = f'{month} Сервисы'
    for deal_info in file_data:
        if deal_info['Ответственный'] in ignore_names or not deal_info['Ответственный']:
            continue

        user_name = deal_info['Ответственный'].split()[0]
        try:
            user_last_name = deal_info['Ответственный'].split()[1]
        except IndexError:
            user_last_name = ''

        deal_info['Сумма'] = int(float(deal_info['Сумма']))
        if deal_info['Ответственный'] not in handled_data and deal_info['Ответственный']:
            handled_data.setdefault(deal_info['Ответственный'], {})
            user_info = list(filter(lambda x: x['NAME'] == user_name and x['LAST_NAME'] == user_last_name, users_info))
            if user_info[0]['UF_DEPARTMENT'][0] not in department_names:
                continue
            if deal_info['Ответственный'] == 'Светлана Ридкобород':
                handled_data[deal_info['Ответственный']].setdefault('Подразделение', 'ОВ')
            else:
                handled_data[deal_info['Ответственный']].setdefault('Подразделение', department_names[user_info[0]['UF_DEPARTMENT'][0]])

        if its_deal_value_field not in deal_info['Ответственный']:
            handled_data[deal_info['Ответственный']].setdefault(its_deal_value_field, 0)
            handled_data[deal_info['Ответственный']].setdefault(service_deal_value_field, 0)
            handled_data[deal_info['Ответственный']].setdefault(f'Сервисы', {})
            service_deal_types_with_month = copy.deepcopy(service_deal_types)
            service_deal_types_with_month.append('ЭДО')
            service_zero_values = [0 for _ in service_deal_types_with_month]
            service_deal_types_with_month = list(map(lambda x: f'{month} {x}', service_deal_types_with_month))
            service_keys_values = dict(zip(service_deal_types_with_month, service_zero_values))

            for key in service_keys_values:
                handled_data[deal_info['Ответственный']]['Сервисы'].setdefault(key, service_keys_values[key])

        titles_for_sorting.append(month)


def add_month_titles_to_responsibles():
    for month in titles_for_sorting:
        its_deal_value_field = f'{month} ИТС'
        service_deal_value_field = f'{month} Сервисы'
        for responsible in handled_data:
            if its_deal_value_field not in handled_data[responsible]:
                handled_data[responsible].setdefault(its_deal_value_field, 0)
                handled_data[responsible].setdefault(service_deal_value_field, 0)
                handled_data[responsible].setdefault(f'Сервисы', {})
                service_deal_types_with_month = copy.deepcopy(service_deal_types)
                service_deal_types_with_month.append('ЭДО')
                service_zero_values = [0 for _ in service_deal_types_with_month]
                service_deal_types_with_month = list(map(lambda x: f'{month} {x}', service_deal_types_with_month))
                service_keys_values = dict(zip(service_deal_types_with_month, service_zero_values))

                for key in service_keys_values:
                    handled_data[responsible]['Сервисы'].setdefault(key, service_keys_values[key])


def sort_handled_data_keys():
    sorted_handled_data = {}
    for responsible in handled_data:
        service_dct = {}
        sorted_handled_data.setdefault(responsible, {})
        for title in titles_for_sorting:
            for key in handled_data[responsible]:
                if key == 'Сервисы':
                    for service in handled_data[responsible]['Сервисы']:
                        if title in service:
                            service_dct[service] = handled_data[responsible]['Сервисы'][service]
                if title in key:
                    sorted_handled_data[responsible][key] = handled_data[responsible][key]
        sorted_handled_data[responsible].setdefault('Сервисы', service_dct)

    return sorted_handled_data


def create_report_service_sales():
    global handled_data
    current_month = ''
    current_year = ''
    users_data = b.get_all('user.get')
    edo_list_elements = get_edo_list_elements()
    filenames = ['DEAL_20220731_6d490055_62e6c5642f727TONIGHT.xlsx', 'DEAL_20220831_66154768_630ef62fa45bb.xlsx', '300920222.xlsx', 'DEAL_20221031_d0ba71b5_635f5c9e736a7.xlsx', 'DEAL_20221130_a67f2b14_6386e8c656b60.xlsx', 'DEAL_20221231_d33ea5db_63afc361a0eba.xlsx']
    file_months = {'DEAL_20220731_6d490055_62e6c5642f727TONIGHT.xlsx': 'Июль', 'DEAL_20220831_66154768_630ef62fa45bb.xlsx': 'Август', '300920222.xlsx': 'Сентябрь', 'DEAL_20221031_d0ba71b5_635f5c9e736a7.xlsx': 'Октябрь', 'DEAL_20221130_a67f2b14_6386e8c656b60.xlsx': 'Ноябрь', 'DEAL_20221231_d33ea5db_63afc361a0eba.xlsx': 'Декабрь'}
    for filename in filenames:
        print(f'Поиск ответственных за {file_months[filename]}')
        file_data = read_deals_from_xlsx(filename)
        find_all_responsibles(file_data, file_months[filename], users_data)
    add_month_titles_to_responsibles()
    titles_for_sorting.insert(0, 'Подразделение')
    handled_data = sort_handled_data_keys()
    for filename in filenames:
        print(f'Подсчет сделок за {file_months[filename]}')
        file_data = read_deals_from_xlsx(filename)
        deal_info_handler(file_data, users_data, file_months[filename], edo_list_elements=edo_list_elements)
    month_titles = ['Сотрудник', 'Подразделение', 'Июль ИТС', 'Июль Сервисы', 'Август ИТС', 'Август Сервисы', 'Сентябрь ИТС', 'Сентябрь Сервисы', 'Октябрь ИТС', 'Октябрь Сервисы', 'Ноябрь ИТС', 'Ноябрь Сервисы', 'Декабрь ИТС', 'Декабрь Сервисы']
    service_titles = ['Подразделение', 'ФИО', 'Сервис', 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь', 'Сумма']
    write_data_to_xlsx(handled_data, month_titles=month_titles, service_titles=service_titles, month_count=len(filenames))


create_report_service_sales()

