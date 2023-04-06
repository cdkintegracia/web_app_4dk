from fast_bitrix24 import Bitrix
import openpyxl
from time import sleep
from authentication import authentication



b = Bitrix(authentication('Bitrix'))


reporting_deal_types = {
    'UC_O99QUW': 'Отчетность',
    'UC_OV4T7K': 'Отчетность (в рамках ИТС)',
}
result = [['Регномер', 'Отчетность', 'Ответственный за ИТС', 'Название компании']]
workbook = openpyxl.load_workbook('Анализ регномеров (1) (1).xlsx')
worksheet = workbook.active
max_rows = worksheet.max_row
regnumbers = []
for row in range(2, max_rows + 1):
    regnumber = worksheet.cell(row=row, column=1).value
    regnumbers.append(regnumber)
no_its = []
deals = b.get_all('crm.deal.list', {'select': ['*', 'UF_*'], 'filter': {'UF_CRM_1640523562691': regnumbers, '!STAGE_ID': ['C1:WON']}})
companies_id = list(map(lambda x: x['COMPANY_ID'], deals))
companies = b.get_all('crm.company.list', {'filter': {'ID': companies_id}})
users = b.get_all('user.get')
for regnumber in regnumbers:
    reporting_deal_name = 'Нет'
    its_deal_responsible = 'Нет'
    regnumber_its_stage = '1491'
    regnumber_reporting_deals = list(filter(lambda x: x['TYPE_ID'] in ['UC_O99QUW', 'UC_OV4T7K'] and str(x['UF_CRM_1640523562691']) == str(regnumber), deals))
    if regnumber_reporting_deals:
        reporting_deal_name = reporting_deal_types[regnumber_reporting_deals[0]['TYPE_ID']]
    regnumber_its = list(filter(lambda x: x['UF_CRM_1657878818384'] == '859' and str(x['UF_CRM_1640523562691']) == str(regnumber), deals))
    if regnumber_its:
        user_info = list(filter(lambda x: x['ID'] == regnumber_its[0]['ASSIGNED_BY_ID'], users))
        its_deal_responsible = f"{user_info[0]['NAME']} {user_info[0]['LAST_NAME']}"
        company_name = list(filter(lambda x: x['ID'] == regnumber_its[0]['COMPANY_ID'], companies))[0]['TITLE']
        row = [regnumber, reporting_deal_name, its_deal_responsible, company_name]
        if regnumber_its[0]['STAGE_ID'] in ['C1:NEW', 'C1:UC_0KJKTY', 'C1:UC_3J0IH6']:
            regnumber_its_stage = '1463'

        new_deal = b.call('crm.deal.add', {'fields': {
            'CATEGORY_ID': '15',
            'TITLE': 'Продажа',
            'COMPANY_ID': regnumber_its[0]['COMPANY_ID'],
            'UF_CRM_1680181516033': '1461',
            #'UF_CRM_1680191120': regnumber_its[0]['ASSIGNED_BY_ID'],
            'UF_CRM_1640523562691': regnumber,
            'UF_CRM_1680253793699': regnumber_its_stage,
            'ASSIGNED_BY_ID': regnumber_its[0]['ASSIGNED_BY_ID'],
            'CLOSEDATE': '2023-04-21'
        }})
        sleep(5)
        b.call('crm.deal.update', {'ID': new_deal, 'fields': {
            'UF_CRM_1680181516033': '1461',
            #'UF_CRM_1680191120': regnumber_its[0]['ASSIGNED_BY_ID'],
            'UF_CRM_1680253793699': regnumber_its_stage,
            'UF_CRM_1670409896': regnumber,
        }})
    else:
        no_its.append(regnumber)
for i in no_its:
    print(i)
exit()

    #result.append([regnumber, reporting_deal_name, its_deal_responsible, company_name])

workbook = openpyxl.Workbook()
worksheet = workbook.active
for row in result:
    worksheet.append(row)
workbook.save('Анализ регномеров.xlsx')





