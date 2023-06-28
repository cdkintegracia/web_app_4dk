from fast_bitrix24 import Bitrix
import openpyxl

from authentication import authentication


b = Bitrix(authentication('Bitrix'))


def read_xlsx(xlsx_name):
    data = []
    workbook = openpyxl.load_workbook(xlsx_name)
    worksheet = workbook.active
    for row in range(1, worksheet.max_row + 1):
        for column in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row, column).value
            if row == 1 or column != 5 or not cell:
                continue
            else:
                cell = cell.split('\n')
                data.append(cell[1])
    return data


def create_summer_2023_sales_tasks():

    group_a = []
    group_b = []
    group_v = []
    group_g = []
    group_d = []
    new_group_a = []

    revenue_elements = b.get_all('lists.element.get', {
        'IBLOCK_TYPE_ID': 'lists',
        'IBLOCK_ID': '257',
        'filter': {
            'PROPERTY_1643': '2022'
        }
    })
    for element in revenue_elements:
        if 'PROPERTY_1627' in element:
            okved = list(element['PROPERTY_1627'].values())[0]
            if 10 <= int(okved.split('.')[0]) <= 33:
                group_a.append(list(element['PROPERTY_1631'].values())[0])

    unf_company_names = read_xlsx('УНФ.xlsx')
    ut_company_names = read_xlsx('УТ.xlsx')

    unf_company_info = b.get_all('crm.company.list', {
        'select': ['ID'],
        'filter':
            {'TITLE': unf_company_names}
    })
    unf_companies_id = list(set(map(lambda x: x['ID'], unf_company_info)))

    ut_company_info = b.get_all('crm.company.list', {
        'select': ['ID'],
        'filter':
            {'TITLE': ut_company_names}
    })
    ut_companies_id = list(set(map(lambda x: x['ID'], ut_company_info)))

    for company in ut_companies_id:
        if company in group_a:
            group_b.append(company)

    for company in unf_companies_id:
        if company in group_a:
            group_v.append(company)

    for company in ut_companies_id:
        if company not in group_b:
            group_g.append(company)

    for company in unf_companies_id:
        if company not in group_v:
            group_d.append(company)

    for company in group_a:
        if company in group_b or company in group_v:
            continue
        else:
            new_group_a.append(company)

    companies_info = b.get_all('crm.company.list', {
        'select': ['TITLE'],
        'filter': {
            'ID': group_a + group_b + group_v + group_g + group_d
        }
    })

    deals = b.get_all('crm.deal.list', {
        'filter': {
            'CATEGORY_ID': '1',
            'STAGE_ID': 'C1:NEW',
            'COMPANY_ID': group_a + group_b + group_v + group_g + group_d
        }
    })
    deals = list(set(map(lambda x: x['COMPANY_ID'], deals)))
    '''
    for company in new_group_a:
        if company not in deals:
            continue

        company_revenue = 'неизвестно'
        revenue = b.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '257',
            'filter': {
                'PROPERTY_1643': '2022',
                'PROPERTY_1631': company,
            }
        })
        if revenue and 'PROPERTY_1635' in revenue[0]:
            company_revenue = list(revenue[0]['PROPERTY_1635'].values())[0]
            if company_revenue == '-0':
                company_revenue = 'неизвестно'

        b.call('tasks.task.add', {
            'fields': {
                'GROUP_ID': '133',
                'TITLE': f"{list(filter(lambda x: x['ID'] == company, companies_info))[0]['TITLE']} летние продажи 2023",
                'UF_CRM_TASK': ['CO_' + company],
                'DEADLINE': '2023-07-31 19:00:00',
                'DESCRIPTION': f'У этого клиента основная деятельность - производство. Клиент не использует УТ или УНФ. Предложите ему автоматизацию производства, уточните требования к автоматизации Гособоронзаказа (это может быть проект или кейс по ГОЗ).\n\n Выручка (в млн) = {company_revenue}',
                'CREATED_BY': '159',
                'RESPONSIBLE_ID': '405'
            }
        })
    '''
    for company in group_b:
        if company not in deals:
            continue
        company_revenue = 'неизвестно'
        revenue = b.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '257',
            'filter': {
                'PROPERTY_1643': '2022',
                'PROPERTY_1631': company,
            }
        })
        if revenue and 'PROPERTY_1635' in revenue[0]:
            company_revenue = list(revenue[0]['PROPERTY_1635'].values())[0]
            if company_revenue == '-0':
                company_revenue = 'неизвестно'

        b.call('tasks.task.add', {
                'fields': {
                    'GROUP_ID': '133',
                    'TITLE': f"{list(filter(lambda x: x['ID'] == company, companies_info))[0]['TITLE']} летние продажи 2023",
                    'UF_CRM_TASK': ['CO_' + company],
                    'DEADLINE': '2023-07-31 19:00:00',
                    'DESCRIPTION': f'У этого клиента основная деятельность - производство. Клиент использует УТ. Уточните версию продукта, предложите ему переход на УТ 11 и\или автоматизацию производства (проект\консалтинг), уточните требования к автоматизации Гособоронзаказа (это может быть проект или кейс по ГОЗ).\n\n Выручка (в млн) = {company_revenue}',
                    'CREATED_BY': '159',
                    'RESPONSIBLE_ID': '405'
                }
            })

    for company in group_v:
        if company not in deals:
            continue
            
        company_revenue = 'неизвестно'
        revenue = b.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '257',
            'filter': {
                'PROPERTY_1643': '2022',
                'PROPERTY_1631': company,
            }
        })
        if revenue and 'PROPERTY_1635' in revenue[0]:
            company_revenue = list(revenue[0]['PROPERTY_1635'].values())[0]
            if company_revenue == '-0':
                company_revenue = 'неизвестно'

        b.call('tasks.task.add', {
                'fields': {
                    'GROUP_ID': '133',
                    'TITLE': f"{list(filter(lambda x: x['ID'] == company, companies_info))[0]['TITLE']} летние продажи 2023",
                    'UF_CRM_TASK': ['CO_' + company],
                    'DEADLINE': '2023-07-31 19:00:00',
                    'DESCRIPTION': f'У этого клиента основная деятельность - производство. Клиент использует УНФ. Предложите ему автоматизацию производства (кейс\проект\консалтинг), уточните требования к автоматизации Гособоронзаказа (это может быть проект или кейс по ГОЗ).\n\n Выручка (в млн) = {company_revenue}',
                    'CREATED_BY': '159',
                    'RESPONSIBLE_ID': '405'
                }
            })

    for company in group_g:
        if company not in deals:
            continue

        company_revenue = 'неизвестно'
        revenue = b.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '257',
            'filter': {
                'PROPERTY_1643': '2022',
                'PROPERTY_1631': company,
            }
        })
        if revenue and 'PROPERTY_1635' in revenue[0]:
            company_revenue = list(revenue[0]['PROPERTY_1635'].values())[0]
            if company_revenue == '-0':
                company_revenue = 'неизвестно'

        b.call('tasks.task.add', {
                'fields': {
                    'GROUP_ID': '133',
                    'TITLE': f"{list(filter(lambda x: x['ID'] == company, companies_info))[0]['TITLE']} летние продажи 2023",
                    'UF_CRM_TASK': ['CO_' + company],
                    'DEADLINE': '2023-07-31 19:00:00',
                    'DESCRIPTION': f'У клиента куплена УТ, задача - развить клиента\n\n Выручка (в млн) = {company_revenue}',
                    'CREATED_BY': '159',
                    'RESPONSIBLE_ID': '393'
                }
            })

    for company in group_d:
        if company not in deals:
            continue
            
        company_revenue = 'неизвестно'
        revenue = b.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '257',
            'filter': {
                'PROPERTY_1643': '2022',
                'PROPERTY_1631': company,
            }
        })
        if revenue and 'PROPERTY_1635' in revenue[0]:
            company_revenue = list(revenue[0]['PROPERTY_1635'].values())[0]
            if company_revenue == '-0':
                company_revenue = 'неизвестно'

        b.call('tasks.task.add', {
                'fields': {
                    'GROUP_ID': '133',
                    'TITLE': f"{list(filter(lambda x: x['ID'] == company, companies_info))[0]['TITLE']} летние продажи 2023",
                    'UF_CRM_TASK': ['CO_' + company],
                    'DEADLINE': '2023-07-31 19:00:00',
                    'DESCRIPTION': f'У клиента куплена УНФ, задача - развить клиента\n\n Выручка (в млн) = {company_revenue}',
                    'CREATED_BY': '159',
                    'RESPONSIBLE_ID': '393'
                }
            })


if __name__ == '__main__':
    create_summer_2023_sales_tasks()
