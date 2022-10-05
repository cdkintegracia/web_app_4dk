import datetime
from time import time

from fast_bitrix24 import Bitrix
import openpyxl

from web_app_4dk.authentication import authentication

workbook = openpyxl.load_workbook("/root/web_app_4dk/web_app_4dk/new_call_statistic.xlsx")
worksheet = workbook.active
b = Bitrix(authentication('Bitrix'))

def rewrite_call_statistic(month, year):
    month_codes = {
            'Январь': '2215',
            'Февраль': '2217',
            'Март': '2219',
            'Апрель': '2221',
            'Май': '2223',
            'Июнь': '2225',
            'Июль': '2227',
            'Август': '2229',
            'Сентябрь': '2231',
            'Октябрь': '2233',
            'Ноябрь': '2235',
            'Декабрь': '2237'
        }
    year_codes = {
        '2022': '2239',
        '2023': '2241'
    }
    elements = b.get_all('lists.element.get', {'IBLOCK_TYPE_ID': 'lists', 'IBLOCK_ID': '175'})
    for element in elements:
        print(element)
        return
    return
    companies = b.get_all('crm.company.list', {'select': ['TITLE']})

    errors = []
    data = []
    for row in range(2, worksheet.max_row + 1):
        temp = []
        for col in range(1, 4):
            value = worksheet.cell(row, col).value
            if isinstance(value, datetime.time):
                value = value.strftime('%H:%M:%S')
            elif type(value) is str:
                company_name = value.strip('ООО')
                for company in companies:
                    if company_name in company['TITLE']:
                        value = company['ID']
            if value is None:
                break
            temp.append(value)
        if len(temp) != 3 or temp[2] == 0:
            continue
        if not temp[0].isdigit():
            errors.append(temp)
        else:
            data.append(temp)

    for d in data:
        b.call('lists.element.add', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '175',
            'ELEMENT_CODE': time(),
            'fields': {
                'NAME': 'Август 2022',  # Название == месяц и год
                'PROPERTY_1303': d[1],  # Продолжительность звонка
                'PROPERTY_1299': d[0],  # Привязка к компании
                'PROPERTY_1305': d[2],  # Количество звонков
                'PROPERTY_1339': month_codes[month],    # Месяц
                'PROPERTY_1341': year_codes[year],    # Год
            }
        }
               )