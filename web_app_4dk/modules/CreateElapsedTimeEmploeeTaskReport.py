from datetime import datetime, timedelta
import base64
import os

from fast_bitrix24 import Bitrix
import openpyxl
from openpyxl.utils import get_column_letter

from web_app_4dk.modules.authentication import authentication


b = Bitrix(authentication('Bitrix'))


def create_eladsed_time_employee_report(req):
    user_id = req['user'][5:]
    print('user_id')
    user_info = b.get_all('user.get', {
        'filter': {
            'ID': user_id
        }
    })

    company_info = b.get_all('crm.company.get', {
        'ID': req['company_id'],
        'select': ['TITLE'],
    })

    tasks = b.get_all('tasks.task.list', {
        'filter': {
            '>=CREATED_DATE': req['date_filter'],
            'UF_CRM_TASK': ['CO_' + req['company_id']],
            'RESPONSIBLE_ID': user_id
        }
    })

    tasks_id = list(map(lambda x: x['id'], tasks))
    tasks_results = list()
    for i in tasks_id:
        results = b.get_all('task.elapseditem.getlist', {
            'TASKID': i
        })
        tasks_results.extend(results)
    tasks_results = list(sorted(tasks_results, key=lambda x: datetime.fromisoformat(x['CREATED_DATE'])))

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    users_name = dict()
    worksheet.append([
        'Дата начала', 'Ответственный', 'Затраченное время', 'Работа'
    ])
    for result in tasks_results:
        if result['USER_ID'] == user_id:
            if result['USER_ID'] not in users_name:
                user_info = list(filter(lambda x: x['ID'] == result['USER_ID'], user_info))[0]
                users_name[result['USER_ID']] = f'{user_info["NAME"]} {user_info["LAST_NAME"]}'
            worksheet.append([
            datetime.fromisoformat(result['CREATED_DATE']).strftime('%d.%m.%Y %H:%M'),
            users_name[result['USER_ID']],
            str(timedelta(seconds=int(result['SECONDS']))),
            result['COMMENT_TEXT']
            ])

        columns_width = {
            0: 25,
            1: 30,
            2: 25,
            3: 50
        }
        for index, column_cells in enumerate(worksheet.columns):
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = columns_width[index]

        for index, column_cells in enumerate(worksheet.columns):
            for cell in column_cells:
                cell.alignment = cell.alignment.copy(wrapText=True)
    report_name = f'{company_info["TITLE"].replace(" ", "_")}_анализ_трудозатрат_на_{datetime.now().strftime("%d.%m.%Y_%S")}.xlsx'
    workbook.save(report_name)

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '1077329'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    b.call('im.notify.system.add', {
        'USER_ID': req['user_id'][5:],
        'MESSAGE': f'Анализ трудозатрат для {company_info["TITLE"]} сформирован. {upload_report["DETAIL_URL"]}'})
    os.remove(report_name)
