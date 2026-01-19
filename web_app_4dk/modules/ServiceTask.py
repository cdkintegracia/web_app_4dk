from calendar import monthrange
from datetime import datetime, timedelta
import base64
from os import remove as os_remove
from time import sleep

from fast_bitrix24 import Bitrix
import openpyxl
from openpyxl.utils import get_column_letter
import requests

try:
    from web_app_4dk.modules.authentication import authentication
    from web_app_4dk.tools import send_bitrix_request
except ModuleNotFoundError:
    from authentication import authentication


    def send_bitrix_request(method: str, data: dict):
        try:
            return requests.post(f"{authentication('Bitrix')}{method}", json=data).json()['result']
        except KeyError:
            return



        # Считывание файла authentication.txt

webhook = authentication('Bitrix')
b = Bitrix(webhook)


months = {
    'Январь': '01',
    'Февраль': '02',
    'Март': '03',
    'Апрель': '04',
    'Май': '05',
    'Июнь': '06',
    'Июль': '07',
    'Август': '08',
    'Сентябрь': '09',
    'Октябрь': '10',
    'Ноябрь': '11',
    'Декабрь': '12'
}

int_to_months = {
    1: 'Январь',
    2: 'Февраль',
    3: 'Март',
    4: 'Апрель',
    5: 'Май',
    6: 'Июнь',
    7: 'Июль',
    8: 'Август',
    9: 'Сентябрь',
    10: 'Октябрь',
    11: 'Ноябрь',
    12: 'Декабрь',
}


def get_employee_id(employees: str) -> list:
    id_list = set()
    id_employees = employees.split(', ')  # Строка с сотрудниками и отделами в список
    for id in id_employees:
        if 'user' in id:  # Если в массиве найден id сотрудника
            id_list.add(id[5:])
        elif 'group' in id:  # Если в массиве найден id отдела
            department_users = b.get_all('user.get', {'filter': {'UF_DEPARTMENT': id[8:]}})
            for user in department_users:
                id_list.add(user['ID'])
    id_list = list(id_list)
    return(id_list)


def get_quarter_deals_for_service_tasks(date_start, date_end, type_deals, employees, stages):
    deal_types = [
                    'UC_HT9G9H',    # ПРОФ Земля
                    'UC_XIYCTV',    # ПРОФ Земля+Помощник
                    'UC_5T4MAW',    # ПРОФ Земля+Облако+Помощник
                    'UC_N113M9',    # ПРОФ Земля+Облако
                    'UC_ZKPT1B',    # ПРОФ Облако
                    'UC_2SJOEJ',    # ПРОФ Облако+Помощник
                    'UC_92H9MN',    # Индивидуальный
                    'UC_7V8HWF',    # Индивидуальный+Облако
                ]

    if employees == '':     # Если не были выбраны сотрудники в параметрах БП

        # Начались в сентябре 2022 и заканчиваются после сентября 2022

        deals_start_in_end_after = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '>BEGINDATE': date_start,
                    '<BEGINDATE': date_end,
                    '>CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '0',    # Помощник
                    'TYPE_ID': deal_types,
                    'STAGE_ID': stages
                }
            }
        )

        # начались до сентября 2022 и заканчиваются в сентябре 2022

        deals_start_before_end_in = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '<BEGINDATE': date_start,
                    '>CLOSEDATE': date_start,
                    '<CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '0',    # Помощник
                    'TYPE_ID': deal_types,
                    'STAGE_ID': stages
                }
            }
        )

        # начались до сентября 2022 и заканчиваются после сентября 2022

        deals_start_before_end_after = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '<BEGINDATE': date_start,
                    '>CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '0',    # Помощник
                    'TYPE_ID': deal_types,
                    'STAGE_ID': stages
                }
            }
        )
    else:   # Если были выбраны сотрудники в параметрах БП
        id_list = get_employee_id(employees)

        # Начались в сентябре 2022 и заканчиваются после сентября 2022

        deals_start_in_end_after = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '>BEGINDATE': date_start,
                    '<BEGINDATE': date_end,
                    '>CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '0',    # Помощник
                    'TYPE_ID': deal_types,
                    'ASSIGNED_BY_ID': id_list,
                    'STAGE_ID': stages
                }
            }
        )

        # начались до сентября 2022 и заканчиваются в сентябре 2022

        deals_start_before_end_in = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '<BEGINDATE': date_start,
                    '>CLOSEDATE': date_start,
                    '<CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '0',    # Помощник
                    'TYPE_ID': deal_types,
                    'ASSIGNED_BY_ID': id_list,
                    'STAGE_ID': stages
                }
            }
        )

        # начались до сентября 2022 и заканчиваются после сентября 2022

        deals_start_before_end_after = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '<BEGINDATE': date_start,
                    '>CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '0',    # Помощник
                    'TYPE_ID': deal_types,
                    'ASSIGNED_BY_ID': id_list,
                    'STAGE_ID': stages
                }
            }
        )

    return deals_start_in_end_after + deals_start_before_end_after + deals_start_before_end_in


def get_deals_for_service_tasks(date_start, date_end, type_deals, employees, stages):

    """
    Функция, которая вызывается из функции create_task_service

    :param date_start: Дата начала фильтрации сделок
    :param date_end: Дата конца фильтрации сделок
    :param type_deals: Типы сделок для фильтрации
    :param employees: Сотрудники и отделы для фильтрации сделок
    :param stages: Стадии сделок
    :return: Массив найденных сделок по фильтру (состоит из 3 массивов)
    :return:
    """

    if employees == '':     # Если не были выбраны сотрудники в параметрах БП

        # Начались в сентябре 2022 и заканчиваются после сентября 2022

        deals_start_in_end_after = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '>BEGINDATE': date_start,
                    '<BEGINDATE': date_end,
                    '>CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '1',    # Помощник
                    'STAGE_ID': stages
                }
            }
        )

        # начались до сентября 2022 и заканчиваются в сентябре 2022

        deals_start_before_end_in = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '<BEGINDATE': date_start,
                    '>CLOSEDATE': date_start,
                    '<CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '1',    # Помощник
                    'STAGE_ID': stages
                }
            }
        )

        # начались до сентября 2022 и заканчиваются после сентября 2022

        deals_start_before_end_after = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '<BEGINDATE': date_start,
                    '>CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '1',    # Помощник
                    'STAGE_ID': stages
                }
            }
        )
    else:   # Если были выбраны сотрудники в параметрах БП
        id_list = get_employee_id(employees)

        # Начались в сентябре 2022 и заканчиваются после сентября 2022

        deals_start_in_end_after = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '>BEGINDATE': date_start,
                    '<BEGINDATE': date_end,
                    '>CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '1',    # Помощник
                    'ASSIGNED_BY_ID': id_list,
                    'STAGE_ID': stages
                }
            }
        )

        # начались до сентября 2022 и заканчиваются в сентябре 2022

        deals_start_before_end_in = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '<BEGINDATE': date_start,
                    '>CLOSEDATE': date_start,
                    '<CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '1',    # Помощник
                    'ASSIGNED_BY_ID': id_list,
                    'STAGE_ID': stages
                }
            }
        )

        # начались до сентября 2022 и заканчиваются после сентября 2022

        deals_start_before_end_after = b.get_all(
            'crm.deal.list', {
                'filter': {
                    '<BEGINDATE': date_start,
                    '>CLOSEDATE': date_end,
                    'UF_CRM_1662365565770': '1',    # Помощник
                    'ASSIGNED_BY_ID': id_list,
                    'STAGE_ID': stages
                }
            }
        )

    return deals_start_in_end_after + deals_start_before_end_after + deals_start_before_end_in


def create_quarter_subtasks(task_id, check_list_id, employee, quarter_deals, year, month, current_month_days, task_text, dct, companies_name, task_deadline):
    deals = list(filter(lambda x: x['ASSIGNED_BY_ID'] == employee, quarter_deals))
    deals = list(map(lambda x: {'ASSIGNED_BY_ID': x['ASSIGNED_BY_ID'], 'COMPANY_ID': x['COMPANY_ID'], 'ID': x['ID'], 'TITLE': x['TITLE'], 'COMPANY_NAME': list(filter(lambda y: y['ID'] == x['COMPANY_ID'], companies_name))[0]['TITLE']}, deals))
    deals = list(sorted(deals, key=lambda x: x['COMPANY_NAME']))
    for deal in deals:

        company = b.get_all('crm.company.list', {
            'filter': {
                'ID': deal['COMPANY_ID']
            }
        })[0]

        is_quarter_sub_task_exists = b.get_all('tasks.task.list', {
            'select': ['ID'],
            'filter': {'TITLE': f"СВ (К): {company['TITLE']} {dct['month']} {str(year)}",
                       'GROUP_ID': '71'
                       }
        }
                                               )

        if is_quarter_sub_task_exists:
            continue
        '''
        # Создание пунктов чек-листа для созданной задачи на сотрудника
        b.call('task.checklistitem.add', [
            task_id, {
                # <Название компании> <Название сделки> <Ссылка на сделку>
                'TITLE': f"{company['TITLE']} {deal['TITLE']} https://vc4dk.bitrix24.ru/crm/deal/details/{deal['ID']}/",
                'PARENT_ID': check_list_id,
            }
        ], raw=True
               )
        '''
        sub_checklist = send_bitrix_request('task.checklistitem.add', {
            'taskId': check_list_id,
            'FIELDS': {
                'TITLE': f"{company['TITLE']} {deal['TITLE']} https://vc4dk.bitrix24.ru/crm/deal/details/{deal['ID']}/     \n\n "
                         f"Задача: https://vc4dk.bitrix24.ru/workgroups/group/71/tasks/task/view/{task_id}/",
            }
        })

        # Создание подзадачи для основной задачи
        send_bitrix_request('tasks.task.add', {
            'fields': {
                'TITLE': f"СВ (К): {company['TITLE']} {dct['month']} {str(year)}",
                'DEADLINE': task_deadline,
                'RESPONSIBLE_ID': employee,
                'ALLOW_CHANGE_DEADLINE': 'N',
                'GROUP_ID': '71',
                'DESCRIPTION': f"{task_text}\n",
                'PARENT_ID': task_id,
                'UF_CRM_TASK': [f"CO_{company['ID']}", f"D_{deal['ID']}"],
                'CREATED_BY': '173',
            }
        }
               )


def create_service_tasks(dct):
    """
    :param dct: Словарь из url POST запроса, в котором есть ключи 'year', 'month'

    :return: Создает задачи "Сервисный выезд" для сделок уровня ПРОФ для выбранного диапазона дат и с чек-листами,
    где каждый пункт в виде <Название компании> <Название сделки>
    Выборка по датам если выбран сентябрь:

    начались до сентября 2022 и заканчиваются после сентября 2022

    начались в сентябре 2022 и заканчиваются после сентября 2022

    начались до сентября 2022 и заканчиваются в сентябре 2022

    """
    '''
    try:
    '''
    companies_name = b.get_all('crm.company.list', {'select': ['TITLE']})
    task_text = b.get_all('tasks.task.get', {'taskId': dct['task_id']})['task']['description']
    task_text = task_text.replace('[LIST]', '').replace('/[LIST]', '')
    employees = {}  # Dct сотрудников, значения которых - ID сделок для задачи
    type_deals = [
                    'UC_XIYCTV',  # ПРОФ Земля + Помощник
                    'UC_5T4MAW',  # ПРОФ Земля + Облако + Помощник
                    'UC_2SJOEJ',  # ПРОФ Облако+Помощник
                    'UC_81T8ZR',  # АОВ
                    'UC_SV60SP',  # АОВ + Облако
                ]
    stage_ids = [
        'C1:NEW',
        'C1:UC_0KJKTY',
        'C1:UC_3J0IH6',
    ]

    year = dct['year']
    month = months[dct['month']]
    date_start = datetime.strptime(f'01-{month}-{year}', '%d-%m-%Y') - timedelta(days=1)
    date_end = datetime.strftime(date_start + timedelta(days=32), '%Y-%m') + '-01'
    date_start = datetime.strftime(date_start, '%Y-%m-%d')
    if month[0] == '0':
        month.replace('0', '')
    current_month_days = monthrange(int(year), int(month))[1]  # Количество дней в выбранном месяце
    if dct['deadline']:
        deadline_datetime = datetime.strptime(dct['deadline'], '%d.%m.%Y')
        task_deadline = deadline_datetime.strftime('%Y-%m-%d 19:00:00')
    else:
        task_deadline = f"{str(year)}-{month}-{current_month_days} 19:00:00"

    # Получение массива сделок

    deals = get_deals_for_service_tasks(date_start, date_end, type_deals, dct['employees'], stage_ids)
    quarter_deals = []
    if dct['quarter'] in ['Да', 'Только квартальные']:
        quarter_deals = get_quarter_deals_for_service_tasks(date_start, date_end, type_deals, dct['employees'], stage_ids)

    # Разделение ID сделок по ответственному

    for deal in deals:
        employee = deal['ASSIGNED_BY_ID']   # Ответственный
        if employee not in employees:

            # Создание ключа с ID сотрудника и значение:
            # 0: ID сделки
            # 1: Название сделки
            # 2: ID компании
            employees.setdefault(employee, [[deal['ID'], deal['TITLE'], deal['COMPANY_ID'], deal['ID']]])
        else:
            # Добавление ID сделки к значению dct
            employees[employee].append([deal['ID'], deal['TITLE'], deal['COMPANY_ID'], deal['ID']])

    if not deals:
        for deal in quarter_deals:
            employee = deal['ASSIGNED_BY_ID']  # Ответственный
            if employee not in employees:

                # Создание ключа с ID сотрудника и значение:
                # 0: ID сделки
                # 1: Название сделки
                # 2: ID компании
                employees.setdefault(employee, [[deal['ID'], deal['TITLE'], deal['COMPANY_ID'], deal['ID']]])
            else:
                # Добавление ID сделки к значению dct
                employees[employee].append([deal['ID'], deal['TITLE'], deal['COMPANY_ID'], deal['ID']])

    # Формирование задач

    for employee in employees:
        if employee not in ['None', None]:

            employee_fields = b.get_all('user.get', {"ID": employee})
            employee_name = employee_fields[0]['NAME'] + ' ' + employee_fields[0]['LAST_NAME']
            if dct['quarter'] == 'Только квартальные':
                current_month_name = int_to_months[datetime.now().month]
                if dct['deadline']:
                    main_task_name = f"Сервисный выезд (квартал) {employee_name} {current_month_name} {datetime.now().year} - {int_to_months[deadline_datetime.month]} {deadline_datetime.year}"
                else:
                    main_task_name = f"Сервисный выезд (квартал) {employee_name} {current_month_name} {datetime.now().year}"
            else:
                main_task_name = f"Сервисный выезд {employee_name} {dct['month']} {str(year)}"
            is_main_task_exists = b.get_all('tasks.task.list', {
                'select': ['ID'],
                'filter': {'TITLE': main_task_name,
                           'GROUP_ID': '71'
                           }
            }
                                            )
            quarter_check_list_flag = False
            if not is_main_task_exists:
                task = send_bitrix_request('tasks.task.add', {
                    'fields': {
                        'TITLE': main_task_name,
                        'DEADLINE': task_deadline,
                        'RESPONSIBLE_ID': employee,
                        'ALLOW_CHANGE_DEADLINE': 'N',
                        'GROUP_ID': '71',
                        'DESCRIPTION': task_text,
                        'CREATED_BY': '173',
                    }
                }
                              )

                main_task = task['task']['id']
                quarter_check_list = ''
                if dct['quarter'] in ['Да', 'Только квартальные']:
                    quarter_check_list = b.call('task.checklistitem.add', [
                        main_task, {
                            'TITLE': 'Квартальные', 'PARENT_ID': main_task,
                        }
                    ], raw=True
                                                )['result']
                    quarter_check_list_flag = True

                if dct['quarter'] != 'Только квартальные':
                    main_check_list = b.call('task.checklistitem.add', [
                        main_task, {
                            #'TITLE': 'Ежемесячные', 'PARENT_ID': main_task,
                            #ИБС 2026-01-19
                            'TITLE': 'Ежемесячные','PARENT_ID':main_task
                        }
                    ], raw=True
                                                )['result']

            else:
                main_task = is_main_task_exists[0]['id']
                check_lists = b.call('task.checklistitem.getlist', [main_task], raw=True)['result']
                for check_list in check_lists:
                    if check_list['TITLE'] == 'Ежемесячные' or check_list['TITLE'] == 'BX_CHECKLIST_1':
                        main_check_list = check_list['ID']
                    elif check_list['TITLE'] == 'Квартальные':
                        quarter_check_list = check_list['ID']
                        quarter_check_list_flag = True

            if quarter_check_list_flag is False and dct['quarter'] in ['Да', 'Только квартальные']:
                quarter_check_list = b.call('task.checklistitem.add', [
                    main_task, {
                        'TITLE': 'Квартальные', 'PARENT_ID': 0,
                    }
                ], raw=True
                                            )['result']

            if dct['quarter'] in ['Да', 'Только квартальные']:
                if dct['quarter'] == 'Только квартальные':
                    quarter_check_list = main_task
                create_quarter_subtasks(main_task, quarter_check_list, employee, quarter_deals, year, month,
                                           current_month_days, task_text, dct, companies_name, task_deadline)

        # Перебор значений выбранного выше ключа
        employees[employee] = list(map(lambda x: [x[0], x[1], x[2], x[3], list(filter(lambda y: y['ID'] == x[2], companies_name))[0]['TITLE']], employees[employee]))
        employees[employee] = list(sorted(employees[employee], key=lambda x: x[4]))
        for value in employees[employee]:
            sleep(1)
            if employee in [None, 'None'] or not deals:
                continue

            company = b.get_all('crm.company.list', {
                'filter': {
                    'ID': value[2]
                }
            })

            # Проверка была ли создана подзадача, для возможности допостановки
            is_sub_task_exists = b.get_all('tasks.task.list', {
                'select': ['ID'],
                'filter': {'TITLE': f"СВ: {company[0]['TITLE']} {dct['month']} {str(year)}",
                           'GROUP_ID': '71'
                           }
            }
                                       )
            if not is_sub_task_exists and dct['quarter'] != 'Только квартальные':
                '''
                # Создание подзадачи для основной задачи
                sub_task = send_bitrix_request('tasks.task.add', {
                    'fields': {
                        'TITLE': f"СВ: {company[0]['TITLE']} {dct['month']} {str(year)}",
                        'DEADLINE': task_deadline,
                        'RESPONSIBLE_ID': employee,
                        'ALLOW_CHANGE_DEADLINE': 'N',
                        'GROUP_ID': '71',
                        'DESCRIPTION': f"{task_text}\n",
                        'PARENT_ID': main_task,
                        'UF_CRM_TASK': [f"CO_{company[0]['ID']}", f"D_{value[3]}"],
                        'CREATED_BY': '173',
                    }
                }
                              )
                sleep(1)
                '''
                # Создание пунктов чек-листа для созданной задачи на сотрудника
                send_bitrix_request('task.checklistitem.add', {
                    'taskId': main_task,
                    'FIELDS': {
                        'TITLE': f"{company[0]['TITLE']} {value[1]} https://vc4dk.bitrix24.ru/crm/deal/details/{value[0]}/     \n\n "
                                 #f"Задача: https://vc4dk.bitrix24.ru/workgroups/group/71/tasks/task/view/{sub_task['task']['id']}/",
                    }
                }
                       )


        # Защита от дублирования задач

        updated_task = b.get_all('tasks.task.get', {'taskId': main_task})
        if len(updated_task['task']['checklist']) == 0:
            send_bitrix_request('tasks.task.delete', {'taskId': main_task})
            print('Удалена пустая задача')

    b.call('im.notify.system.add', {'USER_ID': dct['user_id'][5:], 'MESSAGE': f'Задачи на сервисный выезд поставлены'})
    '''
    except:
        b.call('im.notify.system.add',
               {'USER_ID': dct['user_id'][5:], 'MESSAGE': f'Не удалось поставить задачи на сервисный выезд'})
    '''


def get_report_comment(task_id):
    report_comments = requests.get(f'{authentication("Bitrix")}task.commentitem.getlist?ID={task_id}').json()['result']
    for report_comment in report_comments:
        if 'Отчет Сервисный выезд' in report_comment['POST_MESSAGE'] and 'вы добавлены наблюдателем' not in report_comment:
            return report_comment['POST_MESSAGE'].replace('[USER=333]Отчет Сервисный выезд[/USER]', '')


def create_service_tasks_report(req):
    try:
        month_last_day = monthrange(int(req['year']), int(months[req['month']]))[1]
        tasks = b.get_all('tasks.task.list', {
            'filter': {
                '>=CREATED_DATE': f"{req['year']}-{months[req['month']]}-01",
                '<=CREATED_DATE': f"{req['year']}-{months[req['month']]}-{month_last_day}",
                'GROUP_ID': '71',
                'RESPONSIBLE_ID': get_employee_id(req['employees']),
                'REAL_STATUS': '5',
            }
        })
        tasks = list(map(lambda x: [
            x['responsible']['name'],
            ' '.join(x['title'].split(' ')[:-2]),
            get_report_comment(x['id']),
            f"https://vc4dk.bitrix24.ru/workgroups/group/71/tasks/task/view/{x['id']}/",
            x['title'],
        ], tasks))

        # Создание xlsx файла отчета
        report_created_time = datetime.now()
        report_name_time = report_created_time.strftime('%d-%m-%Y %H %M %S %f')
        report_name = f'Отчет по ЗСВ {report_name_time}.xlsx'.replace(' ', '_')
        workbook = openpyxl.Workbook()
        worklist = workbook.active

        worklist.append([f"{req['month']} {req['year']}"])
        worklist.append(['Ответственный', 'Компания', 'Комментарий', 'Ссылка на задачу'])
        for task in tasks:
            if 'Сервисный выезд' in task[-1]:
                continue
            worklist.append(task[:-1])
        for idx, col in enumerate(worklist.columns, 1):
            worklist.column_dimensions[get_column_letter(idx)].auto_size = True
        workbook.save(report_name)

        # Загрузка отчета в Битрикс
        bitrix_folder_id = '189467'
        with open(report_name, 'rb') as file:
            report_file = file.read()
        report_file_base64 = str(base64.b64encode(report_file))[2:]
        upload_report = b.call('disk.folder.uploadfile', {
            'id': bitrix_folder_id,
            'data': {'NAME': report_name},
            'fileContent': report_file_base64
        })
        b.call('im.notify.system.add', {
            'USER_ID': req['user_id'][5:],
            'MESSAGE': f'Отчет по ЗСВ за {req["month"]} {req["year"]} сформирован. {upload_report["DETAIL_URL"]}'})
        os_remove(report_name)
    except:
        b.call('im.notify.system.add', {
            'USER_ID': req['user_id'][5:],
            'MESSAGE': f'Не удалось сформировать отчет по задачам на сервисный выезд'})



