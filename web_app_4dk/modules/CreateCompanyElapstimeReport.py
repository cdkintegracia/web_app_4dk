import os
from datetime import timedelta
from datetime import datetime
import base64
from time import strptime

import openpyxl
from openpyxl.utils import get_column_letter
from fast_bitrix24 import Bitrix
from time import sleep

from web_app_4dk.modules.authentication import authentication


webhook = authentication('Bitrix')
b = Bitrix(webhook)

def change_sheet_style(sheet) -> None:

    # Изменение ширины
    for column_cells in sheet.columns:
        length = max(len(str(cell.value) if cell.value else '') for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 1.1

def create_company_elapstime_report(req):

    # Формирование заголовков отчета
    report_created_time = datetime.now()

    company_id = req['company_id']
    company_info = b.get_all('crm.company.get', {
        'ID': company_id,
        'select': ['TITLE'],
    })
    company_name = company_info['TITLE']

    report_data = []

    total_duration = timedelta()

    # Формирование отчета

    users_info = b.get_all('user.get', {
        'filter': {
            'UF_DEPARTMENT': ['5', '27', '29', '231', '458'] # ЦС и ЛК
        }
    })

    users_id = list(map(lambda x: x['ID'], users_info))

    date_task_filter = (datetime.strptime(req['date_start'], "%d.%m.%Y") - timedelta(days=30)).strftime("%Y-%m-%d")

    tasks = b.get_all('tasks.task.list', {
        'filter': {
            '>=CREATED_DATE': date_task_filter,
            'UF_CRM_TASK': ['CO_' + company_id],
        },
        'select': ['*', 'UF_CRM_TASK', 'TAGS']
    })
    tasks_id = [x['id'] for x in tasks]

    date_start = req['date_start']
    date_end = req.get('date_end')

    start_iso = f"{date_start}T00:00:00"

    if date_end:
        end_dt = datetime.strptime(date_end, "%d.%m.%Y") + timedelta(days=1)
        end_iso = end_dt.strftime("%Y-%m-%dT00:00:00")
    else:
        end_iso = datetime.now().isoformat()

    all_times = []
    page = 1

    while True:
        response = b.call(
            'task.elapseditem.getlist',
            {
                "order": {"ID": "asc"},
                "filter": {
                    "TASK_ID": tasks_id,
                    "USER_ID": users_id,
                    ">=CREATED_DATE": start_iso,
                    "<CREATED_DATE": end_iso,
                },
                "select": ["*"],
                "params": {
                    "NAV_PARAMS": {
                        "nPageSize": 50,
                        "iNumPage": page,
                    }
                },
            },
            raw=True
        )

        result = response.get("result", [])
        if not result:
            break # если страницы закончились — прерываем

        all_times.extend(result)

        page += 1 # двигаем страницу

        if page > 500: # защита от возможной бесконечной петли
            print(f"Прерывание: превышено максимальное число страниц затрат ({company_name})")
            break

        sleep(0.3)

    titles = [

        [
            company_name, '',
            f'{req["date_start"]} - {req["date_end"] if req["date_end"] else datetime.now().strftime("%d.%m.%Y")}',
        ],
        [],
        [
            'Время отнесения',
            'Id задачи',
            'Группа',
            'Теги ',
            'Автор трудозатраты',
            'ЧЧ:ММ:СС',
            'Комментарий',
        ]
    ]

    tasks_map = {t['id']: t for t in tasks}

    for eltime in all_times:

        task_id = eltime.get("TASK_ID")
        task_info = tasks_map.get(task_id, {})
        
        created_dt = datetime.fromisoformat(eltime["CREATED_DATE"]).replace(tzinfo=None)
        created_str = created_dt.strftime("%d.%m.%y %H:%M")

        try:
            group_name = task_info['group']['name']
        except:
            group_name = ''

        try:
            tags_str = ', '.join(tag['title'] for tag in task_info.get('tags', {}).values())
        except:
            tags_str = ''

        user_info = list(filter(lambda x: x['ID'] == eltime['USER_ID'], users_info))[0]
        user_name = f'{user_info["NAME"]} {user_info["LAST_NAME"]}'

        seconds = int(eltime.get("SECONDS", 0))
        duration = timedelta(seconds=seconds)
        total_duration += duration

        # формат времени
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        secs = seconds % 60
        duration_str = f"{hours:02}:{minutes:02}:{secs:02}"

        report_data.append([
            created_dt,
            created_str,
            task_id,
            group_name,
            tags_str,
            user_name,
            duration_str,
            eltime.get("COMMENT_TEXT", "")
        ])


    report_data_sorted = sorted(report_data, key=lambda x: x[0])

    # Создание xlsx файла отчета
    report_name_time = report_created_time.strftime('%d-%m-%Y %H %M %S %f')
    report_name = f'Отчет по трудозатратам {company_name} {report_name_time}.xlsx'.replace(' ', '_')
    workbook = openpyxl.Workbook()
    worklist = workbook.active

    for data in titles:
        worklist.append(data)

    for data in report_data_sorted:
        worklist.append([data[1]] + data[2:])
        # Добавляем итоговую строку
    worklist.append([])
    worklist.append(['', '', '', '', 'Итого', str(total_duration), ''])

    for idx, col in enumerate(worklist.columns, 1):
        worklist.column_dimensions[get_column_letter(idx)].auto_size = True

    change_sheet_style(worklist)
    workbook.save(report_name)

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '1947902'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })

    b.call('im.notify.system.add', {
        'USER_ID': req['user_id'][5:],
        'MESSAGE': f'Отчет по трудозатратам сформирован. {upload_report["DETAIL_URL"]}'})
    os.remove(report_name)