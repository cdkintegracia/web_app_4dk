from datetime import datetime, timedelta
from calendar import monthrange
import base64
import os

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from fast_bitrix24 import Bitrix

from web_app_4dk.modules.authentication import authentication
from web_app_4dk.modules.EdoInfoHandler import month_codes, year_codes
from web_app_4dk.modules.field_values import month_int_names


b = Bitrix(authentication('Bitrix'))
deals_info_files_directory = f'/root/web_app_4dk/web_app_4dk/modules/deals_info_files/'
#deals_info_files_directory = f'C:\\Users\\Максим\\Documents\\GitHub\\web_app_4dk\\web_app_4dk\\deals_info_files\\'


def get_employee_id(users: str) -> list:
    """
    Приводит строку с id пользователей и id подразделений к единому списку, состоящему только из id сотрудников

    :param users: Строка из параметра запроса Б24 состоящая из {user_...} и|или {group_...}, которые разделены ', '
    """

    users_id_set = set()

    # Строка с сотрудниками и отделами в список
    users = users.split(', ')

    # Если в массиве найден id сотрудника
    for user_id in users:
        if 'user' in user_id:
            users_id_set.add(user_id[5:])

        # Если в массиве найден id отдела
        elif 'group' in user_id:
            department_users = b.get_all('user.get', {'filter': {'UF_DEPARTMENT': user_id[8:]}})
            for user in department_users:
                users_id_set.add(user['ID'])

    return list(users_id_set)


def get_fio_from_user_info(user_info: dict) -> str:
    """
    Возвращает строку, состоящую из фамилии и имени пользователя Б24

    :param user_info: Словарь, полученный запросом с методом user.get
    """

    '''
    return f'{user_info["LAST_NAME"] if "LAST_NAME" in user_info else ""}'\
           f' {user_info["NAME"] if "NAME" in user_info else ""}'.strip()
    '''

    return (f'{user_info["NAME"] if "NAME" in user_info else ""}'
            f' {user_info["LAST_NAME"] if "LAST_NAME" in user_info else ""}').strip()


def change_sheet_style(sheet) -> None:

    # Изменение ширины
    for column_cells in sheet.columns:
        length = max(len(str(cell.value) if cell.value else '') for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 1.1


def get_quarter_filter(month_number):
    quarters = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9],
        [10, 11, 12]
    ]
    quarter = list(filter(lambda x: month_number in x, quarters))[0]
    if month_number + 1 in quarter:
        quarter.remove(month_number + 1)
    #добавлено 01 01 2024
    if month_number == 12:
        quarter_start_filter = datetime(day=1, month=quarter[0], year=datetime.now().year-1)
    else:
        quarter_start_filter = datetime(day=1, month=quarter[0], year=datetime.now().year)

    month_end = quarter[-1] + 1
    year_end=datetime.now().year
    if month_end == 13:
        month_end = 1
        #ibs 2025-01-11
        #year_end = year_end + 1
    quarter_end_filter = datetime(day=1, month=month_end, year=year_end)

    return {
        'start_date': quarter_start_filter,
        'end_date': quarter_end_filter
    }


def read_deals_data_file(month, year):
    filename = f'{month_int_names[month]}_{year}.xlsx'
    workbook = openpyxl.load_workbook(f'{deals_info_files_directory}{filename}')
    worksheet = workbook.active
    file_data = []
    for index, row in enumerate(worksheet.rows):
        if index == 0:
            file_titles = list(map(lambda x: x.value, row))
        else:
            row_data = list(map(lambda x: x.value, row))
            row_data = dict(zip(file_titles, row_data))
            if not row_data['Тип'] or row_data['Тип'] == 'Тестовый':
                continue
            file_data.append(row_data)
    return file_data


def create_employees_report(req):
    users_id = get_employee_id(req['users'])
    users_info = b.get_all('user.get', {
        'filter': {
            'ACTIVE': 'true',
            'ID': users_id,
        }
    })

    deal_fields = b.get_all('crm.deal.fields')

    report_year = datetime.now().year
    report_month = datetime.now().month - 1

    if report_month == 0:
        report_month = 12
        report_year -= 1
    report_month_range = monthrange(report_year, report_month)[1]
    report_month_last_day_date = (f'{report_month_range}.'
                                  f'{report_month if len(str(report_month)) == 2 else "0" + str(report_month)}.'
                                  f'{report_year}')
    before_last_month = report_month - 1
    before_last_month_year = report_year
    if before_last_month == 0:
        before_last_month = 12
        before_last_month_year -= 1
    before_last_month_range_days = monthrange(before_last_month_year, before_last_month)[1]

    before_before_last_month = before_last_month - 1
    before_before_last_month_year = before_last_month_year
    if before_before_last_month == 0:
        before_before_last_month = 12
        before_before_last_month_year -= 1

    before_before_before_last_month = before_last_month - 1
    before_before_before_last_month_year = before_last_month_year
    if before_before_before_last_month == 0:
        before_before_before_last_month = 12
        before_before_before_last_month_year -= 1

    month_filter_start = datetime(day=1, month=report_month, year=report_year)
    month_filter_end = datetime(day=1, month=datetime.now().month, year=datetime.now().year)
    ddmmyyyy_pattern = '%d.%m.%Y'
    if datetime.now().month == 1:
        quarter_filters = get_quarter_filter(12)
    else:
        quarter_filters = get_quarter_filter(datetime.now().month - 1)

    deal_group_field = deal_fields['UF_CRM_1657878818384']['items']
    deal_group_field.append({'ID': None, 'VALUE': 'Лицензии'})
    deal_group_field.append({'ID': None, 'VALUE': 'Остальные'})

    workbook = openpyxl.Workbook()
    report_name = f'Отчет_по_сотрудникам_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
    month_names = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август', 9: 'Сентябрь',
        10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }

    for index, user_info in enumerate(users_info):
        user_name = get_fio_from_user_info(user_info)
        if index == 0:
            worksheet = workbook.active
            worksheet.title = user_name
        else:
            worksheet = workbook.create_sheet(user_name)

        worksheet.append([user_name, '', f'{month_names[report_month]} {report_year}'])
        worksheet.append([])
        worksheet.append([])
    

        # Закрытые задачи
        tasks = b.get_all('tasks.task.list', {
            'filter': {
                'RESPONSIBLE_ID': user_info['ID'],
                '>=CREATED_DATE': month_filter_start.strftime(ddmmyyyy_pattern),
                '<CREATED_DATE': month_filter_end.strftime(ddmmyyyy_pattern),
            },
            'select': ['ID', 'GROUP_ID', 'STATUS', 'UF_AUTO_177856763915']
        })


        tlp_tasks = list(filter(lambda x: x['groupId'] == '1' and x['status'] == '5', tasks))
        tlp_ids = list(set(map(lambda x: x['id'], tlp_tasks)))

        worksits_tasks = list(filter(lambda x: x['groupId'] == '321' and x['status'] == '5', tasks))
        worksits_ids = list(set(map(lambda x: x['id'], worksits_tasks)))

        pu_tasks = list(filter(lambda x: x['groupId'] == '408' and x['status'] == '5', tasks))
        pu_ids = list(set(map(lambda x: x['id'], pu_tasks)))

        others_tasks = list(filter(lambda x: x['groupId'] not in ['1', '321', '408'] and x['status'] == '5', tasks))
        others_ids = list(set(map(lambda x: x['id'], others_tasks)))

        all_tasks = list(filter(lambda x: x['status'] == '5', tasks))
        all_ids = list(set(map(lambda x: x['id'], all_tasks)))
    
        if all_ids:
            all_timespent = []
            page = 1
            page_size = 50

            while True:
                response = b.call(
                    'task.elapseditem.getlist',
                    {
                        'order': {'ID': 'asc'},
                        'filter': {
                            'USER_ID': user_info['ID'],
                            'TASK_ID': all_ids, 
                            #'>=CREATED_DATE': month_filter_start.strftime(ddmmyyyy_pattern),
                            #'<CREATED_DATE': month_filter_end.strftime(ddmmyyyy_pattern),
                        },
                        'select': ["*"],
                        'params': {
                            "NAV_PARAMS": {
                                "nPageSize": page_size,
                                "iNumPage": page,
                            }
                        },
                    },
                    raw=True
                )

                result = response.get("result", [])
                if not result:
                    break # если страницы закончились — прерываем

                all_timespent.extend(result)

                page += 1 # двигаем страницу

                if page > 100: # защита от возможной бесконечной петли
                    print(f"Прерывание: превышено максимальное число страниц (user_id={user_info['ID']})")
                    break


        # фильтрация трудозатрат по группам
        tlp_timespent = list(filter(lambda x: x['TASK_ID'] in tlp_ids, all_timespent))
        worksits_timespent = list(filter(lambda x: x['TASK_ID'] in worksits_ids, all_timespent))
        pu_timespent = list(filter(lambda x: x['TASK_ID'] in pu_ids, all_timespent))
        others_timespent = list(filter(lambda x: x['TASK_ID'] in others_ids, all_timespent))

        def calc_timespent(timespent): # функция для подсчета времени
            total_seconds = sum(int(item.get('SECONDS', 0)) for item in timespent)

            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60

            return hours, minutes

        # итог время
        tlp_hours, tlp_minutes = calc_timespent(tlp_timespent)
        worksits_hours, worksits_minutes = calc_timespent(worksits_timespent)
        pu_hours, pu_minutes = calc_timespent(pu_timespent)
        others_hours, others_minutes = calc_timespent(others_timespent)
        all_hours, all_minutes = calc_timespent(all_timespent)

        # средняя оценка
        tasks_ratings = list(map(lambda x: int(x['ufAuto177856763915']) if x['ufAuto177856763915'] else 0, tlp_tasks))
        tasks_ratings = list(filter(lambda x: x != 0, tasks_ratings))
        try:
            average_tasks_ratings = round(sum(tasks_ratings) / len(tasks_ratings), 2)
        except ZeroDivisionError:
            average_tasks_ratings = '-'

        # попытка Дежурства
        days_duty = b.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '301',
            'filter': {
                'PROPERTY_1753': user_info['ID'],
                'PROPERTY_1769': int(report_month),
                'PROPERTY_1771': int(report_year),
            }
        })
        days_duty_amount = len(days_duty)


        worksheet.append(['Закрытые задачи', 'ТЛП', 'РаботыИТС', 'Платные работы', 'Остальные', 'Всего'])
        worksheet.append(['Кол-во', len(tlp_tasks), len(worksits_tasks), len(pu_tasks), len(others_tasks), len(all_tasks)])
        worksheet.append(['Трудозатраты', f'{tlp_hours} ч {tlp_minutes} мин', f'{worksits_hours} ч {worksits_minutes} мин', 
                          f'{pu_hours} ч {pu_minutes} мин', f'{others_hours} ч {others_minutes} мин', f'{all_hours} ч {all_minutes} мин'])
        row = worksheet.max_row
        for col in range(2, 7):
            worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='right') # выравнивание по правому краю
        worksheet.append(['Средняя оценка', average_tasks_ratings])
        worksheet.append(['Дней дежурства', days_duty_amount])
        worksheet.append([])


        # Звонки и письма

        # сбор инфо по входящим более 10 секунд за день
        incalls = b.get_all('voximplant.statistic.get', {
            'filter': {
                'CALL_TYPE': 2,
                'PORTAL_USER_ID': user_info['ID'],
                '>CALL_DURATION': 10,
                '>=CALL_START_DATE': month_filter_start.strftime(ddmmyyyy_pattern),
                '<CALL_START_DATE': month_filter_end.strftime(ddmmyyyy_pattern),
                'CALL_FAILED_CODE': '200',
            }})

        # сбор инфо по исходящим более 10 секунд за день
        outcalls = b.get_all('voximplant.statistic.get', {
            'filter': {
                'CALL_TYPE': 1,
                'PORTAL_USER_ID': user_info['ID'],
                '>CALL_DURATION': 10,
                '>=CALL_START_DATE': month_filter_start.strftime(ddmmyyyy_pattern),
                '<CALL_START_DATE': month_filter_end.strftime(ddmmyyyy_pattern),
                'CALL_FAILED_CODE': '200',
            }})
        
        #duration_in = sum(list(map(lambda x: int(x['CALL_DURATION']), incalls)))
        #duration_out = sum(list(map(lambda x: int(x['CALL_DURATION']), outcalls)))
        #duration = duration_in + duration_out
        total_seconds = (sum(int(x['CALL_DURATION']) for x in incalls) + sum(int(x['CALL_DURATION']) for x in outcalls))
        duration_hours = total_seconds // 3600
        duration_minutes = (total_seconds % 3600) // 60

        #сбор инфо по исходящим письмам с привязкой к сущностям
        emails = b.get_all('crm.activity.list', {
            'filter': {
                'PROVIDER_TYPE_ID': 'EMAIL_COMPRESSED',
                'AUTHOR_ID': user_info['ID'],
                '>=CREATED': month_filter_start.strftime(ddmmyyyy_pattern),
                '<CREATED': month_filter_end.strftime(ddmmyyyy_pattern),
                'DIRECTION': '2'
                }})

        worksheet.append([])
        worksheet.append(['Входящих звонков (> 10 сек)', len(incalls)])
        worksheet.append(['Исходящих звонков (> 10 сек)', len(outcalls)])
        worksheet.append(['Общее время звонков', f'{duration_hours} ч {duration_minutes} мин'])
        worksheet.append(['Отправлено писем', len(emails)])

        '''
        # ЭДО
        all_its = list(filter(lambda x: 'ГРМ' not in x['Тип'], its_deals_last_month))
        count_its_unigue = len(set(x['Компания'] for x in all_its))

        edo_companies_id = list(map(lambda x: x['Компания'], list(filter(lambda y: 'Компания' in y and y['Компания'], all_its))))

        if edo_companies_id:
            edo_companies = b.get_all('crm.company.list', {
                'select': ['UF_CRM_1638093692254', 'UF_CRM_1638093750742'],
                'filter': {
                    'ID': list(map(lambda x: x['Компания'], list(filter(lambda y: 'Компания' in y and y['Компания'], all_its))))
                }
            })
            edo_companies_count = list(filter(lambda x: x['UF_CRM_1638093692254'] == '69', edo_companies))

            #спецоператоры эдо
            try: 
                operator_2ae = list(filter(lambda x: x['UF_CRM_1638093750742'] == '75', edo_companies_count))
            except ZeroDivisionError:
                operator_2ae = 0
            try: 
                operator_2ae_doki = list(filter(lambda x: x['UF_CRM_1638093750742'] == '1715', edo_companies_count))
            except ZeroDivisionError:
                operator_2ae_doki = 0
            try: 
                operator_2be = list(filter(lambda x: x['UF_CRM_1638093750742'] == '77', edo_companies_count))
            except ZeroDivisionError:
                operator_2be = 0
            try: 
                operator_2bm = list(filter(lambda x: x['UF_CRM_1638093750742'] == '73', edo_companies_count))
            except ZeroDivisionError:
                operator_2bm = 0
            try: 
                operator_2al = list(filter(lambda x: x['UF_CRM_1638093750742'] == '437', edo_companies_count))
            except ZeroDivisionError:
                operator_2al = 0
            try: 
                operator_2lb = list(filter(lambda x: x['UF_CRM_1638093750742'] == '439', edo_companies_count))
            except ZeroDivisionError:
                operator_2lb = 0
            try: 
                operator_2bk = list(filter(lambda x: x['UF_CRM_1638093750742'] == '1357', edo_companies_count))
            except ZeroDivisionError:
                operator_2bk = 0
            try: 
                operator_2lt = list(filter(lambda x: x['UF_CRM_1638093750742'] == '1831', edo_companies_count))
            except ZeroDivisionError:
                operator_2lt = 0

            edo_elements_info = b.get_all('lists.element.get', {
                'IBLOCK_TYPE_ID': 'lists',
                'IBLOCK_ID': '235',
                'filter': {
                    'PROPERTY_1579': edo_companies_id,
                    'PROPERTY_1567': month_codes[month_int_names[report_month]],
                    'PROPERTY_1569': year_codes[str(report_year)],
                }
            })
            edo_elements_info = list(map(lambda x: {
                'ID': x['ID'],
                'Компания': list(x['PROPERTY_1579'].values())[0],
                'Сумма пакетов по владельцу': int(list(x['PROPERTY_1573'].values())[0]),
                'Сумма для клиента': int(list(x['PROPERTY_1575'].values())[0]),
            }, edo_elements_info))
            traffic_more_than_1 = list(filter(lambda x: x['Сумма пакетов по владельцу'] > 1, edo_elements_info))
            edo_elements_paid = b.get_all('lists.element.get', {
                'IBLOCK_TYPE_ID': 'lists',
                'IBLOCK_ID': '235',
                'filter': {
                    'PROPERTY_1581': user_info['ID'],
                    'PROPERTY_1567': month_codes[month_int_names[report_month]],
                    'PROPERTY_1569': year_codes[str(report_year)],
                }
            })

            paid_traffic = list(filter(lambda x: int(list(x['PROPERTY_1573'].values())[0]) > 0 and int(list(x['PROPERTY_1575'].values())[0]) > 0, edo_elements_paid))
            paid_traffic = sum(list(map(lambda x: int(list(x['PROPERTY_1575'].values())[0]), paid_traffic)))

        else:
            edo_companies_count = []
            traffic_more_than_1 = []
            paid_traffic = 0
            operator_2ae = []
            operator_2ae_doki = []
            operator_2be = []
            operator_2bm = []
            operator_2al = []
            operator_2lb = []
            operator_2bk = []
            operator_2lt = []

        try:
            edo_companies_coverage = round((len(edo_companies_count) / len(all_its)) * 100, 2)
        except ZeroDivisionError:
            edo_companies_coverage = 0

        try:
            active_its_coverage = round((len(traffic_more_than_1) / len(all_its)) * 100, 2)
        except ZeroDivisionError:
            active_its_coverage = 0

        try:
            operator_2ae_coverage = round((len(operator_2ae) / len(edo_companies_count)) * 100, 2)
        except ZeroDivisionError:
            operator_2ae_coverage = 0
        try:
            operator_2ae_doki_coverage = round((len(operator_2ae_doki) / len(edo_companies_count)) * 100, 2)
        except ZeroDivisionError:
            operator_2ae_doki_coverage = 0
        try:
            operator_2be_coverage = round((len(operator_2be) / len(edo_companies_count)) * 100, 2)
        except ZeroDivisionError:
            operator_2be_coverage = 0
        try:
            operator_2bm_coverage = round((len(operator_2bm) / len(edo_companies_count)) * 100, 2)
        except ZeroDivisionError:
            operator_2bm_coverage = 0
        try:
            operator_2al_coverage = round((len(operator_2al) / len(edo_companies_count)) * 100, 2)
        except ZeroDivisionError:
            operator_2al_coverage = 0
        try:
            operator_2lb_coverage = round((len(operator_2lb) / len(edo_companies_count)) * 100, 2)
        except ZeroDivisionError:
            operator_2lb_coverage = 0
        try:
            operator_2bk_coverage = round((len(operator_2bk) / len(edo_companies_count)) * 100, 2)
        except ZeroDivisionError:
            operator_2bk_coverage = 0
        try:
            operator_2lt_coverage = round((len(operator_2lt) / len(edo_companies_count)) * 100, 2)
        except ZeroDivisionError:
            operator_2lt_coverage = 0

        worksheet.append(['ЭДО', 'Всего ИТС', 'С ЭДО', '%'])
        worksheet.append(['Охват ЭДО', count_its_unigue, len(edo_companies_count), edo_companies_coverage])
        if len(edo_companies_count) > 0:
            if len(operator_2ae) > 0:
                worksheet.append(['', '2AE', len(operator_2ae), operator_2ae_coverage])
            if len(operator_2ae_doki) > 0:
                worksheet.append(['', '2AE доки', len(operator_2ae_doki), operator_2ae_doki_coverage])
            if len(operator_2be) > 0:
                worksheet.append(['', '2BE', len(operator_2be), operator_2be_coverage])
            if len(operator_2bm) > 0:
                worksheet.append(['', '2BM', len(operator_2bm), operator_2bm_coverage])
            if len(operator_2al) > 0:
                worksheet.append(['', '2AL', len(operator_2al), operator_2al_coverage])
            if len(operator_2lb) > 0:
                worksheet.append(['', '2LB', len(operator_2lb), operator_2lb_coverage])
            if len(operator_2bk) > 0:
                worksheet.append(['', '2BK', len(operator_2bk), operator_2bk_coverage])
            if len(operator_2lt) > 0:
                worksheet.append(['', '2LT', len(operator_2lt), operator_2lt_coverage])
        worksheet.append(['Компании с трафиком больше 1', len(set(map(lambda x: x['Компания'], traffic_more_than_1)))])
        worksheet.append(['% активных ИТС', active_its_coverage])
        worksheet.append(['Сумма платного трафика', paid_traffic])
        worksheet.append([])

        #Сверка 2.0
        all_company = b.get_all('crm.company.list', {
                'select': ['UF_CRM_1735194029'],
                'filter': {
                    'ASSIGNED_BY_ID': user_info['ID'],
                }
            })
        company_sverka = list(filter(lambda x: x['UF_CRM_1735194029'] == '1', all_company))
        
        worksheet.append(['Сверка 2.0'])
        worksheet.append(['Подключено', len(company_sverka)])
        '''

        change_sheet_style(worksheet)
        

    workbook.save(report_name)

    if 'user_id' not in req:
        return

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '600147'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    b.call('im.notify.system.add', {
        'USER_ID': req['user_id'][5:],
        'MESSAGE': f'Отчет по пользователям сформирован. {upload_report["DETAIL_URL"]}'})
    os.remove(report_name)


if __name__ == '__main__':
    create_employees_report({
        'users': 'user_181'  #'user_129'   #'group_dr27', 'user_135'
    })
