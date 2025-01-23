from datetime import datetime, timedelta
from calendar import monthrange
import base64
import os

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from fast_bitrix24 import Bitrix

from web_app_4dk.modules.authentication import authentication
from web_app_4dk.modules.EdoInfoHandler import month_codes, year_codes
from web_app_4dk.modules.field_values import month_int_names


b = Bitrix(authentication('Bitrix'))
deals_info_files_directory = f'/root/web_app_4dk/web_app_4dk/modules/deals_info_files/'


def get_employee_id(users: str) -> list:
    """
    Приводит строку с id пользователей и id подразделений к единому списку, состоящему только из id сотрудников

    :param users: Строка из параметра запроса Б24 состоящая из {user_...} и|или {group_...}, которые разделены ', '
    """

    users_id_set = set()

    # Строка с сотрудниками и отделами в список
    users = users.split(', ')

    # Если в массиве найден id сотрудника
    for user_id in users:
        if 'user' in user_id:
            users_id_set.add(user_id[5:])

        # Если в массиве найден id отдела
        elif 'group' in user_id:
            department_users = b.get_all('user.get', {'filter': {'UF_DEPARTMENT': user_id[8:]}})
            for user in department_users:
                users_id_set.add(user['ID'])

    return list(users_id_set)


def get_fio_from_user_info(user_info: dict) -> str:
    """
    Возвращает строку, состоящую из фамилии и имени пользователя Б24

    :param user_info: Словарь, полученный запросом с методом user.get
    """

    '''
    return f'{user_info["LAST_NAME"] if "LAST_NAME" in user_info else ""}'\
           f' {user_info["NAME"] if "NAME" in user_info else ""}'.strip()
    '''

    return (f'{user_info["NAME"] if "NAME" in user_info else ""}'
            f' {user_info["LAST_NAME"] if "LAST_NAME" in user_info else ""}').strip()


def change_sheet_style(sheet) -> None:

    # Изменение ширины
    for column_cells in sheet.columns:
        length = max(len(str(cell.value) if cell.value else '') for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 1.1


def get_quarter_filter(month_number):
    quarters = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9],
        [10, 11, 12]
    ]
    quarter = list(filter(lambda x: month_number in x, quarters))[0]
    if month_number + 1 in quarter:
        quarter.remove(month_number + 1)  
    #добавлено 01 01 2024
    if month_number == 12:
        quarter_start_filter = datetime(day=1, month=quarter[0], year=datetime.now().year-1)
    else:
        quarter_start_filter = datetime(day=1, month=quarter[0], year=datetime.now().year)
    #quarter_start_filter = datetime(day=1, month=quarter[0], year=datetime.now().year)
    month_end = quarter[-1] + 1
    year_end=datetime.now().year
    if month_end == 13:
        month_end = 1

    quarter_end_filter = datetime(day=1, month=month_end, year=year_end)
    #print(quarter_end_filter)

    return {
        'start_date': quarter_start_filter,
        'end_date': quarter_end_filter
    }


def read_deals_data_file(month, year):
    filename = f'{month_int_names[month]}_{year}.xlsx'
    workbook = openpyxl.load_workbook(f'{deals_info_files_directory}{filename}')
    worksheet = workbook.active
    file_data = []
    for index, row in enumerate(worksheet.rows):
        if index == 0:
            file_titles = list(map(lambda x: x.value, row))
        else:
            row_data = list(map(lambda x: x.value, row))
            row_data = dict(zip(file_titles, row_data))
            if not row_data['Тип'] or row_data['Тип'] == 'Тестовый':
                continue
            file_data.append(row_data)
    return file_data


def create_employees_quarter_report(req):
    users_id = get_employee_id(req['users'])
    users_info = b.get_all('user.get', {
        'filter': {
            'ACTIVE': 'true',
            'ID': users_id,
        }
    })

    deal_fields = b.get_all('crm.deal.fields')

    before_1_month_year = datetime.now().year
    before_1_month = datetime.now().month - 1
    #before_1_month = 3

    if before_1_month == 0:
        before_1_month = 12
        before_1_month_year -= 1
    before_1_month_range = monthrange(before_1_month_year, before_1_month)[1]
    before_1_month_last_day_date = (f'{before_1_month_range}.'
                                  f'{before_1_month if len(str(before_1_month)) == 2 else "0" + str(before_1_month)}.'
                                  f'{before_1_month_year}')
      
    before_2_month = before_1_month - 1
    before_2_month_year = before_1_month_year
    if before_2_month == 0:
        before_2_month = 12
        before_2_month_year -= 1
    before_2_month_range = monthrange(before_2_month_year, before_2_month)[1]

    before_3_month = before_2_month - 1
    before_3_month_year = before_2_month_year
    if before_3_month == 0:
        before_3_month = 12
        before_3_month_year -= 1
    before_3_month_range = monthrange(before_3_month_year, before_3_month)[1]

    before_4_month = before_3_month - 1
    before_4_month_year = before_3_month_year
    if before_4_month == 0:
        before_4_month = 12
        before_4_month_year -= 1
    before_4_month_range = monthrange(before_4_month_year, before_4_month)[1]

    before_5_month = before_4_month - 1
    before_5_month_year = before_4_month_year
    if before_5_month == 0:
        before_5_month = 12
        before_5_month_year -= 1
 
    ddmmyyyy_pattern = '%d.%m.%Y'

    deal_group_field = deal_fields['UF_CRM_1657878818384']['items']
    deal_group_field.append({'ID': None, 'VALUE': 'Лицензии'})
    deal_group_field.append({'ID': None, 'VALUE': 'Остальные'})

    workbook = openpyxl.Workbook()
    report_name = f'Квартальный_отчет_по_сотрудникам_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
    month_names = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август', 9: 'Сентябрь',
        10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }

    for index, user_info in enumerate(users_info):
        user_name = get_fio_from_user_info(user_info)
        if index == 0:
            worksheet = workbook.active
            worksheet.title = user_name
        else:
            worksheet = workbook.create_sheet(user_name)

        quarter_filters = get_quarter_filter(before_1_month)
        start_date_quarter = quarter_filters['start_date'] - timedelta(days=1)
        end_date_quarter = quarter_filters['end_date'] - timedelta(days=1)

        quarter_deals_data = read_deals_data_file(start_date_quarter.month, start_date_quarter.year)
        start_year_deals_data = read_deals_data_file(12, datetime.now().year-1)

        if quarter_filters['start_date'].month == 1: number_quarter = 1
        elif quarter_filters['start_date'].month == 4: number_quarter = 2
        elif quarter_filters['start_date'].month == 7: number_quarter = 3
        else: number_quarter = 4

        before_1_month_deals_data = read_deals_data_file(before_1_month, before_1_month_year)
        before_2_month_deals_data = read_deals_data_file(before_2_month, before_2_month_year)
        before_3_month_deals_data = read_deals_data_file(before_3_month, before_3_month_year)
        before_4_month_deals_data = read_deals_data_file(before_4_month, before_4_month_year)
        before_5_month_deals_data = read_deals_data_file(before_5_month, before_5_month_year)

        worksheet.append([user_name, '', f'{number_quarter} квартал {end_date_quarter.year} г.'])
        worksheet.append([])
        worksheet.append([])

        its_deals_before_1_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                                x['Группа'] == 'ИТС' and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                before_1_month_deals_data))

        start_quarter_its_deals = list(filter(lambda x: x['Ответственный'] == user_name and
                                                x['Группа'] == 'ИТС' and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                quarter_deals_data))

        # Сделки
        # Последний месяц квартала
        its_prof_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                                x['Группа'] == 'ИТС' and
                                                'Базовый' not in x['Тип'] and 'ГРМ' not in x['Тип'] and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                before_1_month_deals_data))

        its_base_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                                x['Группа'] == 'ИТС' and
                                                'Базовый' in x['Тип'] and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                before_1_month_deals_data))

        countragent_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                   'Контрагент' in x['Тип'] and
                                                   x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                   before_1_month_deals_data))

        spark_in_contract_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                         '1Спарк в договоре' == x['Тип'] and
                                                         x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                         before_1_month_deals_data))

        spark_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                  ('1Спарк 3000' == x['Тип'] or '1Спарк' == x['Тип']) and
                                                  x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                  before_1_month_deals_data))

        spark_plus_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                   ('22500' in x['Тип'] or '1СпаркПЛЮС' in x['Тип']) and
                                                   x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                   before_1_month_deals_data))

        rpd_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'РПД' in x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           before_1_month_deals_data))
        
        #добавлено 15-07-2024
        grm_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ГРМ' in x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           before_1_month_deals_data))
        
        doki_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                            'Доки' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            before_1_month_deals_data))
        
        report_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                              'Отчетность' in x['Тип'] and 
                                              x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                              before_1_month_deals_data))
        
        signature_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                                 'Подпись' in x['Тип'] and
                                                 x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                 before_1_month_deals_data))
        
        dop_oblako_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                                  'Допы Облако' == x['Тип'] and
                                                  x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                  before_1_month_deals_data))
        
        link_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                            'Линк' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            before_1_month_deals_data))
        
        unics_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                             'Уникс' == x['Тип'] and
                                             x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                             before_1_month_deals_data))
        
        cab_sotrudnik_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                     'Кабинет сотрудника' == x['Тип'] and
                                                     x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                     before_1_month_deals_data))
        
        cab_sadovod_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                   'Кабинет садовода' == x['Тип'] and
                                                   x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                   before_1_month_deals_data))
        
        edo_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ЭДО' == x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           before_1_month_deals_data))
        
        mdlp_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                            'МДЛП' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            before_1_month_deals_data))
        
        connect_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                               'Коннект' == x['Тип'] and
                                               x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                               before_1_month_deals_data))
                
        its_otrasl_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                  'ИТС Отраслевой' == x['Тип'] and
                                                  x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                  before_1_month_deals_data))
        
        ofd_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ОФД' == x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           before_1_month_deals_data))
        
        bitrix24_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                ('Битрикс24' == x['Тип'] or 'БИТРИКС24' == x['Тип']) and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                before_1_month_deals_data))

        other_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                             x not in its_prof_deals_last_month and x not in its_base_deals_last_month and
                                             x not in countragent_deals_last_month and x not in spark_in_contract_deals_last_month and
                                             x not in spark_deals_last_month and x not in spark_plus_deals_last_month and
                                             x not in rpd_deals_last_month and x not in grm_deals_last_month and
                                             x not in doki_deals_last_month and
                                             x not in report_deals_last_month and x not in signature_deals_last_month and
                                             x not in dop_oblako_deals_last_month and x not in link_deals_last_month and
                                             x not in unics_deals_last_month and x not in cab_sotrudnik_deals_last_month and
                                             x not in cab_sadovod_deals_last_month and x not in edo_deals_last_month and
                                             x not in mdlp_deals_last_month and x not in connect_deals_last_month and
                                             x not in its_otrasl_deals_last_month and x not in ofd_deals_last_month and
                                             x not in bitrix24_deals_last_month and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'], before_1_month_deals_data))        
        
        # Начало квартала
        its_prof_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Группа'] == 'ИТС' and
                                             'Базовый' not in x['Тип'] and 'ГРМ' not in x['Тип'] and
                                             x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                             quarter_deals_data))

        its_base_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Группа'] == 'ИТС' and
                                             'Базовый' in x['Тип'] and
                                             x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                             quarter_deals_data))

        countragent_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                'Контрагент' in x['Тип'] and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                quarter_deals_data))

        spark_in_contract_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                      '1Спарк в договоре' == x['Тип'] and
                                                      x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                      quarter_deals_data))

        spark_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                               ('1Спарк 3000' == x['Тип'] or '1Спарк' == x['Тип']) and
                                               x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                               quarter_deals_data))

        spark_plus_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                ('22500' in x['Тип'] or '1СпаркПЛЮС' in x['Тип']) and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                quarter_deals_data))

        rpd_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                        'РПД' in x['Тип'] and
                                        x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                        quarter_deals_data))
        
        #добавлено 15-07-2024
        grm_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ГРМ' in x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           quarter_deals_data))
        
        doki_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                            'Доки' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            quarter_deals_data))
        
        report_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                              'Отчетность' in x['Тип'] and 
                                              x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                              quarter_deals_data))
        
        signature_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                                 'Подпись' in x['Тип'] and
                                                 x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                 quarter_deals_data))
        
        dop_oblako_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                                  'Допы Облако' == x['Тип'] and
                                                  x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                  quarter_deals_data))
        
        link_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                            'Линк' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            quarter_deals_data))
        
        unics_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                             'Уникс' == x['Тип'] and
                                             x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                             quarter_deals_data))
        
        cab_sotrudnik_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                     'Кабинет сотрудника' == x['Тип'] and
                                                     x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                     quarter_deals_data))
        
        cab_sadovod_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                   'Кабинет садовода' == x['Тип'] and
                                                   x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                   quarter_deals_data))
        
        edo_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ЭДО' == x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           quarter_deals_data))
        
        mdlp_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                            'МДЛП' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            quarter_deals_data))
        
        connect_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                               'Коннект' == x['Тип'] and
                                               x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                               quarter_deals_data))
                
        its_otrasl_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                  'ИТС Отраслевой' == x['Тип'] and
                                                  x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                  quarter_deals_data))
        
        ofd_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ОФД' == x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           quarter_deals_data))
        
        bitrix24_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                ('Битрикс24' == x['Тип'] or 'БИТРИКС24' == x['Тип']) and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                quarter_deals_data))

        other_deals_quarter = list(filter(lambda x: x['Ответственный'] == user_name and
                                             x not in its_prof_deals_quarter and x not in its_base_deals_quarter and
                                             x not in countragent_deals_quarter and x not in spark_in_contract_deals_quarter and
                                             x not in spark_deals_quarter and x not in spark_plus_deals_quarter and
                                             x not in rpd_deals_quarter and x not in grm_deals_quarter and
                                             x not in doki_deals_quarter and
                                             x not in report_deals_quarter and x not in signature_deals_quarter and
                                             x not in dop_oblako_deals_quarter and x not in link_deals_quarter and
                                             x not in unics_deals_quarter and x not in cab_sotrudnik_deals_quarter and
                                             x not in cab_sadovod_deals_quarter and x not in edo_deals_quarter and
                                             x not in mdlp_deals_quarter and x not in connect_deals_quarter and
                                             x not in its_otrasl_deals_quarter and x not in ofd_deals_quarter and
                                             x not in bitrix24_deals_quarter and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'], quarter_deals_data))
        
        worksheet.append(['Сделки', f'на {before_1_month_last_day_date}', f'на {start_date_quarter.strftime("%d.%m.%Y")}', 'Прирост с начала квартала'])
        worksheet.append([
            'ИТС ПРОФ',
            len(its_prof_deals_last_month),
            len(its_prof_deals_quarter),
            len(its_prof_deals_last_month) - len(its_prof_deals_quarter)
        ])
        worksheet.append([
            'ИТС Базовые',
            len(its_base_deals_last_month),
            len(its_base_deals_quarter),
            len(its_base_deals_last_month) - len(its_base_deals_quarter)
        ])
        worksheet.append([
            'Контрагент',
            len(countragent_deals_last_month),
            len(countragent_deals_quarter),
            len(countragent_deals_last_month) - len(countragent_deals_quarter)
        ])
        worksheet.append([
            'Спарк в договоре',
            len(spark_in_contract_deals_last_month),
            len(spark_in_contract_deals_quarter),
            len(spark_in_contract_deals_last_month) - len(spark_in_contract_deals_quarter)
        ])
        worksheet.append([
            'Спарк',
            len(spark_deals_last_month),
            len(spark_deals_quarter),
            len(spark_deals_last_month) - len(spark_deals_quarter)
        ])
        worksheet.append([
            'СпаркПлюс',
            len(spark_plus_deals_last_month),
            len(spark_plus_deals_quarter),
            len(spark_plus_deals_last_month) - len(spark_plus_deals_quarter)
        ])
        worksheet.append([
            'РПД',
            len(rpd_deals_last_month),
            len(rpd_deals_quarter),
            len(rpd_deals_last_month) - len(rpd_deals_quarter)
        ])
        worksheet.append([
            'ГРМ',
            len(grm_deals_last_month),
            len(grm_deals_quarter),
            len(grm_deals_last_month) - len(grm_deals_quarter)
        ])
        worksheet.append([
            'Доки',
            len(doki_deals_last_month),
            len(doki_deals_quarter),
            len(doki_deals_last_month) - len(doki_deals_quarter)
        ])
        worksheet.append([
            'Отчетности',
            len(report_deals_last_month),
            len(report_deals_quarter),
            len(report_deals_last_month) - len(report_deals_quarter)
        ])
        worksheet.append([
            'Подпись',
            len(signature_deals_last_month),
            len(signature_deals_quarter),
            len(signature_deals_last_month) - len(signature_deals_quarter)
        ])
        worksheet.append([
            'Допы Облако',
            len(dop_oblako_deals_last_month),
            len(dop_oblako_deals_quarter),
            len(dop_oblako_deals_last_month) - len(dop_oblako_deals_quarter)
        ])
        worksheet.append([
            'Линк',
            len(link_deals_last_month),
            len(link_deals_quarter),
            len(link_deals_last_month) - len(link_deals_quarter)
        ])
        worksheet.append([
            'Уникс',
            len(unics_deals_last_month),
            len(unics_deals_quarter),
            len(unics_deals_last_month) - len(unics_deals_quarter)
        ])
        worksheet.append([
            'Кабинет сотрудника',
            len(cab_sotrudnik_deals_last_month),
            len(cab_sotrudnik_deals_quarter),
            len(cab_sotrudnik_deals_last_month) - len(cab_sotrudnik_deals_quarter)
        ])
        worksheet.append([
            'Кабинет садовода',
            len(cab_sadovod_deals_last_month),
            len(cab_sadovod_deals_quarter),
            len(cab_sadovod_deals_last_month) - len(cab_sadovod_deals_quarter)
        ])
        worksheet.append([
            'ЭДО',
            len(edo_deals_last_month),
            len(edo_deals_quarter),
            len(edo_deals_last_month) - len(edo_deals_quarter)
        ])
        worksheet.append([
            'МДЛП',
            len(mdlp_deals_last_month),
            len(mdlp_deals_quarter),
            len(mdlp_deals_last_month) - len(mdlp_deals_quarter)
        ])
        worksheet.append([
            'Коннект',
            len(connect_deals_last_month),
            len(connect_deals_quarter),
            len(connect_deals_last_month) - len(connect_deals_quarter)
        ])
        worksheet.append([
            'ИТС Отраслевой',
            len(its_otrasl_deals_last_month),
            len(its_otrasl_deals_quarter),
            len(its_otrasl_deals_last_month) - len(its_otrasl_deals_quarter)
        ])
        worksheet.append([
            'ОФД',
            len(ofd_deals_last_month),
            len(ofd_deals_quarter),
            len(ofd_deals_last_month) - len(ofd_deals_quarter)
        ])
        if (len(bitrix24_deals_last_month) != 0) or (len(bitrix24_deals_quarter) != 0):
            worksheet.append([
                'Битрикс24',
                len(bitrix24_deals_last_month),
                len(bitrix24_deals_quarter),
                len(bitrix24_deals_last_month) - len(bitrix24_deals_quarter)
            ])
        worksheet.append([
            'Прочие',
            len(other_deals_last_month),
            len(other_deals_quarter),
            len(other_deals_last_month) - len(other_deals_quarter)
        ])
        worksheet.append([])
        
               
        # Продление
        title = ['Продление']

        #первый месяц
        deals_ended_before_3_month_dpo = list(filter(lambda x: x['Дата проверки оплаты'] and x['Ответственный'] == user_name, before_3_month_deals_data))
        deals_ended_before_3_month_dk = list(filter(lambda x: x['Ответственный'] == user_name, before_3_month_deals_data))
        deals_ended_before_3_month_dpo = list(map(lambda x: {'ID': x['ID'], 'Дата проверки оплаты': datetime.strptime(x['Дата проверки оплаты'], '%d.%m.%Y %H:%M:%S'), 'Предполагаемая дата закрытия': datetime.strptime(x['Предполагаемая дата закрытия'], '%d.%m.%Y'), 'Группа': x['Группа'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, deals_ended_before_3_month_dpo))
        deals_ended_before_3_month_dk = list(map(lambda x: {'ID': x['ID'], 'Предполагаемая дата закрытия': datetime.strptime(x['Предполагаемая дата закрытия'], '%d.%m.%Y'), 'Дата проверки оплаты': '', 'Группа': x['Группа'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, deals_ended_before_3_month_dk))
        deals_ended_before_3_month_dpo = list(filter(lambda x: datetime(day=1, month=before_2_month, year=before_2_month_year) <= x['Дата проверки оплаты'] <= datetime(day=before_2_month_range, month=before_2_month, year=before_2_month_year, hour=3), deals_ended_before_3_month_dpo))
        deals_ended_before_3_month_dpo_id = list(map(lambda x: x['ID'], deals_ended_before_3_month_dpo))
        deals_ended_before_3_month_dk = list(filter(lambda x: x['ID'] not in deals_ended_before_3_month_dpo_id and (datetime(day=1, month=before_2_month, year=before_2_month_year) <= x['Предполагаемая дата закрытия'] <= datetime(day=before_2_month_range, month=before_2_month, year=before_2_month_year)), deals_ended_before_3_month_dk))
        deals_ended_before_3_month = deals_ended_before_3_month_dk + deals_ended_before_3_month_dpo

        before_3_month_deals_data_datetime_dpo = list(filter(lambda x: x['Ответственный'] == user_name and x['Дата проверки оплаты'], before_1_month_deals_data))
        before_3_month_deals_data_datetime_dpo = list(map(lambda x: {'ID': x['ID'], 'Дата проверки оплаты': datetime.strptime(x['Дата проверки оплаты'], '%d.%m.%Y %H:%M:%S'), 'Предполагаемая дата закрытия': '', 'Стадия': x['Стадия сделки'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, before_3_month_deals_data_datetime_dpo))
        before_3_month_deals_data_datetime_dk = list(filter(lambda x: x['Ответственный'] == user_name, before_1_month_deals_data))
        before_3_month_deals_data_datetime_dk = list(map(lambda x: {'ID': x['ID'], 'Предполагаемая дата закрытия': datetime.strptime(x['Предполагаемая дата закрытия'], '%d.%m.%Y'), 'Дата проверки оплаты': '', 'Стадия': x['Стадия сделки'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, before_3_month_deals_data_datetime_dk))
        non_extended_date_deals_3 = []
        before_3_month_deals_data_datetime = before_3_month_deals_data_datetime_dpo + before_3_month_deals_data_datetime_dk

        for deal in deals_ended_before_3_month:
            before_3_month_deal = list(filter(lambda x: x['ID'] == deal['ID'], before_3_month_deals_data_datetime))
            if not before_3_month_deal:
                non_extended_date_deals_3.append(deal)
            else:
                if not before_3_month_deal[0]['Дата проверки оплаты'] or not deal['Дата проверки оплаты']:
                    continue
                if before_3_month_deal[0]['Дата проверки оплаты'] <= deal['Дата проверки оплаты'] and before_3_month_deal[0]['Стадия'] != 'Услуга завершена':
                    non_extended_date_deals_3.append(deal)

        for deal in deals_ended_before_3_month:
            if deal in non_extended_date_deals_3:
                continue
            before_3_month_deal = list(filter(lambda x: x['ID'] == deal['ID'], before_3_month_deals_data_datetime))
            if not before_3_month_deal:
                non_extended_date_deals_3.append(deal)
            else:
                if not before_3_month_deal[0]['Предполагаемая дата закрытия']:
                    continue
                if before_3_month_deal[0]['Предполагаемая дата закрытия'] <= deal['Предполагаемая дата закрытия'] and before_3_month_deal[0]['Стадия'] != 'Услуга завершена':
                    non_extended_date_deals_3.append(deal)

        ended_its_2 = set(map(lambda x: x['ID'], list(filter(lambda x: x['Группа'] == 'ИТС', deals_ended_before_3_month))))
        ended_reporting_2 = set(map(lambda x: x['ID'], list(filter(lambda x: x['Группа'] == 'Сервисы ИТС', deals_ended_before_3_month))))
        ended_others_2 = set(map(lambda x: x['ID'], list(filter(lambda x: x['Группа'] not in ['Сервисы ИТС', 'ИТС'], deals_ended_before_3_month))))
        non_extended_date_deals_id_2 = set(map(lambda x: x['ID'], non_extended_date_deals_3))

        title.extend([f'Заканчивалось на {datetime(day=before_2_month_range, month=before_2_month, year=before_2_month_year).strftime("%d.%m.%Y")}', 'Из них продлено', 'Не продлено'])
        table_its =  ['ИТС', len(ended_its_2), len(set(filter(lambda x: x not in non_extended_date_deals_id_2, ended_its_2))), len(set(filter(lambda x: x in non_extended_date_deals_id_2, ended_its_2)))]
        table_service = ['Сервисы', len(ended_reporting_2), len(set(filter(lambda x: x not in non_extended_date_deals_id_2, ended_reporting_2))), len(set(filter(lambda x: x in non_extended_date_deals_id_2, ended_reporting_2)))]
        table_other = ['Остальное', len(ended_others_2), len(set(filter(lambda x: x not in non_extended_date_deals_id_2, ended_others_2))), len(set(filter(lambda x: x in non_extended_date_deals_id_2, ended_others_2)))]

        count_its = len(ended_its_2)
        count_service = len(ended_reporting_2)
        count_other = len(ended_others_2)
        count_its_non = len(set(filter(lambda x: x in non_extended_date_deals_id_2, ended_its_2)))
        count_service_non = len(set(filter(lambda x: x in non_extended_date_deals_id_2, ended_reporting_2)))
        count_other_non = len(set(filter(lambda x: x in non_extended_date_deals_id_2, ended_others_2)))
        
        #второй месяц
        if before_1_month not in [1, 4, 7, 10]:
            deals_ended_before_4_month_dpo = list(filter(lambda x: x['Дата проверки оплаты'] and x['Ответственный'] == user_name, before_4_month_deals_data))
            deals_ended_before_4_month_dk = list(filter(lambda x: x['Ответственный'] == user_name, before_4_month_deals_data))
            deals_ended_before_4_month_dpo = list(map(lambda x: {'ID': x['ID'], 'Дата проверки оплаты': datetime.strptime(x['Дата проверки оплаты'], '%d.%m.%Y %H:%M:%S'), 'Предполагаемая дата закрытия': datetime.strptime(x['Предполагаемая дата закрытия'], '%d.%m.%Y'), 'Группа': x['Группа'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, deals_ended_before_4_month_dpo))
            deals_ended_before_4_month_dk = list(map(lambda x: {'ID': x['ID'], 'Предполагаемая дата закрытия': datetime.strptime(x['Предполагаемая дата закрытия'], '%d.%m.%Y'), 'Дата проверки оплаты': '', 'Группа': x['Группа'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, deals_ended_before_4_month_dk))
            deals_ended_before_4_month_dpo = list(filter(lambda x: datetime(day=1, month=before_3_month, year=before_3_month_year) <= x['Дата проверки оплаты'] <= datetime(day=before_3_month_range, month=before_3_month, year=before_3_month_year, hour=3), deals_ended_before_4_month_dpo))
            deals_ended_before_4_month_dpo_id = list(map(lambda x: x['ID'], deals_ended_before_4_month_dpo))
            deals_ended_before_4_month_dk = list(filter(lambda x: x['ID'] not in deals_ended_before_4_month_dpo_id and (datetime(day=1, month=before_3_month, year=before_3_month_year) <= x['Предполагаемая дата закрытия'] <= datetime(day=before_3_month_range, month=before_3_month, year=before_3_month_year)), deals_ended_before_4_month_dk))
            deals_ended_before_4_month = deals_ended_before_4_month_dk + deals_ended_before_4_month_dpo

            before_4_month_deals_data_datetime_dpo = list(filter(lambda x: x['Ответственный'] == user_name and x['Дата проверки оплаты'], before_2_month_deals_data))
            before_4_month_deals_data_datetime_dpo = list(map(lambda x: {'ID': x['ID'], 'Дата проверки оплаты': datetime.strptime(x['Дата проверки оплаты'], '%d.%m.%Y %H:%M:%S'), 'Предполагаемая дата закрытия': '', 'Стадия': x['Стадия сделки'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, before_4_month_deals_data_datetime_dpo))
            before_4_month_deals_data_datetime_dk = list(filter(lambda x: x['Ответственный'] == user_name, before_2_month_deals_data))
            before_4_month_deals_data_datetime_dk = list(map(lambda x: {'ID': x['ID'], 'Предполагаемая дата закрытия': datetime.strptime(x['Предполагаемая дата закрытия'], '%d.%m.%Y'), 'Дата проверки оплаты': '', 'Стадия': x['Стадия сделки'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, before_4_month_deals_data_datetime_dk))
            non_extended_date_deals_4 = []
            before_4_month_deals_data_datetime = before_4_month_deals_data_datetime_dpo + before_4_month_deals_data_datetime_dk

            for deal in deals_ended_before_4_month:
                before_4_month_deal = list(filter(lambda x: x['ID'] == deal['ID'], before_4_month_deals_data_datetime))
                if not before_4_month_deal:
                    non_extended_date_deals_4.append(deal)
                else:
                    if not before_4_month_deal[0]['Дата проверки оплаты'] or not deal['Дата проверки оплаты']:
                        continue
                    if before_4_month_deal[0]['Дата проверки оплаты'] <= deal['Дата проверки оплаты'] and before_4_month_deal[0]['Стадия'] != 'Услуга завершена':
                        non_extended_date_deals_4.append(deal)

            for deal in deals_ended_before_4_month:
                if deal in non_extended_date_deals_4:
                    continue
                before_4_month_deal = list(filter(lambda x: x['ID'] == deal['ID'], before_4_month_deals_data_datetime))
                if not before_4_month_deal:
                    non_extended_date_deals_4.append(deal)
                else:
                    if not before_4_month_deal[0]['Предполагаемая дата закрытия']:
                        continue
                    if before_4_month_deal[0]['Предполагаемая дата закрытия'] <= deal['Предполагаемая дата закрытия'] and before_4_month_deal[0]['Стадия'] != 'Услуга завершена':
                        non_extended_date_deals_4.append(deal)

            ended_its_3 = set(map(lambda x: x['ID'], list(filter(lambda x: x['Группа'] == 'ИТС', deals_ended_before_4_month))))
            ended_reporting_3 = set(map(lambda x: x['ID'], list(filter(lambda x: x['Группа'] == 'Сервисы ИТС', deals_ended_before_4_month))))
            ended_others_3 = set(map(lambda x: x['ID'], list(filter(lambda x: x['Группа'] not in ['Сервисы ИТС', 'ИТС'], deals_ended_before_4_month))))
            non_extended_date_deals_id_3 = set(map(lambda x: x['ID'], non_extended_date_deals_4))

            title.extend([f'Заканчивалось на {datetime(day=before_3_month_range, month=before_3_month, year=before_3_month_year).strftime("%d.%m.%Y")}', 'Из них продлено', 'Не продлено'])
            table_its.extend([len(ended_its_3), len(set(filter(lambda x: x not in non_extended_date_deals_id_3, ended_its_3))), len(set(filter(lambda x: x in non_extended_date_deals_id_3, ended_its_3)))])
            table_service.extend([len(ended_reporting_3), len(set(filter(lambda x: x not in non_extended_date_deals_id_3, ended_reporting_3))), len(set(filter(lambda x: x in non_extended_date_deals_id_3, ended_reporting_3)))])
            table_other.extend([len(ended_others_3), len(set(filter(lambda x: x not in non_extended_date_deals_id_3, ended_others_3))), len(set(filter(lambda x: x in non_extended_date_deals_id_3, ended_others_3)))])

            count_its += len(ended_its_3)
            count_service += len(ended_reporting_3)
            count_other += len(ended_others_3)
            count_its_non += len(set(filter(lambda x: x in non_extended_date_deals_id_3, ended_its_3)))
            count_service_non += len(set(filter(lambda x: x in non_extended_date_deals_id_3, ended_reporting_3)))
            count_other_non += len(set(filter(lambda x: x in non_extended_date_deals_id_3, ended_others_3)))
        
        #третий месяц
        if before_1_month not in [2, 5, 8, 11, 1, 4, 7, 10]:
            deals_ended_before_5_month_dpo = list(filter(lambda x: x['Дата проверки оплаты'] and x['Ответственный'] == user_name, before_5_month_deals_data))
            deals_ended_before_5_month_dk = list(filter(lambda x: x['Ответственный'] == user_name, before_5_month_deals_data))
            deals_ended_before_5_month_dpo = list(map(lambda x: {'ID': x['ID'], 'Дата проверки оплаты': datetime.strptime(x['Дата проверки оплаты'], '%d.%m.%Y %H:%M:%S'), 'Предполагаемая дата закрытия': datetime.strptime(x['Предполагаемая дата закрытия'], '%d.%m.%Y'), 'Группа': x['Группа'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, deals_ended_before_5_month_dpo))
            deals_ended_before_5_month_dk = list(map(lambda x: {'ID': x['ID'], 'Предполагаемая дата закрытия': datetime.strptime(x['Предполагаемая дата закрытия'], '%d.%m.%Y'), 'Дата проверки оплаты': '', 'Группа': x['Группа'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, deals_ended_before_5_month_dk))
            deals_ended_before_5_month_dpo = list(filter(lambda x: datetime(day=1, month=before_4_month, year=before_4_month_year) <= x['Дата проверки оплаты'] <= datetime(day=before_4_month_range, month=before_4_month, year=before_4_month_year, hour=3), deals_ended_before_5_month_dpo))
            deals_ended_before_5_month_dpo_id = list(map(lambda x: x['ID'], deals_ended_before_5_month_dpo))
            deals_ended_before_5_month_dk = list(filter(lambda x: x['ID'] not in deals_ended_before_5_month_dpo_id and (datetime(day=1, month=before_4_month, year=before_4_month_year) <= x['Предполагаемая дата закрытия'] <= datetime(day=before_4_month_range, month=before_4_month, year=before_4_month_year)), deals_ended_before_5_month_dk))
            deals_ended_before_5_month = deals_ended_before_5_month_dk + deals_ended_before_5_month_dpo

            before_5_month_deals_data_datetime_dpo = list(filter(lambda x: x['Ответственный'] == user_name and x['Дата проверки оплаты'], before_3_month_deals_data))
            before_5_month_deals_data_datetime_dpo = list(map(lambda x: {'ID': x['ID'], 'Дата проверки оплаты': datetime.strptime(x['Дата проверки оплаты'], '%d.%m.%Y %H:%M:%S'), 'Предполагаемая дата закрытия': '', 'Стадия': x['Стадия сделки'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, before_5_month_deals_data_datetime_dpo))
            before_5_month_deals_data_datetime_dk = list(filter(lambda x: x['Ответственный'] == user_name, before_3_month_deals_data))
            before_5_month_deals_data_datetime_dk = list(map(lambda x: {'ID': x['ID'], 'Предполагаемая дата закрытия': datetime.strptime(x['Предполагаемая дата закрытия'], '%d.%m.%Y'), 'Дата проверки оплаты': '', 'Стадия': x['Стадия сделки'], 'Регномер': x['Регномер'], 'Компания': x['Компания'], 'Тип': x['Тип']}, before_5_month_deals_data_datetime_dk))
            non_extended_date_deals_5 = []
            before_5_month_deals_data_datetime = before_5_month_deals_data_datetime_dpo + before_5_month_deals_data_datetime_dk

            for deal in deals_ended_before_5_month:
                before_5_month_deal = list(filter(lambda x: x['ID'] == deal['ID'], before_5_month_deals_data_datetime))
                if not before_5_month_deal:
                    non_extended_date_deals_5.append(deal)
                else:
                    if not before_5_month_deal[0]['Дата проверки оплаты'] or not deal['Дата проверки оплаты']:
                        continue
                    if before_5_month_deal[0]['Дата проверки оплаты'] <= deal['Дата проверки оплаты'] and before_5_month_deal[0]['Стадия'] != 'Услуга завершена':
                        non_extended_date_deals_5.append(deal)

            for deal in deals_ended_before_5_month:
                if deal in non_extended_date_deals_5:
                    continue
                before_5_month_deal = list(filter(lambda x: x['ID'] == deal['ID'], before_5_month_deals_data_datetime))
                if not before_5_month_deal:
                    non_extended_date_deals_5.append(deal)
                else:
                    if not before_5_month_deal[0]['Предполагаемая дата закрытия']:
                        continue
                    if before_5_month_deal[0]['Предполагаемая дата закрытия'] <= deal['Предполагаемая дата закрытия'] and before_5_month_deal[0]['Стадия'] != 'Услуга завершена':
                        non_extended_date_deals_5.append(deal)

            ended_its_4 = set(map(lambda x: x['ID'], list(filter(lambda x: x['Группа'] == 'ИТС', deals_ended_before_5_month))))
            ended_reporting_4 = set(map(lambda x: x['ID'], list(filter(lambda x: x['Группа'] == 'Сервисы ИТС', deals_ended_before_5_month))))
            ended_others_4 = set(map(lambda x: x['ID'], list(filter(lambda x: x['Группа'] not in ['Сервисы ИТС', 'ИТС'], deals_ended_before_5_month))))
            non_extended_date_deals_id_4 = set(map(lambda x: x['ID'], non_extended_date_deals_5))

            title.extend([f'Заканчивалось на {datetime(day=before_4_month_range, month=before_4_month, year=before_4_month_year).strftime("%d.%m.%Y")}', 'Из них продлено', 'Не продлено'])
            table_its.extend([len(ended_its_4), len(set(filter(lambda x: x not in non_extended_date_deals_id_4, ended_its_4))), len(set(filter(lambda x: x in non_extended_date_deals_id_4, ended_its_4)))])
            table_service.extend([len(ended_reporting_4), len(set(filter(lambda x: x not in non_extended_date_deals_id_4, ended_reporting_4))), len(set(filter(lambda x: x in non_extended_date_deals_id_4, ended_reporting_4)))])
            table_other.extend([len(ended_others_4), len(set(filter(lambda x: x not in non_extended_date_deals_id_4, ended_others_4))), len(set(filter(lambda x: x in non_extended_date_deals_id_4, ended_others_4)))])\
            
            count_its += len(ended_its_4)
            count_service += len(ended_reporting_4)
            count_other += len(ended_others_4)
            count_its_non += len(set(filter(lambda x: x in non_extended_date_deals_id_4, ended_its_4)))
            count_service_non += len(set(filter(lambda x: x in non_extended_date_deals_id_4, ended_reporting_4)))
            count_other_non += len(set(filter(lambda x: x in non_extended_date_deals_id_4, ended_others_4)))
        
        #добавляем заголовки и сновную часть в таблицу
        title.extend(['Всего к продлению', 'Всего не продлено'])
        worksheet.append(title)

        table_its.extend([count_its, count_its_non])
        table_service.extend([count_service, count_service_non])
        table_other.extend([count_other, count_other_non])

        worksheet.append(table_its)
        worksheet.append(table_service)
        worksheet.append(table_other)
        
        worksheet.append([])
        ''' 
        '''
        # Охват сервисами
        #Последний месяц квартала
        companies = set(map(lambda x: x['Компания'], list(filter(lambda x: x['Ответственный за компанию'] == user_name, before_1_month_deals_data))))
        companies_without_services_last_month = 0
        companies_without_paid_services_last_month = 0
        for company in companies:
            company_regnumbers = set(map(lambda x: x['Регномер'], list(filter(lambda x: x['Компания'] == company, before_1_month_deals_data)))) #читаем все сделки с заданным регномером за последний месяц
            company_its = list(filter(lambda x: x['Группа'] == 'ИТС' and company == x['Компания'], before_1_month_deals_data)) #читаем все сделки из группы итс за последний месяц
            if not company_its:
                continue

            non_prof_its = list(filter(lambda x: x['Группа'] == 'ИТС' and company == x['Компания'] and 'ПРОФ' not in x['Тип'] and 'Облако' not in x['Тип'] and 'ГРМ' not in x['Тип'], before_1_month_deals_data))
            if non_prof_its:
                company_its_services = list(filter(lambda x: (company == x['Компания'] or x['Регномер'] in company_regnumbers) and
                                                   ('Контрагент' in x['Тип'] or
                                                   'Спарк' in x['Тип'] or
                                                   'РПД' in x['Тип'] or
                                                   'Отчетность' in x['Тип'] or
                                                   'Допы Облако' in x['Тип'] or
                                                   'Кабинет сотрудника' in x['Тип']),
                                                    before_1_month_deals_data))

                if not company_its_services:
                    companies_without_services_last_month += 1

            company_its_paid_services = list(filter(lambda x: (company == x['Компания'] or x['Регномер'] in company_regnumbers) and
                                                    ('Контрагент' in x['Тип'] or
                                                    'Спарк' in x['Тип'] or
                                                    'РПД' in x['Тип'] or
                                                    'Отчетность' == x['Тип'] or
                                                    'Допы Облако' in x['Тип'] or
                                                    'Кабинет сотрудника' in x['Тип']), before_1_month_deals_data))

            if not company_its_paid_services:
                companies_without_paid_services_last_month += 1

        try:
            coverage_its_without_services_last_month = round(companies_without_services_last_month /
                                                                    len(its_deals_before_1_month) * 100, 2)
        except ZeroDivisionError:
            coverage_its_without_services_last_month = 0

        try:
            coverage_its_without_paid_services_last_month = round(companies_without_paid_services_last_month /
                                                                    len(its_deals_before_1_month) * 100, 2)
        except ZeroDivisionError:
            coverage_its_without_paid_services_last_month = 0

        #Начало квартала
        companies = set(map(lambda x: x['Компания'], list(filter(lambda x: x['Ответственный за компанию'] == user_name, quarter_deals_data))))
        companies_without_services_start_quarter = 0
        companies_without_paid_services_start_quarter = 0
        for company in companies:
            company_regnumbers = set(map(lambda x: x['Регномер'], list(filter(lambda x: x['Компания'] == company, quarter_deals_data))))
            company_its = list(filter(lambda x: x['Группа'] == 'ИТС' and company == x['Компания'], quarter_deals_data))
            if not company_its:
                continue

            non_prof_its = list(filter(lambda x: x['Группа'] == 'ИТС' and company == x['Компания'] and 'ПРОФ' not in x[
                'Тип'] and 'Облако' not in x['Тип'] and 'ГРМ' not in x['Тип'], quarter_deals_data))
            if non_prof_its:
                company_its_services = list(filter(lambda x: (company == x['Компания'] or x['Регномер'] in company_regnumbers) and
                                                    ('Контрагент' in x['Тип'] or
                                                    'Спарк' in x['Тип'] or
                                                    'РПД' in x['Тип'] or
                                                    'Отчетность' in x['Тип'] or
                                                    'Допы Облако' in x['Тип'] or
                                                    'Кабинет сотрудника' in x['Тип']), quarter_deals_data))

                if not company_its_services:
                    companies_without_services_start_quarter += 1

            company_its_paid_services = list(filter(lambda x: (company == x['Компания'] or x['Регномер'] in company_regnumbers) and
                                                    ('Контрагент' in x['Тип'] or
                                                    'Спарк' in x['Тип'] or
                                                    'РПД' in x['Тип'] or
                                                    'Отчетность' == x['Тип'] or
                                                    'Допы Облако' in x['Тип'] or
                                                    'Кабинет сотрудника' in x['Тип']), quarter_deals_data))

            if not company_its_paid_services:
                companies_without_paid_services_start_quarter += 1

        try:
            coverage_its_without_services_start_quarter = round(companies_without_services_start_quarter /
                                                                   len(start_quarter_its_deals) * 100, 2)
        except ZeroDivisionError:
            coverage_its_without_services_start_quarter = 0

        try:
            coverage_its_without_paid_services_start_quarter = round(companies_without_paid_services_start_quarter /
                                                                        len(start_quarter_its_deals) * 100, 2)            
        except ZeroDivisionError:
            coverage_its_without_paid_services_start_quarter = 0

        
        worksheet.append(['Охват сервисами', f'на {before_1_month_last_day_date}', f'на {start_date_quarter.strftime("%d.%m.%Y")}', 'Прирост с начала квартала'])
        worksheet.append([
            'ИТС без сервисов',
            companies_without_services_last_month,
            companies_without_services_start_quarter,
            companies_without_services_last_month - companies_without_services_start_quarter,
        ])
        worksheet.append([
            '% ИТС без сервисов',
            f'{coverage_its_without_services_last_month}%',
            f'{coverage_its_without_services_start_quarter}%',
            f'{round(coverage_its_without_services_last_month - coverage_its_without_services_start_quarter, 2)}%'
        ])
        worksheet.append([
            'ИТС без платных сервисов',
            companies_without_paid_services_last_month,
            companies_without_paid_services_start_quarter,
            companies_without_paid_services_last_month - companies_without_paid_services_start_quarter
        ])
        worksheet.append([
            '% ИТС без платных сервисов',
            f'{coverage_its_without_paid_services_last_month}%',
            f'{coverage_its_without_paid_services_start_quarter}%',
            f'{round(coverage_its_without_paid_services_last_month - coverage_its_without_paid_services_start_quarter, 2)}%'
        ])
        worksheet.append([])


        # Отчетность
        # Прошлый месяц
        free_reporting_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                                      x['Тип'] == 'Отчетность (в рамках ИТС)' and
                                                      x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                      before_1_month_deals_data))

        try:
            coverage_free_reporting_deals_last_month = round(len(free_reporting_deals_last_month) /
                                                             len(its_prof_deals_last_month) * 100, 2)
        except ZeroDivisionError:
            coverage_free_reporting_deals_last_month = 0

        paid_reporting_deals_last_month_num = 0
        paid_reporting_deals_last_month = 0
        for its_deal in its_deals_before_1_month:
            its_paid_reporting = list(filter(lambda x: (x['Регномер'] == its_deal['Регномер'] and x['Тип'] == 'Отчетность' and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту']), before_1_month_deals_data))
            if its_paid_reporting:
                paid_reporting_deals_last_month += 1
                for i in range(len(its_paid_reporting)):
                    paid_reporting_deals_last_month_num +=1
        try:
            coverage_paid_reporting_deals_last_month = round(paid_reporting_deals_last_month /
                                                             len(its_deals_before_1_month) * 100, 2)
        except ZeroDivisionError:
            coverage_paid_reporting_deals_last_month = 0

        
        any_reporting_deals_last_month = 0
        for its_deal in its_deals_before_1_month:
            any_reporting = list(filter(lambda x: (x['Регномер'] == its_deal['Регномер'] and 'Отчетность' in x['Тип'] and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту']), before_1_month_deals_data))
            if any_reporting:
                any_reporting_deals_last_month += 1
        try:
            coverage_any_reporting_deals_last_month = round(any_reporting_deals_last_month /
                                                             len(its_deals_before_1_month) * 100, 2)
        except ZeroDivisionError:
            coverage_any_reporting_deals_last_month = 0

        # Начало квартала
        free_reporting_deals_start_quarter = list(filter(lambda x: x['Ответственный'] == user_name and
                                                      x['Тип'] == 'Отчетность (в рамках ИТС)' and
                                                      x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                      quarter_deals_data))

        prof_deals_start_quarter = list(filter(lambda x: x['Ответственный'] == user_name and
                                            x['Группа'] == 'ИТС' and
                                            'Базовый' not in x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            quarter_deals_data))

        try:
            coverage_free_reporting_deals_start_quarter = round(len(free_reporting_deals_start_quarter) /
                                                             len(prof_deals_start_quarter) * 100, 2)
        except ZeroDivisionError:
            coverage_free_reporting_deals_start_quarter = 0

        paid_reporting_deals_start_quarter = 0
        paid_reporting_deals_start_quarter_num = 0
        for its_deal in start_quarter_its_deals:
            its_paid_reporting = list(
                filter(lambda x: (x['Регномер'] == its_deal['Регномер'] and x['Тип'] == 'Отчетность' and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту']),
                       quarter_deals_data))
            if its_paid_reporting:
                paid_reporting_deals_start_quarter += 1
                for i in range(len(its_paid_reporting)):
                    paid_reporting_deals_start_quarter_num +=1
        try:
            coverage_paid_reporting_deals_start_quarter = round(paid_reporting_deals_start_quarter /
                                                                len(start_quarter_its_deals) * 100, 2)
        except ZeroDivisionError:
            coverage_paid_reporting_deals_start_quarter = 0

        any_reporting_deals_start_quarter = 0
        for its_deal in start_quarter_its_deals:
            any_reporting = list(filter(lambda x: (x['Регномер'] == its_deal['Регномер'] and 'Отчетность' in x['Тип'] and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту']), quarter_deals_data))
            if any_reporting:
                any_reporting_deals_start_quarter += 1
        try:
            coverage_any_reporting_deals_start_quarter = round(any_reporting_deals_start_quarter /
                                                             len(start_quarter_its_deals) * 100, 2)
        except ZeroDivisionError:
            coverage_any_reporting_deals_start_quarter = 0


        worksheet.append(['Отчетность', f'на {before_1_month_last_day_date}', f'на {start_date_quarter.strftime("%d.%m.%Y")}', 'Прирост с начала квартала'])
        worksheet.append([
            'Льготных отчетностей',
            len(free_reporting_deals_last_month),
            len(free_reporting_deals_start_quarter),
            len(free_reporting_deals_last_month) - len(free_reporting_deals_start_quarter),
        ])
        worksheet.append([
            'Охват льготной отчетностью',
            f'{coverage_free_reporting_deals_last_month}%',
            f'{coverage_free_reporting_deals_start_quarter}%',
            f'{round(coverage_free_reporting_deals_last_month - coverage_free_reporting_deals_start_quarter, 2)}%'
        ])
        worksheet.append([
            'Сделок платных отчетностей',
            paid_reporting_deals_last_month_num,
            paid_reporting_deals_start_quarter_num,
            paid_reporting_deals_last_month_num - paid_reporting_deals_start_quarter_num,
        ])
        worksheet.append([
            'ИТС с платной отчетностью',
            paid_reporting_deals_last_month,
            paid_reporting_deals_start_quarter,
            paid_reporting_deals_last_month - paid_reporting_deals_start_quarter,
        ])
        worksheet.append([
            'Охват платных отчетностей',
            f'{coverage_paid_reporting_deals_last_month}%',
            f'{coverage_paid_reporting_deals_start_quarter}%',
            f'{round(coverage_paid_reporting_deals_last_month - coverage_paid_reporting_deals_start_quarter, 2)}%'
        ])
        worksheet.append([
            'Любая отчетность',
            any_reporting_deals_last_month,
            any_reporting_deals_start_quarter,
            any_reporting_deals_last_month - any_reporting_deals_start_quarter,
        ])
        worksheet.append([
            'Охват любой отчетностью',
            f'{coverage_any_reporting_deals_last_month}%',
            f'{coverage_any_reporting_deals_start_quarter}%',
            f'{round(coverage_any_reporting_deals_last_month - coverage_any_reporting_deals_start_quarter, 2)}%',
        ])

        worksheet.append([])

        # Продажи
        sales = b.get_all('crm.item.list', {
            'entityTypeId': '133',
            'filter': {
                'assignedById': user_info['ID'],
                '>=ufCrm3_1654248264': quarter_filters['start_date'].strftime(ddmmyyyy_pattern),
                '<ufCrm3_1654248264': quarter_filters['end_date'].strftime(ddmmyyyy_pattern),
            }
        })

        list_of_sales = ([{'NAME_DEAL': 'Название сделки', 'COMPANY': 'Компания', 'OPPORTUNITY': 'Сумма'}])

        if sales:
            sold_deals = b.get_all('crm.deal.list', {
                'select': ['ID', 'COMPANY_ID', 'UF_CRM_1657878818384', 'OPPORTUNITY', 'TYPE_ID'],
                'filter': {
                    'ID': list(map(lambda x: x['parentId2'], sales))
                }
            })
            company_titles = b.get_all('crm.company.list', {
                'select': ['ID', 'TITLE'],
                'filter': {
                    'ID': list(map(lambda x: x['COMPANY_ID'], sold_deals))
                }
            })
            for deal_quarter in sold_deals:
                #print (deal_quarter)
                try:
                    deal = list(filter(lambda x: x['ID'] == deal_quarter['ID'], before_1_month_deals_data))
                    if not deal:
                        deal = list(filter(lambda x: x['ID'] == deal_quarter['ID'], before_2_month_deals_data))
                    if not deal:
                        deal = list(filter(lambda x: x['ID'] == deal_quarter['ID'], before_3_month_deals_data))
                    if deal:
                        deal = deal[0]
                        title = list(set(map(lambda x: x['TITLE'], list(filter(lambda x: x['ID'] == deal['Компания'], company_titles)))))
                        list_of_sales.append({'NAME_DEAL': deal['Название сделки'], 'COMPANY': title[0], 'OPPORTUNITY': deal['Сумма']})
                except:
                    #2024-09-10 saa
                    users_id = ['1391', '1']
                    for user_id in users_id:
                        b.call('im.notify.system.add', {
                            'USER_ID': user_id,
                            'MESSAGE': f'Проблемы при поиске сделки в файле по источнику продаж\n\n{deal_quarter}'})
        else:
            sold_deals = []

        worksheet.append(['Продажи', f'{number_quarter} квартал {end_date_quarter.year} шт.', f'{number_quarter} квартал {end_date_quarter.year} руб'])
        for field_value in deal_group_field:
            if field_value['VALUE'] == 'Лицензии':
                grouped_deals = list(filter(lambda x: x['TYPE_ID'] in ['UC_YIAJC8', 'UC_QQPYF0'], sold_deals)) # тип лицензия с купоном или лицензия
            else:
                grouped_deals = list(filter(lambda x: x['UF_CRM_1657878818384'] == field_value['ID'] and x['TYPE_ID'] not in ['UC_YIAJC8', 'UC_QQPYF0'], sold_deals)) # если группы равны и тип не лицензии
            worksheet.append([field_value['VALUE'], len(grouped_deals), sum(list(map(lambda x: float(x['OPPORTUNITY'] if x['OPPORTUNITY'] else 0.0), grouped_deals)))])

        worksheet.append(['Всего по источникам', len(sales), sum(list(map(lambda x: x['opportunity'], sales)))])
        worksheet.append(['Всего по сделкам', len(sold_deals), sum(list(map(lambda x: float(x['OPPORTUNITY']), sold_deals)))])
        if len(list_of_sales) > 1:
            worksheet.append(['', 'Перечень продаж', ''])
            for selling in list_of_sales:
                worksheet.append([selling['NAME_DEAL'], selling['COMPANY'], selling['OPPORTUNITY']])
        worksheet.append([])
        

        # Долги по документам
        documents_debts = b.get_all('crm.item.list', { #все выписанные долги
            'entityTypeId': '161',
            'filter': {
                'assignedById': user_info['ID'],
                '<ufCrm41_1689101272': quarter_filters['end_date'].strftime(ddmmyyyy_pattern)
            }
        })

        non_documents_debts = list(filter(lambda x: x['stageId'] in ['DT161_53:NEW', 'DT161_53:1'], documents_debts)) #все незакрытые долги

        quarter_documents_debts = list(filter(lambda x: #все выписанные долги за квартал
                                              quarter_filters['start_date'].timestamp() <=
                                              datetime.fromisoformat(x['ufCrm41_1689101272']).timestamp(), documents_debts))
        
        non_quarter_documents_debts = list(filter(lambda x: #все незакрытые долги за квартал
                                        quarter_filters['start_date'].timestamp() <=
                                        datetime.fromisoformat(x['ufCrm41_1689101272']).timestamp(), non_documents_debts))


        non_last_documents_debts = len(non_documents_debts) - len(non_quarter_documents_debts) #все незакрытые долги за предыдущие периоды

        worksheet.append(['Долги по документам', 'Всего выписано за квартал', 'Не сдано за квартал', 'Не сдано за предыдущие периоды'])
        worksheet.append(['Штук', len(quarter_documents_debts), len(non_quarter_documents_debts), non_last_documents_debts])
        worksheet.append([])
       
        # Задачи
        tasks = b.get_all('tasks.task.list', {
            'filter': {
                'RESPONSIBLE_ID': user_info['ID'],
                '>=CREATED_DATE': quarter_filters['start_date'].strftime(ddmmyyyy_pattern),
                '<CREATED_DATE': quarter_filters['end_date'].strftime(ddmmyyyy_pattern),
            },
            'select': ['GROUP_ID', 'STATUS', 'UF_AUTO_177856763915']
        })
        completed_tasks = list(filter(lambda x: x['status'] == '5', tasks))
        service_tasks = list(filter(lambda x: x['groupId'] == '71', tasks))
        completed_service_tasks = list(filter(lambda x: x['status'] == '5', service_tasks))
        completed_other_tasks = list(filter(lambda x: x['status'] == '5' and x['groupId'] != '71', tasks))
        non_completed_other_tasks = list(filter(lambda x: x['status'] != '5' and x['groupId'] != '71', tasks))
        completed_tlp_tasks = list(filter(lambda x: x['groupId'] == '1' and x['status'] == '5', tasks))
        tasks_ratings = list(map(lambda x: int(x['ufAuto177856763915']) if x['ufAuto177856763915'] else 0, completed_tlp_tasks))
        tasks_ratings = list(filter(lambda x: x != 0, tasks_ratings))
        try:
            average_tasks_ratings = round(sum(tasks_ratings) / len(tasks_ratings), 2)
        except ZeroDivisionError:
            average_tasks_ratings = '-'

        #Дежурства
        days_duty = b.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '301',
            'filter': {
                'PROPERTY_1753': user_info['ID'],
                '>=PROPERTY_1769': int(quarter_filters['start_date'].month),
                '<=PROPERTY_1769': int(end_date_quarter.month),
                'PROPERTY_1771': int(end_date_quarter.year),
            }
        })
        days_duty_amount = len(days_duty)

        worksheet.append(['', 'Незакрытых (и созд. в этом кварт)', 'Всего создано в квартале'])
        worksheet.append(['Незакрытые задачи', len(tasks) - len(completed_tasks), len(tasks)])
        worksheet.append(['СВ', len(service_tasks) - len(completed_service_tasks), len(service_tasks)])
        worksheet.append(['Остальные', len(non_completed_other_tasks), len(completed_other_tasks) + len(non_completed_other_tasks)])
        worksheet.append([])
        worksheet.append(['Закрытые задачи ТЛП', len(completed_tlp_tasks)])
        worksheet.append(['Средняя оценка', average_tasks_ratings])
        worksheet.append(['Дней дежурства', days_duty_amount])
        worksheet.append([])
        

        # ЭДО
        all_its_last_month = its_prof_deals_last_month + its_base_deals_last_month
        edo_companies_id_last_month = list(map(lambda x: x['Компания'], list(filter(lambda y: 'Компания' in y and y['Компания'], all_its_last_month))))

        if edo_companies_id_last_month:
            edo_companies_last_month = b.get_all('crm.company.list', {
                'select': ['UF_CRM_1638093692254', 'UF_CRM_1638093750742'],
                'filter': {
                    'ID': list(map(lambda x: x['Компания'], list(filter(lambda y: 'Компания' in y and y['Компания'], all_its_last_month))))
                }
            })
            edo_companies_count_last_month = list(filter(lambda x: x['UF_CRM_1638093692254'] == '69', edo_companies_last_month))

            #спецоператоры эдо
            try: 
                operator_2ae = list(filter(lambda x: x['UF_CRM_1638093750742'] == '75', edo_companies_count_last_month))
            except ZeroDivisionError:
                operator_2ae = 0
            try: 
                operator_2ae_doki = list(filter(lambda x: x['UF_CRM_1638093750742'] == '1715', edo_companies_count_last_month))
            except ZeroDivisionError:
                operator_2ae_doki = 0
            try: 
                operator_2be = list(filter(lambda x: x['UF_CRM_1638093750742'] == '77', edo_companies_count_last_month))
            except ZeroDivisionError:
                operator_2be = 0
            try: 
                operator_2bm = list(filter(lambda x: x['UF_CRM_1638093750742'] == '73', edo_companies_count_last_month))
            except ZeroDivisionError:
                operator_2bm = 0
            try: 
                operator_2al = list(filter(lambda x: x['UF_CRM_1638093750742'] == '437', edo_companies_count_last_month))
            except ZeroDivisionError:
                operator_2al = 0
            try: 
                operator_2lb = list(filter(lambda x: x['UF_CRM_1638093750742'] == '439', edo_companies_count_last_month))
            except ZeroDivisionError:
                operator_2lb = 0
            try: 
                operator_2bk = list(filter(lambda x: x['UF_CRM_1638093750742'] == '1357', edo_companies_count_last_month))
            except ZeroDivisionError:
                operator_2bk = 0
            try: 
                operator_2lt = list(filter(lambda x: x['UF_CRM_1638093750742'] == '1831', edo_companies_count_last_month))
            except ZeroDivisionError:
                operator_2lt = 0
        else:
            edo_companies_count_last_month = []

        all_its = its_prof_deals_last_month + its_base_deals_last_month
        name_month = [month_codes[month_int_names[before_1_month]]]

        if before_1_month not in [2, 5, 8, 11]:
            name_month.append(month_codes[month_int_names[before_2_month]])
            all_its_start_quarter = its_prof_deals_quarter + its_base_deals_quarter
            for deals_1 in all_its_start_quarter:
                if deals_1 not in all_its:
                    all_its.append(deals_1)      
            if before_1_month not in [1, 4, 7, 10]:
                name_month.append(month_codes[month_int_names[before_3_month]])
                all_its_deals_middle_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                                x['Группа'] == 'ИТС' and
                                                'ГРМ' not in x['Тип'] and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                before_2_month_deals_data))
                for deals_2 in all_its_deals_middle_month:
                    if deals_2 not in all_its: 
                        all_its.append(deals_2)
        edo_companies_id = list(map(lambda x: x['Компания'], list(filter(lambda y: 'Компания' in y and y['Компания'], all_its))))

        if edo_companies_id:
            edo_elements_info = b.get_all('lists.element.get', {
                'IBLOCK_TYPE_ID': 'lists',
                'IBLOCK_ID': '235',
                'filter': {
                    'PROPERTY_1579': edo_companies_id,
                    'PROPERTY_1567': name_month,
                    'PROPERTY_1569': year_codes[str(before_1_month_year)],
                }
            })
            edo_elements_info = list(map(lambda x: {
                'ID': x['ID'],
                'Компания': list(x['PROPERTY_1579'].values())[0],
                'Сумма пакетов по владельцу': int(list(x['PROPERTY_1573'].values())[0]),
                'Сумма для клиента': int(list(x['PROPERTY_1575'].values())[0]),
            }, edo_elements_info))
            
            traffic_more_than_1 = list(filter(lambda x: x['Сумма пакетов по владельцу'] > 1, edo_elements_info))
            traffic_more_than_1 = len(set(map(lambda x: x['Компания'], traffic_more_than_1)))

            edo_elements_paid = b.get_all('lists.element.get', {
                'IBLOCK_TYPE_ID': 'lists',
                'IBLOCK_ID': '235',
                'filter': {
                    'PROPERTY_1581': user_info['ID'],
                    'PROPERTY_1567': name_month,
                    'PROPERTY_1569': year_codes[str(before_1_month_year)],
                }
            })
            paid_traffic = list(filter(lambda x: int(list(x['PROPERTY_1573'].values())[0]) > 0 and int(list(x['PROPERTY_1575'].values())[0]) > 0, edo_elements_paid))
            paid_traffic = sum(list(map(lambda x: int(list(x['PROPERTY_1575'].values())[0]), paid_traffic)))

        else:
            traffic_more_than_1 = 0
            paid_traffic = 0
        try:
            edo_companies_coverage = round((len(edo_companies_count_last_month) / len(all_its_last_month)) * 100, 2)
        except ZeroDivisionError:
            edo_companies_coverage = 0

        try:
            active_its_coverage = round((traffic_more_than_1 / len(all_its_last_month)) * 100, 2)
        except ZeroDivisionError:
            active_its_coverage = 0

        try:
            operator_2ae_coverage = round((len(operator_2ae) / len(edo_companies_count_last_month)) * 100, 2)
        except ZeroDivisionError:
            operator_2ae_coverage = 0
        try:
            operator_2ae_doki_coverage = round((len(operator_2ae_doki) / len(edo_companies_count_last_month)) * 100, 2)
        except ZeroDivisionError:
            operator_2ae_doki_coverage = 0
        try:
            operator_2be_coverage = round((len(operator_2be) / len(edo_companies_count_last_month)) * 100, 2)
        except ZeroDivisionError:
            operator_2be_coverage = 0
        try:
            operator_2bm_coverage = round((len(operator_2bm) / len(edo_companies_count_last_month)) * 100, 2)
        except ZeroDivisionError:
            operator_2bm_coverage = 0
        try:
            operator_2al_coverage = round((len(operator_2al) / len(edo_companies_count_last_month)) * 100, 2)
        except ZeroDivisionError:
            operator_2al_coverage = 0
        try:
            operator_2lb_coverage = round((len(operator_2lb) / len(edo_companies_count_last_month)) * 100, 2)
        except ZeroDivisionError:
            operator_2lb_coverage = 0
        try:
            operator_2bk_coverage = round((len(operator_2bk) / len(edo_companies_count_last_month)) * 100, 2)
        except ZeroDivisionError:
            operator_2bk_coverage = 0
        try:
            operator_2lt_coverage = round((len(operator_2lt) / len(edo_companies_count_last_month)) * 100, 2)
        except ZeroDivisionError:
            operator_2lt_coverage = 0

        worksheet.append(['ЭДО', 'Всего ИТС на конец квартала', 'С ЭДО', '%'])
        worksheet.append(['Охват ЭДО', len(all_its_last_month), len(edo_companies_count_last_month), edo_companies_coverage])
        if len(edo_companies_count_last_month) > 0:
            if len(operator_2ae) > 0:
                worksheet.append(['', '2AE', len(operator_2ae), operator_2ae_coverage])
            if len(operator_2ae_doki) > 0:
                worksheet.append(['', '2AE доки', len(operator_2ae_doki), operator_2ae_doki_coverage])
            if len(operator_2be) > 0:
                worksheet.append(['', '2BE', len(operator_2be), operator_2be_coverage])
            if len(operator_2bm) > 0:
                worksheet.append(['', '2BM', len(operator_2bm), operator_2bm_coverage])
            if len(operator_2al) > 0:
                worksheet.append(['', '2AL', len(operator_2al), operator_2al_coverage])
            if len(operator_2lb) > 0:
                worksheet.append(['', '2LB', len(operator_2lb), operator_2lb_coverage])
            if len(operator_2bk) > 0:
                worksheet.append(['', '2BK', len(operator_2bk), operator_2bk_coverage])
            if len(operator_2lt) > 0:
                worksheet.append(['', '2LT', len(operator_2lt), operator_2lt_coverage])
        worksheet.append(['Компании с трафиком больше 1', traffic_more_than_1])
        worksheet.append(['% активных ИТС', active_its_coverage])
        worksheet.append(['Сумма платного трафика', paid_traffic])
        worksheet.append([])

        #Сверка 2.0
        all_company = b.get_all('crm.company.list', {
                'select': ['UF_CRM_1735194029'],
                'filter': {
                    'ASSIGNED_BY_ID': user_info['ID'],
                }
            })
        company_sverka = list(filter(lambda x: x['UF_CRM_1735194029'] == '1', all_company))
        
        worksheet.append(['Сверка 2.0'])
        worksheet.append(['Подключено', len(company_sverka)])

        change_sheet_style(worksheet)
       
    workbook.save(report_name)

    if 'user_id' not in req:
        return

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '828809'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    b.call('im.notify.system.add', {
        'USER_ID': req['user_id'][5:],
        'MESSAGE': f'Отчет по пользователям за квартал сформирован. {upload_report["DETAIL_URL"]}'})
    os.remove(report_name) 


if __name__ == '__main__':
    create_employees_quarter_report({
        'users': 'user_1391'
    })
