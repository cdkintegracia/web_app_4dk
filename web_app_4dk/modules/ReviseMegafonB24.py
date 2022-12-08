from fast_bitrix24 import Bitrix
import openpyxl
from datetime import datetime

b = Bitrix('https://vc4dk.bitrix24.ru/rest/311/wkq0a0mvsvfmoseo/')


workbook = openpyxl.load_workbook('Отчет_Мегафон_07-12-2022_17_14_40_685685.xlsx')
worklist = workbook.active
max_rows = worklist.max_row
max_columns = worklist.max_column
megafon_data = []
titles = {}
reverse_titles = {}
for row in range(1, max_rows + 1):
    temp = {}
    for column in range(1, max_columns + 1):
        if row == 1:
            titles.setdefault(worklist.cell(row, column).value, column)
            reverse_titles.setdefault(column, worklist.cell(row, column).value)
        else:
            temp.setdefault(reverse_titles[column], worklist.cell(row, column).value)
    if temp:
        megafon_data.append(temp)

b24_elements = b.get_all('lists.element.get', {
    'IBLOCK_TYPE_ID': 'lists',
    'IBLOCK_ID': '175',
    'filter': {'NAME': 'Ноябрь 2022'}
     }
                         )
companies = b.get_all('crm.company.list')
for element in b24_elements:
    company_id = list(element['PROPERTY_1299'].values())[0]
    company_name = list(filter(lambda x: x['ID'] == company_id, companies))[0]['TITLE']
    duration = list(element['PROPERTY_1303'].values())[0]
    for i in range(len(megafon_data)):
        if megafon_data[i]['Компания'] == company_name:
            megafon_data[i].setdefault('Длительность Б24', duration)

workbook = openpyxl.Workbook()
worklist = workbook.active
worklist.append(['Компания', 'Длительность Мегафон', 'Длительность Б24', 'Разница'])
for i in megafon_data:
    if 'Длительность Б24' not in i:
        print(i)
        continue
    duration_megafon = datetime.strptime(i['ЛК минут'], '%H:%M:%S')
    duration_b24 = datetime.strptime(i['Длительность Б24'], '%H:%M:%S')
    difference = duration_megafon - duration_b24
    data = [i['Компания'], i['ЛК минут'], i['Длительность Б24'], difference]
    worklist.append(data)
workbook.save('Сверка звонков.xlsx')


