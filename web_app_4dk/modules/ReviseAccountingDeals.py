from fast_bitrix24 import Bitrix
import openpyxl
import os
import base64
from datetime import datetime

from web_app_4dk.modules.authentication import authentication


b = Bitrix(authentication('Bitrix'))


def read_report_file(filename: str) -> dict:
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    max_rows = worksheet.max_row
    max_cols = worksheet.max_column
    file_data = []
    titles = []
    for row in range(1, max_rows + 1):
        temp = {}
        for column in range(1, max_cols + 1):
            if row == 1:
                titles.append(str(worksheet.cell(row=row, column=column).value))
            else:
                temp.setdefault(titles[column - 1], worksheet.cell(row=row, column=column).value)
        if temp and temp['Наименование тарифа'] not in [
            'Информационно-технологическое сопровождение 1С',
            'Обслуживание учетной записи 1С в рамках ранее заключенного договора',
            'Промо (1 месяц бесплатно)',
            'Промо Кадровые решения (ФСС и ПФР)',
        ]:
            file_data.append(temp)
    titles.append('Расхождение')
    titles.append('Цена из Битрикса')
    titles.append('Цена')
    return {'titles': titles, 'data': file_data}


def revise_accounting_deals(filename, b24_user_id):
    what_remove = [
        'ИНН партнёра',
        'КПП партнёра',
        'Наименование партнёра',
        'КПП абонента',
        'Дата регистрации',
        'Регион',
        'Тарифная зона'
    ]
    job_counter = 0
    file_data = read_report_file(filename)
    for file_line in file_data['data']:
        registration_data_list = file_line['Дата регистрации'].split('.')
        registration_data = f"{registration_data_list[2]}-{registration_data_list[1]}-{registration_data_list[0]}"
        company_info = b.get_all('crm.company.list', {
            'filter': {'UF_CRM_1656070716': file_line['ИНН абонента'],
                       'UF_CRM_1654682057': registration_data
                       }})
        if not company_info:
            file_line.setdefault('Расхождение', 'Нет ИНН')
            file_line.pop('Цена')
            continue
        company_id = company_info[0]['ID']
        company_deals = b.get_all('crm.deal.list', {'filter': {'TYPE_ID': 'UC_O99QUW', 'COMPANY_ID': company_id}})
        if not company_deals:
            file_line.setdefault('Расхождение', 'Нет отчетности')
            file_line.pop('Цена')
            continue
        company_deal = company_deals[0]
        if int(float(company_deal['OPPORTUNITY'])) < int(float(file_line['Цена'])):
            file_line.setdefault('Расхождение', 'Некорректная сумма')
        else:
            file_line.setdefault('Расхождение', 'Нет')
        price = file_line.pop('Цена')
        file_line.setdefault('Сумма из Битрикса', int(float(company_deal['OPPORTUNITY'])))
        file_line.setdefault('Цена', price)
        job_counter += 1
        print(f"{job_counter} | {len(file_data['data'])}")

    ind = 0
    for i in range(len(file_data['titles'])):
        if file_data['titles'][i] == 'Цена':
            ind = i
            break
    del file_data['titles'][ind]
    for title in what_remove:
        file_data['titles'].remove(title)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(file_data['titles'])
    for data in file_data['data']:
        for value in what_remove:
            data.pop(value)
        worksheet.append(list(data.values()))
    create_time = datetime.now().strftime('%d-%m-%Y-%f')
    report_name = f'Сверка_отчетности_{create_time}.xlsx'
    workbook.save(report_name)

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '193689'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    b.call('tasks.task.add', {
        'fields': {
            'TITLE': 'Сверка отчетности',
            'RESPONSIBLE_ID': b24_user_id,
            'DESCRIPTION': upload_report["DETAIL_URL"],
            'CREATED_BY': '173'
        }})
    os.remove(f'Сверка_отчетности_{create_time}.xlsx')











