from time import time

from fast_bitrix24 import Bitrix
import openpyxl

from web_app_4dk.modules.authentication import authentication


b = Bitrix(authentication('Bitrix'))


def read_file(filename: str) -> list:
    workbook = openpyxl.load_workbook(filename)
    worklist = workbook.active
    max_rows = worklist.max_row
    max_columns = worklist.max_column
    titles = {}
    file_data = []
    for row in range(1, max_rows + 1):
        temp = {}
        for column in range(1, max_columns + 1):
            cell_value = worklist.cell(row, column).value
            if row == 1:
                titles.setdefault(column, cell_value)
            else:
                temp.setdefault(titles[column], cell_value)
        if temp:
            file_data.append(temp)
    return file_data


def get_questionnaire_answers(registrant_questionnaire: dict) -> dict:
    ignore_answer_titles = [
        'Дата создания',
        'Дата обновления',
        'ФИО',
        'E-mail',
        'Телефон',
        'Название организации',
        'Вид деятельности',
        'Должность',
        'ИНН организации',
    ]
    answers_data = []
    company_inn = ''
    for title in registrant_questionnaire:
        if title == 'ИНН организации':
            company_inn = str(registrant_questionnaire[title])
        if title in ignore_answer_titles or 'Оценка' in title:
            continue
        answers_data.append(f"{title}: {registrant_questionnaire[title]}\n")
    return {'Ответы': answers_data, 'ИНН': company_inn}


def phone_handler(phone) -> str:
    phone = str(phone)
    if len(phone) == 10 and (phone[0] == '9' or phone[0] == '8'):
        phone = '+7' + phone
    elif len(phone) == 11 and phone[0] == '7':
        phone = '+' + phone
    elif len(phone) == 11 and phone[0] != '7':
        phone = '+7' + phone[1:]
    return phone


def contact_in_b24(registrant, contacts, phone, email):
    registrant_phone = phone_handler(registrant['Телефон'])
    for registrant_name_word in registrant['ФИО'].split(' '):
        registrant_name_word = registrant_name_word.lower().capitalize()
        for contact in contacts:
            if registrant_name_word != contact['LAST_NAME']:
                continue
            if 'EMAIL' in contact:
                emails = list(map(lambda x: x['VALUE'], contact['EMAIL']))
                if registrant['E-mail'] in emails:
                    return contact['ID']
            if 'PHONE' in contact:
                phones = list(map(lambda x: x['VALUE'], contact['PHONE']))
                if registrant_phone in phones:
                    return contact['ID']

    for contact in contacts:
        if 'PHONE' in contact:
            phones = list(map(lambda x: x['VALUE'], contact['PHONE']))
            if registrant_phone in phones:
                return contact['ID']

    if phone:
        contacts_phones = list(filter(lambda x: 'PHONE' in x, contacts))
        contacts_phones = list(map(lambda x: {'ID': x['ID'], 'PHONE': list(map(lambda y: y['VALUE'], x['PHONE']))}, contacts_phones))
        is_phone_in_contacts = list(filter(lambda x: phone in x['PHONE'], contacts_phones))
        if len(is_phone_in_contacts) == 1:
            return is_phone_in_contacts[0]['ID']
    if email:
        contacts_emails = list(filter(lambda x: 'EMAIL' in x, contacts))
        contacts_emails = list(map(lambda x: {'ID': x['ID'], 'EMAIL': list(map(lambda y: y['VALUE'], x['EMAIL']))}, contacts_emails))
        is_email_in_contacts = list(filter(lambda x: email in x['EMAIL'], contacts_emails))
        if len(is_email_in_contacts) == 1:
            return is_email_in_contacts[0]['ID']


def company_in_b24(companies, company_inn=None, company_name=None, phone=None, email=None):
    if company_inn:
        company_info = list(filter(lambda x: x['UF_CRM_1656070716'] == company_inn, companies))
        if company_info:
            return company_info[0]['ID']
    elif company_name:
        replace_elements = ['ООО', '"', "'"]
        for replace_element in replace_elements:
            company_name = str(company_name).replace(replace_element, '')
        company_name = company_name.strip()
        companies_names = list(map(lambda x: {'ID': x['ID'], 'TITLE': x['TITLE']}, companies))
        company_info = list(filter(lambda x: company_name.upper() in x['TITLE'], companies_names))
        if len(company_info) == 1:
            return company_info[0]['ID']
    if phone:
        companies_phones = list(filter(lambda x: 'PHONE' in x, companies))
        companies_phones = list(map(lambda x: {'ID': x['ID'], 'PHONE': list(map(lambda y: y['VALUE'], x['PHONE']))}, companies_phones))
        is_phone_in_companies = list(filter(lambda x: phone in x['PHONE'], companies_phones))
        if len(is_phone_in_companies) == 1:
            return is_phone_in_companies[0]['ID']
    if email:
        companies_emails = list(filter(lambda x: 'EMAIL' in x, companies))
        companies_emails = list(map(lambda x: {'ID': x['ID'], 'EMAIL': list(map(lambda y: y['VALUE'], x['EMAIL']))}, companies_emails))
        is_email_in_companies = list(filter(lambda x: email in x['EMAIL'], companies_emails))
        if len(is_email_in_companies) == 1:
            return is_email_in_companies[0]['ID']


def is_company_type_client(company_id, companies):
    company_info = list(filter(lambda x: str(x['ID']) == str(company_id), companies))[0]
    if company_info['COMPANY_TYPE'] == 'CUSTOMER':
        return '311'


def get_company_contact_by_surname(company_id, fio):
    surname = fio.split(' ')[0].lower().capitalize()
    company_contacts = b.get_all('crm.company.contact.items.get', {'id': company_id})
    company_contacts = list(map(lambda x: str(x['CONTACT_ID']), company_contacts))
    if company_contacts:
        contacts = b.get_all('crm.contact.list', {'filter': {'ID': company_contacts}})
        for contact in contacts:
            if contact['LAST_NAME'] in surname:
                return contact['ID']


def seminar_data_handler(event_id, registrants_file, questionnaire_file):
    registrants_data = read_file(registrants_file)
    questionnaire_data = read_file(questionnaire_file)
    companies = b.get_all('crm.company.list', {'select': ['*', 'UF_*', 'EMAIL', 'PHONE']})
    contacts = b.get_all('crm.contact.list', {'select': ['*', 'UF_*', 'EMAIL', 'PHONE']})
    for registrant in registrants_data:
        registrant.setdefault('Ответы', '')
        registrant.setdefault('Компания ID', '')
        registrant.setdefault('Контакт ID', '')
        registrant.setdefault('Анкета заполнена', '307')
        registrant.setdefault('Участвовал', '2175')
        registrant.setdefault('Потенциальный', '309')
        registrant_questionnaire = list(filter(lambda x: x['ФИО'] == registrant['ФИО'] and x['E-mail'] == registrant['E-mail'], questionnaire_data))

        if registrant_questionnaire:
            registrant_questionnaire = registrant_questionnaire[0]
            questionnaire_answers = get_questionnaire_answers(registrant_questionnaire)
            registrant['Ответы'] = questionnaire_answers['Ответы']
            b24_company_id = company_in_b24(companies, company_inn=questionnaire_answers['ИНН'])
            registrant['Компания ID'] = b24_company_id

        b24_contact_id = contact_in_b24(registrant, contacts, registrant['Телефон'], registrant['E-mail'])
        registrant['Контакт ID'] = b24_contact_id

        if registrant['Ответы']:
            registrant['Анкета заполнена'] = '305'
            registrant['Участвовал'] = '2173'

        if not registrant['Компания ID']:
            b24_company_id = company_in_b24(companies, company_name=registrant['Название организации'], phone=phone_handler(registrant['Телефон']), email=registrant['E-mail'])
            registrant['Компания ID'] = b24_company_id

        if registrant['Компания ID'] and not registrant['Контакт ID']:
            contact = get_company_contact_by_surname(registrant['Компания ID'], registrant['ФИО'])
            if contact:
                registrant['Контакт ID'] = contact

        if registrant['Контакт ID'] and not registrant['Компания ID']:
            contact_companies = b.get_all('crm.contact.company.items.get', {'id': registrant['Контакт ID']})
            contact_companies = list(map(lambda x: x['COMPANY_ID'], contact_companies))
            contact_companies_info = b.get_all('crm.company.list', {'filter': {'ID': contact_companies}})
            client_companies = list(filter(lambda x: x['COMPANY_TYPE'] == 'CUSTOMER', contact_companies_info))
            if client_companies:
                registrant['Компания ID'] = client_companies[0]['ID']
            elif contact_companies:
                registrant['Компания ID'] = str(contact_companies[0])

        if registrant['Компания ID']:
            potential_client = is_company_type_client(registrant['Компания ID'], companies)
            if potential_client:
                registrant['Потенциальный'] = potential_client


        b.call('lists.element.add', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '49',
            'ELEMENT_CODE': time(),
            'FIELDS': {
                'NAME': registrant['ФИО'],
                'PROPERTY_399': event_id,
                'CREATED_BY': '173',
                'PROPERTY_401': registrant['Контакт ID'],
                'PROPERTY_403': registrant['Компания ID'],
                'PROPERTY_407': registrant['Анкета заполнена'],
                'PROPERTY_409': registrant['Потенциальный'],
                'PROPERTY_1223': registrant['Участвовал'],
                'PROPERTY_1513': registrant['Телефон'],
                'PROPERTY_1515': registrant['Название организации'],
                'PROPERTY_1521': registrant['Ответы'],
                'PROPERTY_1533': registrant['E-mail']
            }})
