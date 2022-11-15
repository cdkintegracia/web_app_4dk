from fast_bitrix24 import Bitrix
import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
import dateutil.parser
import base64
import os

from web_app_4dk.modules.authentication import authentication


b = Bitrix(authentication('Bitrix'))


deal_type_names = {
        'SALE': 'ИТС Земля',
        'COMPLEX': 'СааС',
        'UC_APUWEW': 'ИТС Земля + СааС',
        'UC_1UPOTU': 'ИТС Бесплатный',
        'UC_O99QUW': 'Отчетность',
        'UC_OV4T7K': 'Отчетность (в рамках ИТС)',
        'UC_2B0CK2': '1Спарк в договоре',
        'UC_86JXH1': '1Спарк 3000',
        'UC_WUGAZ7': '1СпаркПЛЮС 22500',
        'UC_A7G0AM': '1С Контрагент',
        'GOODS': 'ГРМ',
        'UC_GZFC63': 'РПД',
        'UC_QQPYF0': 'Лицензия',
        'UC_8Z4N1O': '1С-Подпись',
        'UC_FOKY52': '1С-Подпись 1000',
        'UC_D1DN7U': '1С Кабинет сотрудника',
        'UC_34QFP9': 'Уникс',
        'UC_J426ZW': '1С Садовод',
        'UC_H8S037': 'ЭДО',
        'UC_8LW09Y': 'МДЛП',
        'UC_3SKJ5M': '1С Касса',
        'UC_4B5UQD': 'ЭТП',
        'UC_H7HOD0': 'Коннект',
        'UC_USDKKM': 'Медицина',
        'SERVICE': 'Сервисное обслуживание',
        'SERVICES': 'Услуги',
        'UC_XJFZN4': 'Кабинет садовода',
        'UC_BZYY0D': 'ИТС Отраслевой',
        'UC_66Z1ZF': 'ОФД',
        'UC_40Q6MC': 'Старт ЭДО',
        'UC_74DPBQ': 'Битрикс24',
        'UC_IV3HX1': 'Тестовый',
        'UC_HT9G9H': 'ПРОФ Земля',
        'UC_XIYCTV': 'ПРОФ Земля+Помощник',
        'UC_5T4MAW': 'ПРОФ Земля+Облако+Помощник',
        'UC_N113M9': 'ПРОФ Земля+Облако',
        'UC_ZKPT1B': 'ПРОФ Облако',
        'UC_2SJOEJ': 'ПРОФ Облако+Помощник',
        'UC_AVBW73': 'Базовый Земля',
        'UC_GPT391': 'Базовый Облако',
        'UC_92H9MN': 'Индивидуальный',
        'UC_7V8HWF': 'Индивидуальный+Облако',
        'UC_IUJR81': 'Допы Облако',
        'UC_2R01AE': 'Услуги (без нашего ИТС)',
        'UC_81T8ZR': 'АОВ',
        'UC_SV60SP': 'АОВ+Облако',
        'UC_D7TC4I': 'ГРМ Спец',
        'UC_K9QJDV': 'ГРМ Бизнес',
        'UC_DBLSP5': 'Садовод+Помощник',
        'UC_GP5FR3': 'ДУО',
        'UC_YIAJC8': 'Лицензия с купоном ИТС',
        '1': 'Не указан',
    }
deal_stage_names = {
        'C1:NEW': 'Услуга активна',
        'C1:LOSE': 'Отказ от сопровождения',
        'C1:UC_0KJKTY': 'Счет сформирован',
        'C1:UC_3J0IH6': 'Счет отправлен клиенту',
        'C1:WON': 'Услуга завершена',
        'C1:UC_KZSOR2': 'Нет оплаты',
        'C1:UC_VQ5HJD': 'Ждем решения клиента',
    }
extra_1c_codes = ['161', '162', '2001', '2002', '2003']

def revise_b24_deals(file_name, file_data, titles, companies):
    all_deals = b.get_all('crm.deal.list', {
        'select': [
            'UF_CRM_1640523562691',  # Регномер
            'UF_CRM_1655972832',  # СлужКод1С
            'COMPANY_ID',
            'TITLE',
            'CLOSEDATE',
            'TYPE_ID',
            'STAGE_ID'
        ],
        'filter': {
            'CATEGORY_ID': '1',
            'TYPE_ID': [
                    'UC_HT9G9H',    # ПРОФ Земля
                    'UC_XIYCTV',    # ПРОФ Земля+Помощник
                    'UC_N113M9',    # ПРОФ Земля+Облако
                    'UC_5T4MAW',    # ПРОФ Земля+Облако+Помощник
                    'UC_ZKPT1B',    # ПРОФ Облако
                    'UC_2SJOEJ',    # ПРОФ Облако+Помощник
                    'UC_81T8ZR',    # АОВ
                    'UC_SV60SP',    # АОВ+Облако
                    'UC_92H9MN',    # Индивидуальный
                    'UC_7V8HWF',    # Индивидуальный+Облако
                    'UC_AVBW73',    # Базовый Земля
                    'UC_GPT391',    # Базовый Облако
                    'UC_1UPOTU',    # ИТС Бесплатный
                    'UC_K9QJDV',    # ГРМ Бизнес
                    'GOODS',        # ГРМ
                    'UC_DBLSP5',    # Садовод+Помощник
                    'UC_J426ZW',    # 1С Садовод
                    'UC_2B0CK2',    # 1Спарк в договоре
                    'UC_86JXH1',    # 1Спарк 3000
                    'UC_WUGAZ7',    # 1СпаркПЛЮС 22500
                    'UC_A7G0AM',    # 1С Контрагент
                    'UC_GZFC63',    # РПД
                    'UC_40Q6MC',    # Старт ЭДО
                    'UC_XJFZN4',    # Кабинет садовода
                    'UC_USDKKM',    # Медицина
                    'UC_D1DN7U',    # 1С Кабинет сотрудника
                    'UC_H8S037',    # ЭДО
            ],
            'STAGE_ID': [
                'C1:NEW',           # Услуга активна
                'C1:UC_0KJKTY',     # Счет сформирован
                'C1:UC_3J0IH6',     # Счет отправлен клиенту
                ]
        }})
    report_name = f'Обратная сверка по {file_name}'
    report_data = [
        [
            report_name
        ],
        [
            'Регномер в Б24',
            'Код в Б24',
            'Название в Б24',
            'Тип в Б24',
            'Стадия в Б24',
            'Компания в Б24',
            'Найдено в 1С',
            'Компания в 1С',
            'Название сделки в 1С',
            'Дата завершения в Б24',
            'Дата завершения в 1С',
            'Код в 1С'
        ]
    ]
    for deal in all_deals:
        reg_number_b24 = deal['UF_CRM_1640523562691']
        code_1c_b24 = deal['UF_CRM_1655972832']
        deal_name_b24 = deal['TITLE']
        deal_type_b24 = deal_type_names[deal['TYPE_ID']]
        deal_stage_b24 = deal_stage_names[deal['STAGE_ID']]
        close_date_b24 = dateutil.parser.isoparse(deal['CLOSEDATE'])
        close_date_b24 = datetime.strftime(close_date_b24, "%d.%m.%Y")
        company_name_b24 = ''
        extra_code = ''
        company_info = list(filter(lambda x: x['ID'] == deal['COMPANY_ID'], companies))[0]
        if company_info:
            company_name_b24 = company_info['TITLE']
        data_was_found = 'Нет'
        company_name_1c = ''
        deal_name_1c = ''
        close_date_1c = ''
        data = list(filter(lambda x: str(reg_number_b24).strip(' ') == str(x[titles['Регномер']]).strip(' ') and str(code_1c_b24).strip(' ') == str(x[titles['Вид 1С:ИТС']]).strip(' '), file_data))

        # Поиск сделок по дополнительным кодам
        if not data and str(code_1c_b24).strip(' ') in extra_1c_codes:
            for extra_1c_code in extra_1c_codes:
                data = list(filter(lambda x: str(reg_number_b24).strip(' ') == str(x[titles['Регномер']]).strip(' ') and str(x[titles['Вид 1С:ИТС']]).strip(' ') == extra_1c_code.strip(' '), file_data))
                if data:
                    extra_code = extra_1c_code
                    break

        if data:
            for line in data:
                company_name_1c = line[titles['Пользователь']]
                deal_name_1c = line[titles['Наименование вида 1С:ИТС']]
                close_date_1c = line[titles['Конец']]
                if str(close_date_b24) == str(line[titles['Конец']]):
                    data_was_found = 'Да'
                    break
                else:
                    data_was_found = 'Другая дата завершения'

        if data_was_found == 'Нет':
            data = list(filter(lambda x: str(reg_number_b24).strip(' ') == str(x[titles['Регномер']]).strip(' '), file_data))
            if not data:
                data_was_found = 'Нет регномера'

        report_data.append([
            reg_number_b24,
            code_1c_b24,
            deal_name_b24,
            deal_type_b24,
            deal_stage_b24,
            company_name_b24,
            data_was_found,
            company_name_1c,
            deal_name_1c,
            close_date_b24,
            close_date_1c,
            extra_code
        ])

    return report_data


def revise_new_sub(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
        max_rows = worksheet.max_row
        max_columns = worksheet.max_column
    except:
        b.call('tasks.task.add', {
            'fields': {
                'TITLE': 'Ошибка: Сверка по NewSub',
                'RESPONSIBLE_ID': '173',
                'GROUP_ID': '13',
                'DESCRIPTION': 'Не удалось открыть загруженный файл NewSub. Может помочь его пересохранение в тот же формат.',
                'CREATED_BY': '173'
            }})
        return
    file_name = worksheet.cell(row=1, column=1).value
    report_name = f'Сверка по {worksheet.cell(row=1, column=1).value}'
    titles = {}
    data = []
    all_deals = b.get_all('crm.deal.list', {
        'select': [
            'UF_CRM_1640523562691',     # Регномер
            'UF_CRM_1655972832',        # СлужКод1С
            'COMPANY_ID',
            'TITLE',
            'CLOSEDATE',
            'TYPE_ID',
            'STAGE_ID'
            ],
        'filter': {
            'CATEGORY_ID': '1'
        }
    })
    companies = b.get_all('crm.company.list', {
        'select': [
            'TITLE',
            'UF_CRM_1656070716',    # СлужИНН
        ]
    })
    for row in range(5, max_rows + 1):
        temp_data = []
        for column in range(1, max_columns + 1):
            if row == 5:
                titles.setdefault(worksheet.cell(row=row, column=column).value, column - 1)
            else:
                temp_data.append(worksheet.cell(row=row, column=column).value)
        if temp_data:
            data.append(temp_data)
    report_data = [
        [
            report_name
        ],
        [
            'Регномер в 1С',
            'Код в 1С',
            'Название в 1С',
            'Компания в 1С',
            'Найдено в Б24',
            'Название сделки в Б24',
            'Компания в Б24',
            'Тип в Б24',
            'Дата завершения в 1С',
            'Дата завершения в Б24',
            'Стадия в Б24',
            'Код в Б24'
        ]
    ]

    for line in data:
        reg_number_1c = line[titles['Регномер']]
        code1c_1c = line[titles['Вид 1С:ИТС']]
        deal_name_1c = line[titles['Наименование вида 1С:ИТС']]
        company_name_1c = line[titles['Пользователь']]
        close_date_1c = line[titles['Конец']]
        inn_1c = line[titles['ИНН / ЕДРПОУ(Укр) пользователя']]
        deal_was_found = 'Нет'
        deals = list(filter(lambda x: str(x['UF_CRM_1640523562691']).strip(' ') == str(reg_number_1c).strip(' ') and str(x['UF_CRM_1655972832']).strip(' ') == str(code1c_1c).strip(' '), all_deals))
        close_date_b24 = ''
        deal_name_b24 = ''
        company_name_b24 = ''
        deal_type_b24 = ''
        deal_stage_b24 = ''
        extra_code = ''

        # Поиск сделок по дополнительным кодам
        if not deals and str(code1c_1c) in extra_1c_codes:
            for extra_1c_code in extra_1c_codes:
                deals = list(filter(lambda x: str(x['UF_CRM_1640523562691']).strip(' ') == str(reg_number_1c).strip(' ') and str(x['UF_CRM_1655972832']).strip(' ') == extra_1c_code.strip(' '), all_deals))
                if deals:
                    extra_code = extra_1c_code
                    break
        if deals:
            deals = list(sorted(deals, key=lambda x: dateutil.parser.isoparse(x['CLOSEDATE'])))
            for deal in deals:
                close_date_b24 = dateutil.parser.isoparse(deal['CLOSEDATE'])
                close_date_b24 = datetime.strftime(close_date_b24, "%d.%m.%Y")
                deal_name_b24 = deal['TITLE']
                deal_type_b24 = deal_type_names[deal['TYPE_ID']]
                deal_stage_b24 = deal_stage_names[deal['STAGE_ID']]
                company_name_b24 = list(filter(lambda x: str(x['UF_CRM_1656070716']) == str(inn_1c), companies))
                if company_name_b24:
                    company_name_b24 = company_name_b24[0]['TITLE']
                if not company_name_b24:
                    company_name_b24 = ''
                if str(close_date_b24) == str(close_date_1c):
                    deal_was_found = 'Да'
                    break
                else:
                    deal_was_found = 'Другая дата завершения'
        if deal_was_found == 'Нет':
            deals = list(filter(lambda x: str(x['UF_CRM_1640523562691']).strip(' ') == str(reg_number_1c).strip(' '), all_deals))
            if not deals:
                deal_was_found = 'Нет регномера'
        report_data.append([
            reg_number_1c,
            code1c_1c,
            deal_name_1c,
            company_name_1c,
            deal_was_found,
            deal_name_b24,
            company_name_b24,
            deal_type_b24,
            close_date_1c,
            close_date_b24,
            deal_stage_b24,
            extra_code,
        ])

    # Обратная сверка
    reverse_report = revise_b24_deals(file_name, data, titles, companies)

    # Создание xlsx файла отчета
    report_created_time = datetime.now()
    report_name_time = report_created_time.strftime('%d-%m-%Y %H %M %S %f')
    report_name = f'Сверка по NewSub {report_name_time}.xlsx'.replace(' ', '_')
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Сверка по NewSub'
    for data in report_data:
        worksheet.append(data)
    for idx, col in enumerate(worksheet.columns, 1):
        worksheet.column_dimensions[get_column_letter(idx)].auto_size = True

    worksheet = workbook.create_sheet('Обратная сверка')
    for data in reverse_report:
        worksheet.append(data)
    for idx, col in enumerate(worksheet.columns, 1):
        worksheet.column_dimensions[get_column_letter(idx)].auto_size = True
    workbook.save(report_name)

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '214239'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    task = b.call('tasks.task.add', {
        'fields': {
            'TITLE': 'Сверка по NewSub',
            'RESPONSIBLE_ID': '173',
            'GROUP_ID': '13',
            'DESCRIPTION': upload_report["DETAIL_URL"],
            'CREATED_BY': '173'
        }})
    os.remove(report_name)

