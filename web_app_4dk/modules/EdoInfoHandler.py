from time import time, sleep
import os
import base64
from datetime import datetime
import asyncio

import openpyxl
from fast_bitrix24 import BitrixAsync

from web_app_4dk.modules.authentication import authentication


def run_async(coro):
    """
    Безопасно запускает корутину из sync-кода (WSGI/Flask).
    Создаёт новый loop, делает его current для текущего thread, выполняет и закрывает.
    """
    loop = asyncio.new_event_loop()
    try:
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(coro)
    finally:
        try:
            loop.run_until_complete(loop.shutdown_asyncgens())
        except Exception:
            pass
        loop.close()
        asyncio.set_event_loop(None)


month_codes = {
    'Январь': '2371',
    'Февраль': '2373',
    'Март': '2375',
    'Апрель': '2377',
    'Май': '2379',
    'Июнь': '2381',
    'Июль': '2383',
    'Август': '2385',
    'Сентябрь': '2387',
    'Октябрь': '2389',
    'Ноябрь': '2391',
    'Декабрь': '2393'
}

year_codes = {
    '2022': '2395',
    '2023': '2397',
    '2024': '2741',
    '2025': '2743',
    '2026': '2745',
    '2027': '2747',
    '2028': '2749',
    '2029': '2781',
    '2030': '2783'
}


def read_xlsx(filename: str) -> list:
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    max_rows = worksheet.max_row
    max_columns = worksheet.max_column
    data = []
    titles = {}
    ignore_client_values = [None, '0', 'None']
    ignore_package_count = [None, '0', 'None']

    for row in range(1, max_rows + 1):
        temp = {}
        for column in range(1, max_columns + 1):
            cell_value = str(worksheet.cell(row, column).value)
            if row == 1:
                titles.setdefault(column, cell_value)
            else:
                if not titles:
                    continue
                temp.setdefault(titles[column], cell_value)

        if temp:
            if (
                temp.get('Сумма для клиента') in ignore_client_values
                and temp.get('Сумма пакетов по владельцу') in ignore_package_count
            ):
                continue
            data.append(temp)

    return data


async def get_service_list_elements(b_async: BitrixAsync) -> list:
    elements = await b_async.get_all(
        'lists.element.get',
        {'IBLOCK_TYPE_ID': 'lists', 'IBLOCK_ID': '169'}
    )
    return list(map(lambda x: {
        'COMPANY_ID': list(x['PROPERTY_1291'].values())[0],
        'SUBSCRIBER_CODE': list(x['PROPERTY_1289'].values())[0]
    }, elements))


async def create_edo_list_element(b_async: BitrixAsync, month: str, year: str, data: dict):
    # если ты реально хочешь паузу между запросами — оставляю.
    # но лучше заменить на asyncio.sleep(1), чтобы не блокировать loop.
    await asyncio.sleep(1)

    await b_async.call('lists.element.add', {
        'IBLOCK_TYPE_ID': 'lists',
        'IBLOCK_ID': '235',
        'ELEMENT_CODE': str(time()),
        'FIELDS': {
            'NAME': f"{month} {year}",
            'PROPERTY_1567': month_codes[month],
            'PROPERTY_1569': year_codes[year],
            'PROPERTY_1571': data['Компания'],
            'PROPERTY_1573': data['Сумма пакетов по владельцу'] if data.get('Сумма пакетов по владельцу') != 'None' else '0',
            'PROPERTY_1575': data['Сумма для клиента'] if data.get('Сумма для клиента') != 'None' else '0',
            'PROPERTY_1577': data['Регномер'] if 'Регномер' in data else '',
            'PROPERTY_1579': data['Владелец ИТС'] if 'Владелец ИТС' in data else '',
            'PROPERTY_1581': data['Ответственный за ИТС'] if 'Ответственный за ИТС' in data else ''
        }
    })


async def upload_file_to_b24(b_async: BitrixAsync, report_name: str) -> str:
    bitrix_folder_id = '268033'
    with open(report_name, 'rb') as file:
        report_file = file.read()

    report_file_base64 = str(base64.b64encode(report_file))[2:]

    upload_report = await b_async.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })

    return upload_report["DETAIL_URL"]


async def edo_info_handler_async(month: str, year: str, filename: str, b24_user_id):
    b_async = BitrixAsync(authentication('Bitrix'))

    not_found = []
    file_data = read_xlsx(filename)

    service_list_elements = await get_service_list_elements(b_async)

    companies = await b_async.get_all('crm.company.list', {'select': ['*', 'UF_*']})
    deals = await b_async.get_all('crm.deal.list', {'select': ['*', 'UF_*'], 'filter': {'CATEGORY_ID': '1'}})

    for data in file_data:
        # Поиск компании по ИНН
        company = list(filter(lambda x: x.get('UF_CRM_1656070716') == data.get('ИНН клиента'), companies))
        company_link_company = None

        if company:
            company = company[0]
            data.setdefault('Компания', company['ID'])
            if company.get('UF_CRM_1662385947'):
                company_link_company = company['UF_CRM_1662385947'][0]
        else:
            # Поиск компании по коду подписчика в УС "Отчет по сервисам"
            subscriber_code = data.get('Владелец', '').split()[0] if data.get('Владелец') else ''
            elements = list(filter(lambda x: x['SUBSCRIBER_CODE'] == subscriber_code, service_list_elements))
            if elements:
                data.setdefault('Компания', elements[0]['COMPANY_ID'])
            else:
                data.setdefault('Ошибка', 'Не найдена компания')
                not_found.append(data)
                continue

        # Поиск ИТС по компании
        company_deals = list(filter(
            lambda x: x.get('COMPANY_ID') == data.get('Компания') and x.get('TYPE_ID') not in ['UC_QQPYF0', 'UC_YIAJC8'],
            deals
        ))

        for deal in company_deals:
            data['Регномер'] = deal.get('UF_CRM_1640523562691')

            company_its = list(filter(
                lambda x: x.get('UF_CRM_1640523562691') == data.get('Регномер')
                          and x.get('UF_CRM_1657878818384') == '859'
                          and x.get('STAGE_ID') not in ['C1:WON', 'C1:LOSE'],
                deals
            ))

            if company_its:
                data.setdefault('Владелец ИТС', company_its[0].get('COMPANY_ID'))
                data.setdefault('Ответственный за ИТС', company_its[0].get('ASSIGNED_BY_ID'))
                break

        # Поиск ИТС по значению поля компании "Связан с"
        if 'Владелец ИТС' not in data and company_link_company:
            company_its = list(filter(
                lambda x: x.get('COMPANY_ID') == company_link_company and x.get('UF_CRM_1657878818384') == '859',
                deals
            ))
            if company_its:
                data.setdefault('Владелец ИТС', company_its[0].get('COMPANY_ID'))
                data.setdefault('Ответственный за ИТС', company_its[0].get('ASSIGNED_BY_ID'))
                data.setdefault('Регномер', company_its[0].get('UF_CRM_1640523562691'))

        if company and ('Ответственный за ИТС' not in data or not data.get('Ответственный за ИТС')):
            data.setdefault('Ответственный за ИТС', company.get('ASSIGNED_BY_ID'))

        await create_edo_list_element(b_async, month, year, data)

    task_description = f'Данные за {month} {year} файла по ЭДО загружены '

    if not_found:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        not_found_titles = list(not_found[0].keys())
        worksheet.append(not_found_titles)
        for row in not_found:
            worksheet.append(list(row.values()))

        report_created_time = datetime.now()
        report_name_time = report_created_time.strftime('%d-%m-%Y_%H_%M_%S_%f')
        report_name = f"Ошибки_по_ЭДО_за_{month}_{year}_{report_name_time}.xlsx"
        report_name = report_name.replace(' ', '_')

        workbook.save(report_name)
        file_url = await upload_file_to_b24(b_async, report_name)
        os.remove(report_name)

        task_description += f'не полностью\nОтчет по ошибкам: {file_url}'
    else:
        task_description += 'полностью'

    await b_async.call('tasks.task.add', {'fields': {
        'TITLE': "Файл ЭДО обработан",
        'DESCRIPTION': task_description,
        'GROUP_ID': '13',
        'CREATED_BY': '173',
        'RESPONSIBLE_ID': b24_user_id,
    }})


def edo_info_handler(month: str, year: str, filename: str, b24_user_id):
    # Это остаётся синхронной точкой входа, как у тебя было в routes.py
    return run_async(edo_info_handler_async(month, year, filename, b24_user_id))


