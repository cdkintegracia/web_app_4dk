import datetime
from time import time

from fast_bitrix24 import Bitrix
import openpyxl

from web_app_4dk.modules.authentication import authentication
from web_app_4dk.modules.ElementCallStatistic import rewrite_element


b = Bitrix(authentication('Bitrix'))


def rewrite_call_statistic(month, year):
    try:
        workbook = openpyxl.load_workbook("/root/web_app_4dk/web_app_4dk/new_call_statistic.xlsx")
        worksheet = workbook.active
    except:
        return
    month_codes = {
            'Январь': '2215',
            'Февраль': '2217',
            'Март': '2219',
            'Апрель': '2221',
            'Май': '2223',
            'Июнь': '2225',
            'Июль': '2227',
            'Август': '2229',
            'Сентябрь': '2231',
            'Октябрь': '2233',
            'Ноябрь': '2235',
            'Декабрь': '2237'
        }
    year_codes = {
        '2022': '2239',
        '2023': '2241'
    }
    elements = b.get_all('lists.element.get', {'IBLOCK_TYPE_ID': 'lists', 'IBLOCK_ID': '175', 'filter': {'NAME': f"{month} {year}"}})
    companies = b.get_all('crm.company.list', {'select': ['TITLE']})

    errors = []
    data = []
    for row in range(2, worksheet.max_row + 1):
        temp = []
        for col in range(1, 11):
            if col not in [1, 3, 4]:
                continue
            value = worksheet.cell(row, col).value
            if isinstance(value, datetime.time):
                value = value.strftime('%H:%M:%S')
            elif type(value) is str:
                company_name = value.replace(u'\xa0', ' ')
                for company in companies:
                    if company_name in company['TITLE']:
                        value = company['ID']
            if value is None:
                break
            temp.append(value)
        if len(temp) != 3 or temp[2] == 0:
            continue
        if not temp[0].isdigit():
            errors.append(temp)
        else:
            data.append(temp)
    new_errors = []
    for error in errors:
        value = None
        inn = error[0].split()[-1]
        for company in companies:
            if inn in company['TITLE'].split()[-1]:
                value = company['ID']
                break
        if value is None:
            new_errors.append(error)
        else:
            data.append([value, error[1], error[2]])

    for d in data:
        print(d)
        for element in elements:
            company_id = list(element['PROPERTY_1299'].values())[0]
            if d[0] == company_id:
                rewrite_element(element, d[1], d[2])

    if new_errors:
        task_text = 'Компания Длительность звонков Количество звонков\n'
        for count, error in enumerate(errors, start=1):
            str_error = ' '.join(list(map(lambda x: str(x), error)))
            task_text += f"{count}. {str_error}\n"
        b.call('tasks.task.add', {'fields': {
            'TITLE': f"Ошибки записи статистики звонков за {month} {year}",
            'DESCRIPTION': task_text,
            'GROUP_ID': '13',
            'CREATED_BY': '173',
            'RESPONSIBLE_ID': '173'
        }})
    else:
        b.call('tasks.task.add', {'fields': {
            'TITLE': f"ИЗМЕНИТЬ БП (ЗАДАЧИ НА ЛИМИТ). Cтатистика звонков за {month} {year} успешно перезаписана",
            'GROUP_ID': '13',
            'CREATED_BY': '173',
            'RESPONSIBLE_ID': '173'
        }})

