from datetime import datetime, timedelta
import base64
import os

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from fast_bitrix24 import Bitrix
from web_app_4dk.modules.authentication import authentication

b = Bitrix(authentication('Bitrix'))


def normalize_number(value): # удаление пробелов по краям и внутри строки

    if not value:
        return ''
    return str(value).strip().replace(' ', '')


def format_date(date_string): # нормализация даты

    if not date_string:
        return ''

    try:
        return datetime.fromisoformat(date_string).strftime('%d.%m.%Y')

    except:
        return str(date_string)


def get_year(date_string): # извлечение года из даты

    if not date_string:
        return None

    try:
        return datetime.fromisoformat(date_string).year

    except:
        return None


def change_sheet_style(sheet): # изменение ширины столбцов

    for column_cells in sheet.columns:

        length = max(len(str(cell.value) if cell.value else '') for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length * 1.2, 60)


def write_table(sheet, title, headers, rows): # вывод таблицы на страницу

    sheet.append([])
    sheet.append([title])

    sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True)

    sheet.append(headers)

    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(row)


def report_paid_tasks(req):

    date_from = req['date_from']
    date_to = req['date_to']

    extended_date_from = (datetime.strptime(date_from, '%d.%m.%Y') - timedelta(days=31)).strftime('%Y-%m-%d')
    extended_date_to = (datetime.strptime(date_to, '%d.%m.%Y') + timedelta(days=31)).strftime('%Y-%m-%d')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Сверка'

    ws.append([f'Задачи закрытые за {format_date(date_from)} - {format_date(date_to)}'])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    department_users = b.get_all('user.get', {
        'filter': {
            'UF_DEPARTMENT': ['5', '27', '29', '458'] # ЦС, ЛК '231' не считаются
        }
    })

    department_user_ids = {str(user['ID']) for user in department_users}


    tasks = b.get_all( # задачи из Платных работ
        'tasks.task.list',
        {
            'filter': {
                'GROUP_ID': 408,
                '>=CLOSED_DATE': extended_date_from,
                '<=CLOSED_DATE': extended_date_to
            },
            'select': [
                'ID',
                'RESPONSIBLE_ID',
                'CLOSED_DATE',
                'UF_AUTO_210185778353', # Оплачено
                'UF_AUTO_324696461819', # Номер реализации
                'UF_CRM_TASK'
            ]
        }
    )


    # номера реализаций из задач
    task_keys = set()
    period_task_keys = set()

    user_ids = set()
    company_ids = set()

    for task in tasks:

        number = normalize_number(task.get('ufAuto324696461819'))
        year = get_year(task.get('closedDate'))

        if number and year:

            task_keys.add((year, number)) # Все задачи для проверки существования
            closed_date_obj = datetime.fromisoformat(task.get('closedDate')).date() # Только задачи внутри периода отчета

            if (datetime.strptime(date_from, '%d.%m.%Y').date() <= closed_date_obj <= datetime.strptime(date_to, '%d.%m.%Y').date()):
                period_task_keys.add((year, number))

        responsible = task.get('responsibleId')

        if responsible:
            user_ids.add(str(responsible))

        for link in task.get('ufCrmTask',[]):

            if str(link).startswith('CO_'):
                company_ids.add(str(link).replace('CO_', ''))


    # соответствующие номера в долгах
    numbers = list({key[1] for key in task_keys})
    debts = []
    if numbers:

        debts = b.get_all(
            'crm.item.list',
            {
                'entityTypeId': 161,
                'select': [
                    'id',
                    'ufCrm41_Provider', # Исполнитель услуги
                    'ufCrm41_1689101328', # Сумма документа
                    'ufCrm41_1689101306', # Номер в 1С
                    'ufCrm41_1690546413', # Компания (название)
                    'ufCrm41_1689101272' # Дата создания
                ],
                'filter': {
                    '@ufCrm41_1689101306': numbers, # Номер в 1С
                    '!ufCrm41_1733668627': '1' # СтатусНомераВРеестре
                }
            }
        )

    # долги с датой создания в заданном периоде
    extra_debts = b.get_all(
        'crm.item.list',
        {
            'entityTypeId': 161,
            'select': [
                'id',
                'ufCrm41_Provider',
                'ufCrm41_1689101328',
                'ufCrm41_1689101306',
                'ufCrm41_1690546413',
                'ufCrm41_1689101272'
            ],
            'filter': {
                '>=ufCrm41_1689101272': date_from,
                '<=ufCrm41_1689101272': date_to,
                '!ufCrm41_1733668627': '1'
            }
        }
    )

    # собираем id исполнителей и id компаний
    for debt in debts + extra_debts:

        provider = debt.get('ufCrm41_Provider')
        if provider:
            user_ids.add(str(int(float(provider))))

        company = debt.get('ufCrm41_1690546413')
        if company:
            company_ids.add(str(company))


    # собираем фио исполнителей
    users = {}
    if user_ids:
        users_data = b.get_all('user.get', {'ID': list(user_ids)})

        for user in users_data:
            users[str(user['ID'])] = (f"{user['NAME']} {user['LAST_NAME']}")


    # собираем названия компаний
    companies = {}
    if company_ids:
        companies_data = b.get_all(
            'crm.company.list',
            {
                'filter': {'ID': list(company_ids)},
                'select': ['ID', 'TITLE']
            }
        )

        for company in companies_data:
            companies[str(company['ID'])] = company['TITLE']


    # приводим к ок виду данные из соу
    debts_dict = {}

    def add_debt(debt):

        number = normalize_number(debt.get('ufCrm41_1689101306'))
        if not number:
            return

        act_date = debt.get('ufCrm41_1689101272')
        year = get_year(act_date)

        key = (year, number)

        provider = debt.get('ufCrm41_Provider')
        provider_id = ''
        if provider:
            provider_id = str(int(float(provider)))

        debts_dict[key] = {
            'number': number,
            'id': debt.get('id'),
            'executor': users.get(provider_id, ''),
            'company': debt.get('ufCrm41_1690546413', ''),
            'amount': debt.get('ufCrm41_1689101328'),
            'date': format_date(act_date)
        }

    for debt in debts:
        add_debt(debt)


    # долги по докам без задач
    for debt in extra_debts:

        number = normalize_number(debt.get('ufCrm41_1689101306'))
        if not number:
            continue

        act_date = debt.get('ufCrm41_1689101272')
        year = get_year(act_date)

        key = (year, number)
        if key in debts_dict:
            continue

        provider = debt.get('ufCrm41_Provider')
        provider_id = ''

        if provider:
            provider_id = str(int(float(provider)))

        if provider_id not in department_user_ids:
            continue

        if key in task_keys:
            continue

        add_debt(debt)


    # преобразуем задачи
    tasks_dict = {}
    tasks_without_number = []

    for task in tasks:

        number = normalize_number(task.get('ufAuto324696461819'))

        closed_date = task.get('closedDate')
        year = get_year(closed_date)

        company_name = ''
        for link in task.get('ufCrmTask', []):
            if str(link).startswith('CO_'):
                company_name = companies.get(str(link).replace('CO_', ''), '')
                break

        task_data = {
            'number': number,
            'id': task.get('id'),
            'executor': users.get(str(task.get('responsibleId')), ''),
            'company': company_name,
            'amount': task.get('ufAuto210185778353'),
            'date': format_date(closed_date)
        }

        if not number:
            closed_date_obj = datetime.fromisoformat(closed_date).date()
            if (datetime.strptime(date_from, '%d.%m.%Y').date() 
                <= closed_date_obj 
                <= datetime.strptime(date_to, '%d.%m.%Y').date()):
                tasks_without_number.append(task_data)
            continue

        #tasks_dict[(year, number)] = task_data
        key = (year, number)

        closed_date_obj = datetime.fromisoformat(closed_date).date()

        is_period_task = (datetime.strptime(date_from, '%d.%m.%Y').date()
            <= closed_date_obj
            <= datetime.strptime(date_to, '%d.%m.%Y').date())

        # задачи периода всегда приоритетнее
        if (key not in tasks_dict or is_period_task):
            tasks_dict[key] = task_data


    # таблицы
    matched_rows = []
    no_debt_rows = []
    no_task_rows = []
    no_number_rows = []

    #all_keys = set(debts_dict.keys()) | set(period_task_keys)

    for key in sorted(period_task_keys):

        debt = debts_dict.get(key)
        task = tasks_dict.get(key)

        # есть везде
        if debt and task:

            matched_rows.append([
                debt['number'],

                task['id'],
                debt['id'],

                task['executor'],
                debt['executor'],

                task['amount'],
                debt['amount'],

                task['company'],
                debt['company'],
            ])


        # нет в долгах
        elif task and not debt:

            no_debt_rows.append([
                task['number'],
                task['id'],
                task['executor'],
                task['company'],
                task['amount'],
                task['date']
            ])

    # нет в задачах
    for key, debt in debts_dict.items():

        # если задача существует даже вне периода, то долг НЕ выводим
        if key in task_keys:
            continue

        no_task_rows.append([
            debt['number'],
            debt['id'],
            debt['executor'],
            debt['company'],
            debt['amount'],
            debt['date']
        ])

    # задачи без номера
    for task in tasks_without_number:

        no_number_rows.append([
            task['id'],
            task['executor'],
            task['company'],
            task['amount'],
            task['date']
        ])


    # формируем таблицы

    write_table(
        ws,
        'Работа найдена и в задачах, и в СОУ',
        [
            'Номер реализации',
            'ID задачи',
            'ID элемента СОУ',
            'Исполнитель задача',
            'Исполнитель СОУ',
            'Сумма задача',
            'Сумма СОУ',
            'Компания задача',
            'Компания СОУ',
        ],
        matched_rows
    )

    write_table(
        ws,
        'Работа не найдена в СОУ',
        [
            'Номер реализации',
            'ID задачи',
            'Исполнитель',
            'Компания',
            'Сумма',
            'Дата закрытия'
        ],
        no_debt_rows
    )

    write_table(
        ws,
        'Работа не найдена в задачах',
        [
            'Номер реализации',
            'ID элемента СОУ',
            'Исполнитель',
            'Компания',
            'Сумма',
            'Дата акта'
        ],
        no_task_rows
    )

    write_table(
        ws,
        'Задачи без номера реализации',
        [
            'ID задачи',
            'Исполнитель',
            'Компания',
            'Сумма',
            'Дата закрытия'
        ],
        no_number_rows
    )

    change_sheet_style(ws)

    report_name = f'Отчет_по_платным_задачам_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
    wb.save(report_name)

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '2041440'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })

    b.call('im.notify.system.add', {
        'USER_ID': req['user_id'][5:],
        'MESSAGE': f'Отчет по платным задачам сформирован. {upload_report["DETAIL_URL"]}'})
    os.remove(report_name)


if __name__ == '__main__':

    report_paid_tasks(
        '''{
        'date_from': '2026-05-01',
        'date_to': '2026-05-31',
        'user_id': 'user_1391'
    }''')