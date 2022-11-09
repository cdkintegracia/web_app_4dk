import os
from datetime import timedelta
from datetime import datetime
import base64
from time import strptime

import openpyxl
from openpyxl.utils import get_column_letter
from fast_bitrix24 import Bitrix
import dateutil.parser

from web_app_4dk.modules.authentication import authentication


webhook = authentication('Bitrix')
b = Bitrix(webhook)
month_string = {
        '01': 'Январь',
        '02': 'Февраль',
        '03': 'Март',
        '04': 'Апрель',
        '05': 'Май',
        '06': 'Июнь',
        '07': 'Июль',
        '08': 'Август',
        '09': 'Сентябрь',
        '10': 'Октябрь',
        '11': 'Ноябрь',
        '12': 'Декабрь'
    }

month_codes = {
        'Январь': '1',
        'Февраль': '2',
        'Март': '3',
        'Апрель': '4',
        'Май': '5',
        'Июнь': '6',
        'Июль': '7',
        'Август': '8',
        'Сентябрь': '9',
        'Октябрь': '10',
        'Ноябрь': '11',
        'Декабрь': '12'
}


def create_company_call_report(req):

    # Формирование заголовков отчета
    report_created_time = datetime.now()
    company_id = req['id']
    total_duration = timedelta()
    contacts = b.get_all('crm.company.contact.items.get', {'id': company_id})
    deals = b.get_all('crm.deal.list', {'select': ['ID'], 'filter': {'COMPANY_ID': req['id']}})
    company_info = b.get_all('crm.company.get', {'ID': company_id})
    company_name = company_info['TITLE']
    titles = [
        # 1 строка
        [
            company_name,
            f'{req["month"]} {req["year"]}',  # Месяц и год
            f'Дата формирования {report_created_time.strftime("%d.%m.%Y %H:%M")}',
        ],
        # 2 строка
        [
            'Дата звонка',
            'Контакт (кому звонили)',
            'Номер телефона, на который звонили',
            'Хронометраж',
            'Кто звонил',
        ]
    ]
    report_data = []

    # Формирование отчета
    for contact in contacts:
        contact_info = b.get_all('crm.contact.get', {'id': contact['CONTACT_ID']})
        contact_name = f"{contact_info['LAST_NAME']} {contact_info['NAME']} {contact_info['SECOND_NAME']}"
        not_sorted_activities = b.get_all('crm.activity.list', {
            'filter': {
                'OWNER_TYPE_ID': '3',
                'OWNER_ID': contact['CONTACT_ID'],
                'PROVIDER_TYPE_ID': 'CALL'
            }})
        activities = sorted(not_sorted_activities, key=lambda x: x['ID'])

        for activity in activities:
            if 'Исходящий' in activity['SUBJECT']:
                call_start_time = dateutil.parser.isoparse(activity['START_TIME'])
                if str(call_start_time.month) == month_codes[req['month']] and str(call_start_time.year) == req['year']:
                    if activity['AUTHOR_ID'] == '109':  # Ридкобород
                        continue
                    author = b.get_all('user.get', {'ID': activity['AUTHOR_ID']})[0]
                    if 231 not in author['UF_DEPARTMENT']:
                        continue
                    author_name = f"{author['NAME']} {author['LAST_NAME']}"
                    phone_number = activity['SUBJECT'].split(' на ')[1]
                    check_mobile_phone = phone_number.split(' ')
                    '''
                    if check_mobile_phone[1] == '812':
                        phone_number = ''
                    '''
                    call_end_time = dateutil.parser.isoparse(activity['END_TIME'])
                    call_start_time_formatted = call_start_time.strftime('%d.%m.%Y %H:%M:%S')
                    duration = call_end_time - call_start_time
                    report_data.append([
                        call_start_time_formatted,
                        contact_name,
                        phone_number,
                        duration,
                        author_name,
                    ])
                    duration = call_end_time - call_start_time
                    total_duration += duration

    for deal in deals:
        not_sorted_activities = b.get_all('crm.activity.list', {
            'filter': {
                'OWNER_TYPE_ID': '2',
                'OWNER_ID': deal['ID'],
                'PROVIDER_TYPE_ID': 'CALL'
            }})
        activities = sorted(not_sorted_activities, key=lambda x: x['ID'])
        for activity in activities:
            if 'Исходящий' in activity['SUBJECT']:
                call_start_time = dateutil.parser.isoparse(activity['START_TIME'])
                if str(call_start_time.month) == month_codes[req['month']] and str(call_start_time.year) == req['year']:
                    if activity['AUTHOR_ID'] == '109':  # Ридкобород
                        continue
                    author = b.get_all('user.get', {'ID': activity['AUTHOR_ID']})[0]
                    if 231 not in author['UF_DEPARTMENT']:
                        continue
                    author_name = f"{author['NAME']} {author['LAST_NAME']}"
                    phone_number = activity['SUBJECT'].split(' на ')[1]
                    call_end_time = dateutil.parser.isoparse(activity['END_TIME'])
                    call_start_time_formatted = call_start_time.strftime('%d.%m.%Y %H:%M:%S')
                    duration = call_end_time - call_start_time
                    report_data.append([
                        call_start_time_formatted,
                        '',
                        phone_number,
                        duration,
                        author_name,
                    ])
                    duration = call_end_time - call_start_time
                    total_duration += duration
    report_data = sorted(report_data, key=lambda x: strptime(x[0], "%d.%m.%Y %H:%M:%S"))
    report_data.append(['Итого', '', '', total_duration])
    report_data = titles + report_data

    # Создание xlsx файла отчета
    report_name_time = report_created_time.strftime('%d-%m-%Y %H %M %S %f')
    report_name = f'Отчет по звонкам {company_name} {report_name_time}.xlsx'.replace(' ', '_')
    workbook = openpyxl.Workbook()
    worklist = workbook.active

    for data in report_data:
        worklist.append(data)
    for idx, col in enumerate(worklist.columns, 1):
        worklist.column_dimensions[get_column_letter(idx)].auto_size = True
    workbook.save(report_name)

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '187139'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })

    b.call('crm.timeline.comment.add', {
        'fields': {
            'ENTITY_ID': company_id,
            'ENTITY_TYPE': 'company',
            'COMMENT': f'Отчет по звонкам:\n'
                       f'{upload_report["DETAIL_URL"]}',
            'AUTHOR_ID': '173'
        }})

    os.remove(report_name)
