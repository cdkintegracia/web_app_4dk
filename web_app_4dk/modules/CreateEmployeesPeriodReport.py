from datetime import datetime, timedelta
from calendar import monthrange
import base64
import os

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from fast_bitrix24 import Bitrix

from web_app_4dk.modules.authentication import authentication
from web_app_4dk.modules.EdoInfoHandler import month_codes, year_codes
from web_app_4dk.modules.field_values import month_int_names


b = Bitrix(authentication('Bitrix'))
deals_info_files_directory = f'/root/web_app_4dk/web_app_4dk/modules/deals_info_files/'


def get_employee_id(users: str) -> list:
    """
    Приводит строку с id пользователей и id подразделений к единому списку, состоящему только из id сотрудников

    :param users: Строка из параметра запроса Б24 состоящая из {user_...} и|или {group_...}, которые разделены ', '
    """

    users_id_set = set()

    # Строка с сотрудниками и отделами в список
    users = users.split(', ')

    # Если в массиве найден id сотрудника
    for user_id in users:
        if 'user' in user_id:
            users_id_set.add(user_id[5:])

        # Если в массиве найден id отдела
        elif 'group' in user_id:
            department_users = b.get_all('user.get', {'filter': {'UF_DEPARTMENT': user_id[8:]}})
            for user in department_users:
                users_id_set.add(user['ID'])

    return list(users_id_set)


def get_fio_from_user_info(user_info: dict) -> str:
    """
    Возвращает строку, состоящую из фамилии и имени пользователя Б24

    :param user_info: Словарь, полученный запросом с методом user.get
    """

    '''
    return f'{user_info["LAST_NAME"] if "LAST_NAME" in user_info else ""}'\
           f' {user_info["NAME"] if "NAME" in user_info else ""}'.strip()
    '''

    return (f'{user_info["NAME"] if "NAME" in user_info else ""}'
            f' {user_info["LAST_NAME"] if "LAST_NAME" in user_info else ""}').strip()


def change_sheet_style(sheet) -> None:

    # Изменение ширины
    for column_cells in sheet.columns:
        length = max(len(str(cell.value) if cell.value else '') for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 1.1

def read_deals_data_file(month, year):
    filename = f'{month_int_names[month]}_{year}.xlsx'
    workbook = openpyxl.load_workbook(f'{deals_info_files_directory}{filename}')
    worksheet = workbook.active
    file_data = []
    for index, row in enumerate(worksheet.rows):
        if index == 0:
            file_titles = list(map(lambda x: x.value, row))
        else:
            row_data = list(map(lambda x: x.value, row))
            row_data = dict(zip(file_titles, row_data))
            if not row_data['Тип'] or row_data['Тип'] == 'Тестовый':
                continue
            file_data.append(row_data)
    return file_data


def create_employees_period_report(req):
    users_id = get_employee_id(req['users'])
    users_info = b.get_all('user.get', {
        'filter': {
            'ACTIVE': 'true',
            'ID': users_id,
        }
    })

    deal_fields = b.get_all('crm.deal.fields')

    start_period = req['start_date']
    end_period = req['end_date']
    print(start_period)
    print(end_period)

    before_1_month_year = datetime.now().year
    before_1_month = datetime.now().month - 1
    #before_1_month = 12 #для теста, как будто уже янв 2025

    if before_1_month == 0:
        before_1_month = 12
        before_1_month_year -= 1
    before_1_month_range = monthrange(before_1_month_year, before_1_month)[1]
    before_1_month_last_day_date = (f'{before_1_month_range}.'
                                  f'{before_1_month if len(str(before_1_month)) == 2 else "0" + str(before_1_month)}.'
                                  f'{before_1_month_year}')
 
    ddmmyyyy_pattern = '%d.%m.%Y'

    start_filter = datetime(day=1, month=1, year=before_1_month_year)
    end_filter = datetime(day=1, month=datetime.now().month, year=datetime.now().year)
    #end_filter = datetime(day=1, month=1, year=datetime.now().year+1) #для теста, как будто уже янв 2025
    last_day_of_last_year = datetime(day=31, month=12, year=before_1_month_year-1)

    deal_group_field = deal_fields['UF_CRM_1657878818384']['items']
    deal_group_field.append({'ID': None, 'VALUE': 'Лицензии'})
    deal_group_field.append({'ID': None, 'VALUE': 'Остальные'})

    workbook = openpyxl.Workbook()
    report_name = f'Годовой_отчет_по_сотрудникам_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'

    for index, user_info in enumerate(users_info):
        user_name = get_fio_from_user_info(user_info)
        if index == 0:
            worksheet = workbook.active
            worksheet.title = user_name
        else:
            worksheet = workbook.create_sheet(user_name)

        first_month_deals_data = read_deals_data_file(12, before_1_month_year-1) #конец прошлого года
        before_1_month_deals_data = read_deals_data_file(before_1_month, before_1_month_year) #последний месяц

        worksheet.append([user_name, '', f'{start_filter.year} г.'])
        worksheet.append([])
        worksheet.append([])

        its_deals_before_1_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                                x['Группа'] == 'ИТС' and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                before_1_month_deals_data))

        first_month_its_deals = list(filter(lambda x: x['Ответственный'] == user_name and
                                                x['Группа'] == 'ИТС' and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                first_month_deals_data))

        # Сделки
        # Последний месяц
        its_prof_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                                x['Группа'] == 'ИТС' and
                                                'Базовый' not in x['Тип'] and 'ГРМ' not in x['Тип'] and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                before_1_month_deals_data))

        its_base_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                                x['Группа'] == 'ИТС' and
                                                'Базовый' in x['Тип'] and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                before_1_month_deals_data))

        countragent_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                   'Контрагент' in x['Тип'] and
                                                   x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                   before_1_month_deals_data))

        spark_in_contract_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                         '1Спарк в договоре' == x['Тип'] and
                                                         x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                         before_1_month_deals_data))

        spark_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                  ('1Спарк 3000' == x['Тип'] or '1Спарк' == x['Тип']) and
                                                  x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                  before_1_month_deals_data))

        spark_plus_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                   ('22500' in x['Тип'] or '1СпаркПЛЮС' in x['Тип']) and
                                                   x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                   before_1_month_deals_data))

        rpd_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'РПД' in x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           before_1_month_deals_data))
        
        grm_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ГРМ' in x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           before_1_month_deals_data))
        
        doki_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                            'Доки' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            before_1_month_deals_data))
        
        report_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                              'Отчетность' in x['Тип'] and 
                                              x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                              before_1_month_deals_data))
        
        signature_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                                 'Подпись' in x['Тип'] and
                                                 x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                 before_1_month_deals_data))
        
        dop_oblako_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                                  'Допы Облако' == x['Тип'] and
                                                  x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                  before_1_month_deals_data))
        
        link_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                            'Линк' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            before_1_month_deals_data))
        
        unics_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                             'Уникс' == x['Тип'] and
                                             x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                             before_1_month_deals_data))
        
        cab_sotrudnik_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                     'Кабинет сотрудника' == x['Тип'] and
                                                     x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                     before_1_month_deals_data))
        
        cab_sadovod_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                   'Кабинет садовода' == x['Тип'] and
                                                   x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                   before_1_month_deals_data))
        
        edo_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ЭДО' == x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           before_1_month_deals_data))
        
        mdlp_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                            'МДЛП' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            before_1_month_deals_data))
        
        connect_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                               'Коннект' == x['Тип'] and
                                               x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                               before_1_month_deals_data))
                
        its_otrasl_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                  'ИТС Отраслевой' == x['Тип'] and
                                                  x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                  before_1_month_deals_data))
        
        ofd_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ОФД' == x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           before_1_month_deals_data))
        
        bitrix24_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                ('Битрикс24' == x['Тип'] or 'БИТРИКС24' == x['Тип']) and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                before_1_month_deals_data))

        other_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                             x not in its_prof_deals_last_month and x not in its_base_deals_last_month and
                                             x not in countragent_deals_last_month and x not in spark_in_contract_deals_last_month and
                                             x not in spark_deals_last_month and x not in spark_plus_deals_last_month and
                                             x not in rpd_deals_last_month and x not in grm_deals_last_month and
                                             x not in doki_deals_last_month and
                                             x not in report_deals_last_month and x not in signature_deals_last_month and
                                             x not in dop_oblako_deals_last_month and x not in link_deals_last_month and
                                             x not in unics_deals_last_month and x not in cab_sotrudnik_deals_last_month and
                                             x not in cab_sadovod_deals_last_month and x not in edo_deals_last_month and
                                             x not in mdlp_deals_last_month and x not in connect_deals_last_month and
                                             x not in its_otrasl_deals_last_month and x not in ofd_deals_last_month and
                                             x not in bitrix24_deals_last_month and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'], before_1_month_deals_data))        
        
        # Первый месяц
        its_prof_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Группа'] == 'ИТС' and
                                             'Базовый' not in x['Тип'] and 'ГРМ' not in x['Тип'] and
                                             x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                             first_month_deals_data))

        its_base_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Группа'] == 'ИТС' and
                                             'Базовый' in x['Тип'] and
                                             x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                             first_month_deals_data))

        countragent_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                'Контрагент' in x['Тип'] and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                first_month_deals_data))

        spark_in_contract_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                      '1Спарк в договоре' == x['Тип'] and
                                                      x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                      first_month_deals_data))

        spark_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                               ('1Спарк 3000' == x['Тип'] or '1Спарк' == x['Тип']) and
                                               x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                               first_month_deals_data))

        spark_plus_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                ('22500' in x['Тип'] or '1СпаркПЛЮС' in x['Тип']) and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                first_month_deals_data))

        rpd_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                        'РПД' in x['Тип'] and
                                        x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                        first_month_deals_data))
        
        #добавлено 15-07-2024
        grm_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ГРМ' in x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           first_month_deals_data))
        
        doki_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                            'Доки' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            first_month_deals_data))
        
        report_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                              'Отчетность' in x['Тип'] and 
                                              x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                              first_month_deals_data))
        
        signature_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                                 'Подпись' in x['Тип'] and
                                                 x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                 first_month_deals_data))
        
        dop_oblako_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and 
                                                  'Допы Облако' == x['Тип'] and
                                                  x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                  first_month_deals_data))
        
        link_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                            'Линк' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            first_month_deals_data))
        
        unics_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                             'Уникс' == x['Тип'] and
                                             x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                             first_month_deals_data))
        
        cab_sotrudnik_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                     'Кабинет сотрудника' == x['Тип'] and
                                                     x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                     first_month_deals_data))
        
        cab_sadovod_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                   'Кабинет садовода' == x['Тип'] and
                                                   x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                   first_month_deals_data))
        
        edo_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ЭДО' == x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           first_month_deals_data))
        
        mdlp_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                            'МДЛП' == x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            first_month_deals_data))
        
        connect_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                               'Коннект' == x['Тип'] and
                                               x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                               first_month_deals_data))
                
        its_otrasl_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                  'ИТС Отраслевой' == x['Тип'] and
                                                  x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                  first_month_deals_data))
        
        ofd_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                           'ОФД' == x['Тип'] and
                                           x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                           first_month_deals_data))
        
        bitrix24_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and x['Тип'] and
                                                ('Битрикс24' == x['Тип'] or 'БИТРИКС24' == x['Тип']) and
                                                x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                first_month_deals_data))

        other_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                             x not in its_prof_deals_first_month and x not in its_base_deals_first_month and
                                             x not in countragent_deals_first_month and x not in spark_in_contract_deals_first_month and
                                             x not in spark_deals_first_month and x not in spark_plus_deals_first_month and
                                             x not in rpd_deals_first_month and x not in grm_deals_first_month and
                                             x not in doki_deals_first_month and
                                             x not in report_deals_first_month and x not in signature_deals_first_month and
                                             x not in dop_oblako_deals_first_month and x not in link_deals_first_month and
                                             x not in unics_deals_first_month and x not in cab_sotrudnik_deals_first_month and
                                             x not in cab_sadovod_deals_first_month and x not in edo_deals_first_month and
                                             x not in mdlp_deals_first_month and x not in connect_deals_first_month and
                                             x not in its_otrasl_deals_first_month and x not in ofd_deals_first_month and
                                             x not in bitrix24_deals_first_month and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'], first_month_deals_data))
        
        worksheet.append(['Сделки', f'на {before_1_month_last_day_date}', f'на {last_day_of_last_year.strftime("%d.%m.%Y")}', 'Прирост с начала года'])
        worksheet.append([
            'ИТС ПРОФ',
            len(its_prof_deals_last_month),
            len(its_prof_deals_first_month),
            len(its_prof_deals_last_month) - len(its_prof_deals_first_month)
        ])
        worksheet.append([
            'ИТС Базовые',
            len(its_base_deals_last_month),
            len(its_base_deals_first_month),
            len(its_base_deals_last_month) - len(its_base_deals_first_month)
        ])
        worksheet.append([
            'Контрагент',
            len(countragent_deals_last_month),
            len(countragent_deals_first_month),
            len(countragent_deals_last_month) - len(countragent_deals_first_month)
        ])
        worksheet.append([
            'Спарк в договоре',
            len(spark_in_contract_deals_last_month),
            len(spark_in_contract_deals_first_month),
            len(spark_in_contract_deals_last_month) - len(spark_in_contract_deals_first_month)
        ])
        worksheet.append([
            'Спарк',
            len(spark_deals_last_month),
            len(spark_deals_first_month),
            len(spark_deals_last_month) - len(spark_deals_first_month)
        ])
        worksheet.append([
            'СпаркПлюс',
            len(spark_plus_deals_last_month),
            len(spark_plus_deals_first_month),
            len(spark_plus_deals_last_month) - len(spark_plus_deals_first_month)
        ])
        worksheet.append([
            'РПД',
            len(rpd_deals_last_month),
            len(rpd_deals_first_month),
            len(rpd_deals_last_month) - len(rpd_deals_first_month)
        ])
        worksheet.append([
            'ГРМ',
            len(grm_deals_last_month),
            len(grm_deals_first_month),
            len(grm_deals_last_month) - len(grm_deals_first_month)
        ])
        worksheet.append([
            'Доки',
            len(doki_deals_last_month),
            len(doki_deals_first_month),
            len(doki_deals_last_month) - len(doki_deals_first_month)
        ])
        worksheet.append([
            'Отчетности',
            len(report_deals_last_month),
            len(report_deals_first_month),
            len(report_deals_last_month) - len(report_deals_first_month)
        ])
        worksheet.append([
            'Подпись',
            len(signature_deals_last_month),
            len(signature_deals_first_month),
            len(signature_deals_last_month) - len(signature_deals_first_month)
        ])
        worksheet.append([
            'Допы Облако',
            len(dop_oblako_deals_last_month),
            len(dop_oblako_deals_first_month),
            len(dop_oblako_deals_last_month) - len(dop_oblako_deals_first_month)
        ])
        worksheet.append([
            'Линк',
            len(link_deals_last_month),
            len(link_deals_first_month),
            len(link_deals_last_month) - len(link_deals_first_month)
        ])
        worksheet.append([
            'Уникс',
            len(unics_deals_last_month),
            len(unics_deals_first_month),
            len(unics_deals_last_month) - len(unics_deals_first_month)
        ])
        worksheet.append([
            'Кабинет сотрудника',
            len(cab_sotrudnik_deals_last_month),
            len(cab_sotrudnik_deals_first_month),
            len(cab_sotrudnik_deals_last_month) - len(cab_sotrudnik_deals_first_month)
        ])
        worksheet.append([
            'Кабинет садовода',
            len(cab_sadovod_deals_last_month),
            len(cab_sadovod_deals_first_month),
            len(cab_sadovod_deals_last_month) - len(cab_sadovod_deals_first_month)
        ])
        worksheet.append([
            'ЭДО',
            len(edo_deals_last_month),
            len(edo_deals_first_month),
            len(edo_deals_last_month) - len(edo_deals_first_month)
        ])
        worksheet.append([
            'МДЛП',
            len(mdlp_deals_last_month),
            len(mdlp_deals_first_month),
            len(mdlp_deals_last_month) - len(mdlp_deals_first_month)
        ])
        worksheet.append([
            'Коннект',
            len(connect_deals_last_month),
            len(connect_deals_first_month),
            len(connect_deals_last_month) - len(connect_deals_first_month)
        ])
        worksheet.append([
            'ИТС Отраслевой',
            len(its_otrasl_deals_last_month),
            len(its_otrasl_deals_first_month),
            len(its_otrasl_deals_last_month) - len(its_otrasl_deals_first_month)
        ])
        worksheet.append([
            'ОФД',
            len(ofd_deals_last_month),
            len(ofd_deals_first_month),
            len(ofd_deals_last_month) - len(ofd_deals_first_month)
        ])
        if (len(bitrix24_deals_last_month) != 0) or (len(bitrix24_deals_first_month) != 0):
            worksheet.append([
                'Битрикс24',
                len(bitrix24_deals_last_month),
                len(bitrix24_deals_first_month),
                len(bitrix24_deals_last_month) - len(bitrix24_deals_first_month)
            ])
        worksheet.append([
            'Прочие',
            len(other_deals_last_month),
            len(other_deals_first_month),
            len(other_deals_last_month) - len(other_deals_first_month)
        ])
        worksheet.append([])
        
        # Охват сервисами
        #Последний месяц
        companies = set(map(lambda x: x['Компания'], list(filter(lambda x: x['Ответственный за компанию'] == user_name, before_1_month_deals_data)))) #компании из сделок отв на последний месяц
        companies_without_services_last_month = 0
        companies_without_paid_services_last_month = 0
        for company in companies:
            company_regnumbers = set(map(lambda x: x['Регномер'], list(filter(lambda x: x['Компания'] == company, before_1_month_deals_data)))) #регномера компании из сделок на последний месяц
            company_its = list(filter(lambda x: x['Группа'] == 'ИТС' and company == x['Компания'], before_1_month_deals_data)) #сделки итс компании на последний месяц
            if not company_its:
                continue
            #сделки итс компании КРОМЕ профов, облако и грм (т.е. базовый земля, идвид, беспл, садовод, медицина, аов)
            non_prof_its = list(filter(lambda x: x['Группа'] == 'ИТС' and company == x['Компания'] and 'ПРОФ' not in x['Тип'] and 'Облако' not in x['Тип'] and 'ГРМ' not in x['Тип'], before_1_month_deals_data))
            if non_prof_its:
                company_its_services = list(filter(lambda x: (company == x['Компания'] or x['Регномер'] in company_regnumbers) and
                                                   ('Контрагент' in x['Тип'] or 'Спарк' in x['Тип'] or
                                                   'РПД' in x['Тип'] or 'Отчетность' in x['Тип'] or
                                                   'Допы Облако' in x['Тип'] or 'Кабинет сотрудника' in x['Тип'] or
                                                   'Подпись 1200' in x['Тип'] or 'Коннект' in x['Тип'] or
                                                   'mag1c' in x['Тип'] or 'Облачный архив' in x['Тип'] or
                                                   'Сканер чехов' in x['Тип']),
                                                    before_1_month_deals_data)) #если есть сделки итс компании КРОМЕ профов, то считаем сервисы включая бесплатную отчетность

                if not company_its_services:
                    companies_without_services_last_month += 1

            company_its_paid_services = list(filter(lambda x: (company == x['Компания'] or x['Регномер'] in company_regnumbers) and
                                                    ('Контрагент' in x['Тип'] or 'Спарк' in x['Тип'] or
                                                   'РПД' in x['Тип'] or 'Отчетность' == x['Тип'] or
                                                   'Допы Облако' in x['Тип'] or 'Кабинет сотрудника' in x['Тип'] or
                                                   'Подпись 1200' in x['Тип'] or 'Коннект' in x['Тип'] or
                                                   'mag1c' in x['Тип'] or 'Облачный архив' in x['Тип'] or
                                                   'Сканер чехов' in x['Тип']),
                                                    before_1_month_deals_data)) #если сделки итс только профы, то считаем сервисы без бесплатной отчетности

            if not company_its_paid_services:
                companies_without_paid_services_last_month += 1

        try:
            coverage_its_without_services_last_month = round(companies_without_services_last_month /
                                                                    len(its_deals_before_1_month) * 100, 2)
        except ZeroDivisionError:
            coverage_its_without_services_last_month = 0

        try:
            coverage_its_without_paid_services_last_month = round(companies_without_paid_services_last_month /
                                                                    len(its_deals_before_1_month) * 100, 2)
        except ZeroDivisionError:
            coverage_its_without_paid_services_last_month = 0

        #Первый месяц
        companies = set(map(lambda x: x['Компания'], list(filter(lambda x: x['Ответственный за компанию'] == user_name, first_month_deals_data))))
        companies_without_services_first_month = 0
        companies_without_paid_services_first_month = 0
        for company in companies:
            company_regnumbers = set(map(lambda x: x['Регномер'], list(filter(lambda x: x['Компания'] == company, first_month_deals_data))))
            company_its = list(filter(lambda x: x['Группа'] == 'ИТС' and company == x['Компания'], first_month_deals_data))
            if not company_its:
                continue

            non_prof_its = list(filter(lambda x: x['Группа'] == 'ИТС' and company == x['Компания'] and 'ПРОФ' not in x[
                'Тип'] and 'Облако' not in x['Тип'] and 'ГРМ' not in x['Тип'], first_month_deals_data))
            if non_prof_its:
                company_its_services = list(filter(lambda x: (company == x['Компания'] or x['Регномер'] in company_regnumbers) and
                                                    ('Контрагент' in x['Тип'] or 'Спарк' in x['Тип'] or
                                                   'РПД' in x['Тип'] or 'Отчетность' in x['Тип'] or
                                                   'Допы Облако' in x['Тип'] or 'Кабинет сотрудника' in x['Тип'] or
                                                   'Подпись 1200' in x['Тип'] or 'Коннект' in x['Тип'] or
                                                   'mag1c' in x['Тип'] or 'Облачный архив' in x['Тип'] or
                                                   'Сканер чехов' in x['Тип']),
                                                    first_month_deals_data))

                if not company_its_services:
                    companies_without_services_first_month += 1

            company_its_paid_services = list(filter(lambda x: (company == x['Компания'] or x['Регномер'] in company_regnumbers) and
                                                    ('Контрагент' in x['Тип'] or 'Спарк' in x['Тип'] or
                                                   'РПД' in x['Тип'] or 'Отчетность' == x['Тип'] or
                                                   'Допы Облако' in x['Тип'] or 'Кабинет сотрудника' in x['Тип'] or
                                                   'Подпись 1200' in x['Тип'] or 'Коннект' in x['Тип'] or
                                                   'mag1c' in x['Тип'] or 'Облачный архив' in x['Тип'] or
                                                   'Сканер чехов' in x['Тип']),
                                                    first_month_deals_data))

            if not company_its_paid_services:
                companies_without_paid_services_first_month += 1

        try:
            coverage_its_without_services_first_month = round(companies_without_services_first_month /
                                                                   len(first_month_its_deals) * 100, 2)
        except ZeroDivisionError:
            coverage_its_without_services_first_month = 0

        try:
            coverage_its_without_paid_services_first_month = round(companies_without_paid_services_first_month /
                                                                        len(first_month_its_deals) * 100, 2)            
        except ZeroDivisionError:
            coverage_its_without_paid_services_first_month = 0

        
        worksheet.append(['Охват сервисами', f'на {before_1_month_last_day_date}', f'на {last_day_of_last_year.strftime("%d.%m.%Y")}', 'Прирост с начала года'])
        worksheet.append([
            'ИТС без сервисов',
            companies_without_services_last_month,
            companies_without_services_first_month,
            companies_without_services_last_month - companies_without_services_first_month,
        ])
        worksheet.append([
            '% ИТС без сервисов',
            f'{coverage_its_without_services_last_month}%',
            f'{coverage_its_without_services_first_month}%',
            f'{round(coverage_its_without_services_last_month - coverage_its_without_services_first_month, 2)}%'
        ])
        worksheet.append([
            'ИТС без платных сервисов',
            companies_without_paid_services_last_month,
            companies_without_paid_services_first_month,
            companies_without_paid_services_last_month - companies_without_paid_services_first_month,
        ])
        worksheet.append([
            '% ИТС без платных сервисов',
            f'{coverage_its_without_paid_services_last_month}%',
            f'{coverage_its_without_paid_services_first_month}%',
            f'{round(coverage_its_without_paid_services_last_month - coverage_its_without_paid_services_first_month, 2)}%'
        ])
        worksheet.append([])
    

        # Отчетность
        # Прошлый месяц
        free_reporting_deals_last_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                                      x['Тип'] == 'Отчетность (в рамках ИТС)' and
                                                      x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                      before_1_month_deals_data))

        try:
            coverage_free_reporting_deals_last_month = round(len(free_reporting_deals_last_month) /
                                                             len(its_prof_deals_last_month) * 100, 2)
        except ZeroDivisionError:
            coverage_free_reporting_deals_last_month = 0

        paid_reporting_deals_last_month_num = 0
        paid_reporting_deals_last_month = 0
        for its_deal in its_deals_before_1_month:
            its_paid_reporting = list(filter(lambda x: (x['Регномер'] == its_deal['Регномер'] and x['Тип'] == 'Отчетность' and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту']), before_1_month_deals_data))
            if its_paid_reporting:
                paid_reporting_deals_last_month += 1
                for i in range(len(its_paid_reporting)):
                    paid_reporting_deals_last_month_num +=1
        try:
            coverage_paid_reporting_deals_last_month = round(paid_reporting_deals_last_month /
                                                             len(its_deals_before_1_month) * 100, 2)
        except ZeroDivisionError:
            coverage_paid_reporting_deals_last_month = 0

        
        any_reporting_deals_last_month = 0
        for its_deal in its_deals_before_1_month:
            any_reporting = list(filter(lambda x: (x['Регномер'] == its_deal['Регномер'] and 'Отчетность' in x['Тип'] and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту']), before_1_month_deals_data))
            if any_reporting:
                any_reporting_deals_last_month += 1
        try:
            coverage_any_reporting_deals_last_month = round(any_reporting_deals_last_month /
                                                             len(its_deals_before_1_month) * 100, 2)
        except ZeroDivisionError:
            coverage_any_reporting_deals_last_month = 0

        #Первый месяц
        free_reporting_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                                      x['Тип'] == 'Отчетность (в рамках ИТС)' and
                                                      x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                                      first_month_deals_data))

        prof_deals_first_month = list(filter(lambda x: x['Ответственный'] == user_name and
                                            x['Группа'] == 'ИТС' and
                                            'Базовый' not in x['Тип'] and
                                            x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'],
                                            first_month_deals_data))

        try:
            coverage_free_reporting_deals_first_month = round(len(free_reporting_deals_first_month) /
                                                             len(prof_deals_first_month) * 100, 2)
        except ZeroDivisionError:
            coverage_free_reporting_deals_first_month = 0

        paid_reporting_deals_first_month = 0
        paid_reporting_deals_first_month_num = 0
        for its_deal in first_month_its_deals:
            its_paid_reporting = list(
                filter(lambda x: (x['Регномер'] == its_deal['Регномер'] and x['Тип'] == 'Отчетность' and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту']),
                       first_month_deals_data))
            if its_paid_reporting:
                paid_reporting_deals_first_month += 1
                for i in range(len(its_paid_reporting)):
                    paid_reporting_deals_first_month_num +=1
        try:
            coverage_paid_reporting_deals_first_month = round(paid_reporting_deals_first_month /
                                                                len(first_month_its_deals) * 100, 2)
        except ZeroDivisionError:
            coverage_paid_reporting_deals_first_month = 0

        any_reporting_deals_first_month = 0
        for its_deal in first_month_its_deals:
            any_reporting = list(filter(lambda x: (x['Регномер'] == its_deal['Регномер'] and 'Отчетность' in x['Тип'] and x['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту']),
                                        first_month_deals_data))
            if any_reporting:
                any_reporting_deals_first_month += 1
        try:
            coverage_any_reporting_deals_first_month = round(any_reporting_deals_first_month /
                                                             len(first_month_its_deals) * 100, 2)
        except ZeroDivisionError:
            coverage_any_reporting_deals_first_month = 0

        worksheet.append(['Отчетность', f'на {before_1_month_last_day_date}', f'на {last_day_of_last_year.strftime("%d.%m.%Y")}', 'Прирост с начала года'])
        worksheet.append([
            'Льготных отчетностей',
            len(free_reporting_deals_last_month),
            len(free_reporting_deals_first_month),
            len(free_reporting_deals_last_month) - len(free_reporting_deals_first_month),
        ])
        worksheet.append([
            'Охват льготной отчетностью',
            f'{coverage_free_reporting_deals_last_month}%',
            f'{coverage_free_reporting_deals_first_month}%',
            f'{round(coverage_free_reporting_deals_last_month - coverage_free_reporting_deals_first_month, 2)}%'
        ])
        worksheet.append([
            'Сделок платных отчетностей',
            paid_reporting_deals_last_month_num,
            paid_reporting_deals_first_month_num,
            paid_reporting_deals_last_month_num - paid_reporting_deals_first_month_num,
        ])
        worksheet.append([
            'ИТС с платной отчетностью',
            paid_reporting_deals_last_month,
            paid_reporting_deals_first_month,
            paid_reporting_deals_last_month - paid_reporting_deals_first_month,
        ])
        worksheet.append([
            'Охват платных отчетностей',
            f'{coverage_paid_reporting_deals_last_month}%',
            f'{coverage_paid_reporting_deals_first_month}%',
            f'{round(coverage_paid_reporting_deals_last_month - coverage_paid_reporting_deals_first_month, 2)}%'
        ])
        worksheet.append([
            'Любая отчетность',
            any_reporting_deals_last_month,
            any_reporting_deals_first_month,
            any_reporting_deals_last_month - any_reporting_deals_first_month,
        ])
        worksheet.append([
            'Охват любой отчетностью',
            f'{coverage_any_reporting_deals_last_month}%',
            f'{coverage_any_reporting_deals_first_month}%',
            f'{round(coverage_any_reporting_deals_last_month - coverage_any_reporting_deals_first_month, 2)}%',
        ])

        worksheet.append([])
   
        # Продажи
        sales = b.get_all('crm.item.list', {
            'entityTypeId': '133',
            'filter': {
                'assignedById': user_info['ID'],
                '>=ufCrm3_1654248264': start_filter.strftime(ddmmyyyy_pattern),
                '<ufCrm3_1654248264': end_filter.strftime(ddmmyyyy_pattern),
            }
        })

        if sales:
            sold_deals = b.get_all('crm.deal.list', {
                'select': ['ID', 'COMPANY_ID', 'UF_CRM_1657878818384', 'OPPORTUNITY', 'TYPE_ID'],
                'filter': {
                    'ID': list(map(lambda x: x['parentId2'], sales))
                }
            })            
        else:
            sold_deals = []

        worksheet.append(['Продажи', f'за {start_filter.year} год, шт.', f'за {start_filter.year} год, руб'])
        for field_value in deal_group_field:
            if field_value['VALUE'] == 'Лицензии':
                grouped_deals = list(filter(lambda x: x['TYPE_ID'] in ['UC_YIAJC8', 'UC_QQPYF0'], sold_deals)) # тип лицензия с купоном или лицензия
            else:
                grouped_deals = list(filter(lambda x: x['UF_CRM_1657878818384'] == field_value['ID'] and x['TYPE_ID'] not in ['UC_YIAJC8', 'UC_QQPYF0'], sold_deals)) # если группы равны и тип не лицензии
            worksheet.append([field_value['VALUE'], len(grouped_deals), sum(list(map(lambda x: float(x['OPPORTUNITY'] if x['OPPORTUNITY'] else 0.0), grouped_deals)))])

        worksheet.append(['Всего по источникам', len(sales), sum(list(map(lambda x: x['opportunity'], sales)))])
        worksheet.append(['Всего по сделкам', len(sold_deals), sum(list(map(lambda x: float(x['OPPORTUNITY']), sold_deals)))])
        worksheet.append([])
        

        # Долги по документам
        documents_debts = b.get_all('crm.item.list', { #все выписанные долги
            'entityTypeId': '161',
            'filter': {
                'assignedById': user_info['ID'],
                '<ufCrm41_1689101272': end_filter.strftime(ddmmyyyy_pattern)
            }
        })

        non_documents_debts = list(filter(lambda x: x['stageId'] in ['DT161_53:NEW', 'DT161_53:1'], documents_debts)) #все незакрытые долги

        period_documents_debts = list(filter(lambda x: #все выписанные долги за период
                                              start_filter.timestamp() <=
                                              datetime.fromisoformat(x['ufCrm41_1689101272']).timestamp(), documents_debts))
        
        non_period_documents_debts = list(filter(lambda x: #все незакрытые долги за период
                                        start_filter.timestamp() <=
                                        datetime.fromisoformat(x['ufCrm41_1689101272']).timestamp(), non_documents_debts))


        worksheet.append(['Долги по документам', 'Всего выписано за год', 'Не сдано за год'])
        worksheet.append(['Штук', len(period_documents_debts), len(non_period_documents_debts)])
        worksheet.append([])
       
        # Задачи
        tasks = b.get_all('tasks.task.list', {
            'filter': {
                'RESPONSIBLE_ID': user_info['ID'],
                '>=CREATED_DATE': start_filter.strftime(ddmmyyyy_pattern),
                '<CREATED_DATE': end_filter.strftime(ddmmyyyy_pattern),
            },
            'select': ['GROUP_ID', 'STATUS', 'UF_AUTO_177856763915']
        })
        completed_tlp_tasks = list(filter(lambda x: x['groupId'] == '1' and x['status'] == '5', tasks))
        tasks_ratings = list(map(lambda x: int(x['ufAuto177856763915']) if x['ufAuto177856763915'] else 0, completed_tlp_tasks))
        tasks_ratings = list(filter(lambda x: x != 0, tasks_ratings))
        try:
            average_tasks_ratings = round(sum(tasks_ratings) / len(tasks_ratings), 2)
        except ZeroDivisionError:
            average_tasks_ratings = '-'

        #Дежурства
        days_duty = b.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '301',
            'filter': {
                'PROPERTY_1753': user_info['ID'],
                'PROPERTY_1771': int(start_filter.year),
            }
        })
        days_duty_amount = len(days_duty)

        worksheet.append(['Закрытые задачи ТЛП', len(completed_tlp_tasks)])
        worksheet.append(['Средняя оценка', average_tasks_ratings])
        worksheet.append(['Дней дежурства', days_duty_amount])
        worksheet.append([])
        
    
        # ЭДО
        all_its = its_prof_deals_last_month + its_base_deals_last_month
        edo_companies_id = list(map(lambda x: x['Компания'], list(filter(lambda y: 'Компания' in y and y['Компания'], all_its))))

        if edo_companies_id:
            edo_companies = b.get_all('crm.company.list', {
                'select': ['UF_CRM_1638093692254', 'UF_CRM_1638093750742'],
                'filter': {
                    'ID': list(map(lambda x: x['Компания'], list(filter(lambda y: 'Компания' in y and y['Компания'], all_its))))
                }
            })
            edo_companies_count = list(filter(lambda x: x['UF_CRM_1638093692254'] == '69', edo_companies))

            #спецоператоры эдо
            try: 
                operator_2ae = list(filter(lambda x: x['UF_CRM_1638093750742'] == '75', edo_companies_count))
            except ZeroDivisionError:
                operator_2ae = 0
            try: 
                operator_2ae_doki = list(filter(lambda x: x['UF_CRM_1638093750742'] == '1715', edo_companies_count))
            except ZeroDivisionError:
                operator_2ae_doki = 0
            try: 
                operator_2be = list(filter(lambda x: x['UF_CRM_1638093750742'] == '77', edo_companies_count))
            except ZeroDivisionError:
                operator_2be = 0
            try: 
                operator_2bm = list(filter(lambda x: x['UF_CRM_1638093750742'] == '73', edo_companies_count))
            except ZeroDivisionError:
                operator_2bm = 0
            try: 
                operator_2al = list(filter(lambda x: x['UF_CRM_1638093750742'] == '437', edo_companies_count))
            except ZeroDivisionError:
                operator_2al = 0
            try: 
                operator_2lb = list(filter(lambda x: x['UF_CRM_1638093750742'] == '439', edo_companies_count))
            except ZeroDivisionError:
                operator_2lb = 0
            try: 
                operator_2bk = list(filter(lambda x: x['UF_CRM_1638093750742'] == '1357', edo_companies_count))
            except ZeroDivisionError:
                operator_2bk = 0
            try: 
                operator_2lt = list(filter(lambda x: x['UF_CRM_1638093750742'] == '1831', edo_companies_count))
            except ZeroDivisionError:
                operator_2lt = 0


        else:
            edo_companies_count = []

        traffic_more_than_1 = []
        paid_traffic = 0
        
        edo_elements_info = b.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '235',
            'filter': {
                'PROPERTY_1581': user_info['ID'],
                'PROPERTY_1569': year_codes[str(before_1_month_year)],
                '!PROPERTY_1579': '',
                }
        })

        edo_elements_info = list(map(lambda x: {
            'ID': x['ID'],
            'Компания': list(x['PROPERTY_1579'].values())[0],
            'Сумма пакетов по владельцу': int(list(x['PROPERTY_1573'].values())[0]),
            'Сумма для клиента': int(list(x['PROPERTY_1575'].values())[0]),
        }, edo_elements_info))

        traffic_more_than_1 = list(filter(lambda x: x['Сумма пакетов по владельцу'] > 1, edo_elements_info))

        edo_elements_paid = b.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '235',
            'filter': {
                'PROPERTY_1581': user_info['ID'],
                'PROPERTY_1569': year_codes[str(before_1_month_year)],
            }
        })

        paid_traffic = list(filter(lambda x: int(list(x['PROPERTY_1573'].values())[0]) > 0 and int(list(x['PROPERTY_1575'].values())[0]) > 0, edo_elements_paid))
        paid_traffic = sum(list(map(lambda x: int(list(x['PROPERTY_1575'].values())[0]), paid_traffic)))

        try:
            edo_companies_coverage = round((len(edo_companies_count) / len(all_its)) * 100, 2)
        except ZeroDivisionError:
            edo_companies_coverage = 0

        try:
            active_its_coverage = round((len(traffic_more_than_1) / len(all_its)) * 100, 2)
        except ZeroDivisionError:
            active_its_coverage = 0

        worksheet.append(['ЭДО', 'Всего ИТС', 'С ЭДО', '%'])
        worksheet.append(['Охват ЭДО', len(all_its), len(edo_companies_count), edo_companies_coverage]) #на последний месяц
        if len(edo_companies_count) > 0:
            if len(operator_2ae) > 0:
                worksheet.append(['', '2AE', len(operator_2ae)])
            if len(operator_2ae_doki) > 0:
                worksheet.append(['', '2AE доки', len(operator_2ae_doki)])
            if len(operator_2be) > 0:
                worksheet.append(['', '2BE', len(operator_2be)])
            if len(operator_2bm) > 0:
                worksheet.append(['', '2BM', len(operator_2bm)])
            if len(operator_2al) > 0:
                worksheet.append(['', '2AL', len(operator_2al)])
            if len(operator_2lb) > 0:
                worksheet.append(['', '2LB', len(operator_2lb)])
            if len(operator_2bk) > 0:
                worksheet.append(['', '2BK', len(operator_2bk)])
            if len(operator_2lt) > 0:
                worksheet.append(['', '2LT', len(operator_2lt)])
        worksheet.append(['Компании с трафиком больше 1', len(set(map(lambda x: x['Компания'], traffic_more_than_1)))]) #уникальные за весь год
        worksheet.append(['% активных ИТС', active_its_coverage])
        worksheet.append(['Сумма платного трафика', paid_traffic]) #за весь год
        worksheet.append([])

        #Сверка 2.0
        all_company = b.get_all('crm.company.list', {
                'select': ['UF_CRM_1735194029'],
                'filter': {
                    'ASSIGNED_BY_ID': user_info['ID'],
                }
            })
        company_sverka = list(filter(lambda x: x['UF_CRM_1735194029'] == '1', all_company))
        
        worksheet.append(['Сверка 2.0'])
        worksheet.append(['Подключено', len(company_sverka)])

        change_sheet_style(worksheet)
   
    workbook.save(report_name)

    if 'user_id' not in req:
        return

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '828809'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    b.call('im.notify.system.add', {
        'USER_ID': req['user_id'][5:],
        'MESSAGE': f'Отчет по пользователям за год сформирован. {upload_report["DETAIL_URL"]}'})
    os.remove(report_name) 


if __name__ == '__main__':
    create_employees_period_report({
        'users': 'user_1391'
    })
