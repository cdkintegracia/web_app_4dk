import time
from datetime import datetime
from datetime import timedelta
from datetime import timezone
from os import remove as os_remove
import base64

from fast_bitrix24 import Bitrix
import dateutil.parser
import openpyxl
from openpyxl.utils import get_column_letter

from web_app_4dk.modules.authentication import authentication


# Считывание файла authentication.txt

webhook = authentication('Bitrix')
b = Bitrix(webhook)


def create_companies_activity_report(req):
    deal_type_names = {
            'UC_HT9G9H': 'ПРОФ Земля',
            'UC_XIYCTV': 'ПРОФ Земля+Помощник',
            'UC_N113M9': 'ПРОФ Земля+Облако',
            'UC_5T4MAW': 'ПРОФ Земля+Облако+Помощник',
            'UC_ZKPT1B': 'ПРОФ Облако',
            'UC_2SJOEJ': 'ПРОФ Облако+Помощник',
            'UC_81T8ZR': 'АОВ',
            'UC_SV60SP': 'АОВ+Облако',
            'UC_92H9MN': 'Индивидуальный',
            'UC_7V8HWF': 'Индивидуальный+Облако',
            'UC_34QFP9': 'Уникс',
    }
    deals = b.get_all('crm.deal.list', {
        'select': ['COMPANY_ID', 'TYPE_ID'],
        'filter': {
            'CATEGORY_ID': '1',
            'TYPE_ID': [
                'UC_HT9G9H',    # ПРОФ Земля
                'UC_XIYCTV',    # ПРОФ Земля+Помощник
                'UC_N113M9',    # ПРОФ Земля+Облако
                'UC_5T4MAW',    # ПРОФ Земля+Облако+Помощник
                'UC_ZKPT1B',    # ПРОФ Облако
                'UC_2SJOEJ',    # ПРОФ Облако+Помощник
                'UC_81T8ZR',    # АОВ
                'UC_SV60SP',    # АОВ+Облако
                'UC_92H9MN',    # Индивидуальный
                'UC_7V8HWF',    # Индивидуальный+Облако
                'UC_34QFP9',    # Уникс
            ],
            'STAGE_ID': [
                'C1:NEW',           # Услуга активна
                'C1:UC_0KJKTY',     # Счет сформирован
                'C1:UC_3J0IH6',     # Счет отправлен клиенту
                'C1:UC_KZSOR2',     # Нет оплаты
                'C1:UC_VQ5HJD',     # Ждём решения клиента
            ]
        }})
    company_id_list = list(set(map(lambda x: x['COMPANY_ID'], deals)))
    companies = b.get_all('crm.company.list', {
        'select': ['TITLE', 'ASSIGNED_BY_ID', 'DATE_CREATE'],
        'filter': {'ID': company_id_list}
    })
    users_info = b.get_all('user.get')
    date_now = datetime.now()
    date_check = date_now - timedelta(int(req['days']))
    date_check_string = datetime.strftime(date_check, '%Y-%m-%d') + 'T00:00:00+03:00'
    data_to_write = [
        [
          'Отчет по активности компаний',
          '',
          'Сформирован:',
          datetime.strftime(date_now, '%H:%M:%S %d:%m:%Y')
        ],
        [
        'Компания',
        'Ответственный за компанию',
        'Тип сделки',
        'Дней без взаимодействия',
        'Дата создания компании'
        ]
    ]
    counter = 0
    for company in companies:
        counter += 1
        print(f'{counter} | {len(companies)}')
        contacts = b.get_all('crm.company.contact.items.get', {'id': company['ID']})
        contact_id_list = list(map(lambda x: x['CONTACT_ID'], contacts))
        calls = b.get_all('voximplant.statistic.get', {
            'filter': {
                'CRM_ENTITY_TYPE': 'CONTACT',
                'CRM_ENTITY_ID': contact_id_list,
                '>=CALL_START_DATE': date_check_string,
                'CALL_FAILED_CODE': '200',
            }})
        if not calls:
            emails = b.get_all('crm.activity.list', {
                'filter': {
                    "OWNER_TYPE_ID": '3',
                    "OWNER_ID": contact_id_list,
                    'PROVIDER_TYPE_ID': ['EMAIL'],
                    '>=CREATED': date_check_string,
                }})
            if not emails:
                old_calls = b.get_all('voximplant.statistic.get', {
                    'filter': {
                        'CRM_ENTITY_TYPE': 'CONTACT',
                        'CRM_ENTITY_ID': contact_id_list,
                        'CALL_FAILED_CODE': '200',
                    }})
                old_emails = b.get_all('crm.activity.list', {
                    'filter': {
                        "OWNER_TYPE_ID": '3',
                        "OWNER_ID": contact_id_list,
                        'PROVIDER_TYPE_ID': ['EMAIL'],
                    }})
                last_activity = ''
                if old_calls:
                    old_calls = list(sorted(old_calls, key=lambda x: dateutil.parser.isoparse(x['CALL_START_DATE'])))
                else:
                    calls_activities = b.get_all('crm.activity.list', {'filter': {
                        "OWNER_TYPE_ID": '3',
                        "OWNER_ID": contact_id_list,
                        'PROVIDER_TYPE_ID': ['CALL']
                    }})
                    if calls_activities:
                        calls_activities = list(map(lambda x: x['ID'], calls_activities))
                        old_calls = b.get_all('voximplant.statistic.get', {
                    'filter': {
                        'CRM_ACTIVITY_ID': calls_activities,
                        'CALL_FAILED_CODE': '200',
                    }})
                        old_calls = list(sorted(old_calls, key=lambda x: dateutil.parser.isoparse(x['CALL_START_DATE'])))
                if old_emails:
                    old_emails = list(sorted(old_emails, key=lambda x: dateutil.parser.isoparse(x['CREATED'])))
                if old_emails and old_calls:
                    if dateutil.parser.isoparse(old_emails[-1]['CREATED']) > dateutil.parser.isoparse(old_calls[-1]['CALL_START_DATE']):
                        last_activity = dateutil.parser.isoparse(old_emails[-1]['CREATED'])
                    else:
                        last_activity = dateutil.parser.isoparse(old_calls[-1]['CALL_START_DATE'])
                elif old_emails:
                    last_activity = dateutil.parser.isoparse(old_emails[-1]['CREATED'])
                elif old_calls:
                    last_activity = dateutil.parser.isoparse(old_calls[-1]['CALL_START_DATE'])

                company_deals = list(filter(lambda x: x['COMPANY_ID'] == company['ID'], deals))
                user_info = list(filter(lambda x: x['ID'] == company['ASSIGNED_BY_ID'], users_info))[0]
                user_name = f"{user_info['NAME']} {user_info['LAST_NAME']}"
                if last_activity:
                    last_activity = (datetime.now(timezone.utc) - last_activity).days
                data_to_write.append(
                    [
                     company['TITLE'],
                     user_name,
                     deal_type_names[company_deals[0]['TYPE_ID']],
                     last_activity,
                     datetime.strftime(dateutil.parser.isoparse(company['DATE_CREATE']), '%d.%m.%Y')
                    ]
                )
        time.sleep(1)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for data in data_to_write:
        worksheet.append(data)
    for idx, col in enumerate(worksheet.columns, 1):
        worksheet.column_dimensions[get_column_letter(idx)].auto_size = True
    report_created_time = datetime.now()
    report_name_time = report_created_time.strftime('%d-%m-%Y %H %M %S %f')
    report_name = f'Отчет по активности компаний {report_name_time}.xlsx'.replace(' ', '_')
    workbook.save(report_name)

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '218069'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    b.call('im.notify.system.add', {
        'USER_ID': req['user_id'][5:],
        'MESSAGE': f'Отчет по активности компаний сформирован. {upload_report["DETAIL_URL"]}'})
    os_remove(report_name)
