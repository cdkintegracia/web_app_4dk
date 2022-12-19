from datetime import datetime, timedelta

from fast_bitrix24 import Bitrix
import openpyxl

from web_app_4dk.modules.authentication import authentication


b = Bitrix(authentication('Bitrix'))

def read_xlsx_file() -> dict:
    workbook = openpyxl.load_workbook('График ЧДИ.xlsx')
    worksheet = workbook.active
    max_rows = worksheet.max_row
    max_columns = worksheet.max_column
    file_data = {}
    fio_column = 0
    days_column = 0
    date_column = 0
    fio = ''
    for row in range(1, max_rows + 1):
        if fio_column:
            fio = worksheet.cell(row, fio_column).value
            if fio and not fio.isdigit() and fio != 'Фамилия, имя, отчество':
                file_data.setdefault(fio, [])
            else:
                fio = ''
        temp = {}
        for column in range(1, max_columns + 1):
            cell_value = worksheet.cell(row, column).value
            if cell_value == 'Фамилия, имя, отчество':
                fio_column = column
            elif cell_value == 'Количество календарных дней':
                days_column = column
            elif cell_value == f"запланиро-\nванная":
                date_column = column

            if fio:
                if column == days_column:
                    temp.setdefault('Количество календарных дней', cell_value)
                elif column == date_column:
                    temp.setdefault('Запланированная', cell_value)
        if temp:
            file_data[fio].append(temp)

    return file_data


def find_employee_id(users, fio):
    fio_lst = fio.split()
    last_name = fio_lst[0]
    name = fio_lst[1]
    user_info = list(filter(lambda x: x['NAME'] == name and x['LAST_NAME'] == last_name, users))
    if user_info:
        return user_info[0]['ID']
    b.call('tasks.task.add', {
        'fields': {
            'TITLE': f'Не удалось создать отпуск для {fio}',
            'RESPONSIBLE_ID': '173',
            'GROUP_ID': '13',
            'CREATED_BY': '173'
        }})


def create_vacation():
    file_data = read_xlsx_file()
    users = b.get_all('user.get')
    for fio in file_data:
        employee_id = find_employee_id(users, fio)
        if not employee_id:
            continue
        for vacation in file_data[fio]:
            date = datetime.strptime(vacation['Запланированная'], '%d.%m.%Y')
            start_date = datetime.strftime(date, '%Y-%m-%d')
            days = vacation['Количество календарных дней']
            end_date = date + timedelta(days=int(days) - 1)
            end_date = datetime.strftime(end_date, '%Y-%m-%d')

            b.call('bizproc.workflow.start', {
                'TEMPLATE_ID': '883',
                'DOCUMENT_ID': ['lists', 'BizprocDocument', '82689'],
                'PARAMETERS': {
                    'employee': f'user_{employee_id}',
                    'Parameter1': start_date,
                    'Parameter2': end_date,
                    'old_start': '',
                    'old_end': '',
                }
            })


create_vacation()