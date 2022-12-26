from datetime import datetime, timedelta

from fast_bitrix24 import Bitrix
import openpyxl


b = Bitrix('https://vc4dk.bitrix24.ru/rest/311/wkq0a0mvsvfmoseo/')

service_deal_current_month = ['Контрагент', 'Линк', 'МДЛП', 'СтартЭДО']


def read_deals_from_xlsx(filename):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    max_rows = worksheet.max_row
    max_columns = worksheet.max_column
    data = []
    titles = {}
    for row in range(1, max_rows + 1):
        temp = {}
        for column in range(1, max_columns):
            cell_value = worksheet.cell(row, column).value
            if row == 1:
                titles.setdefault(column, cell_value)
            else:
                temp.setdefault(titles[column], cell_value)
        if temp:
            data.append(temp)
    return data


def get_service_deal_start_dates(month, deal_type, deal_date_end):
    month_number = {
        'Январь': '01',
        'Февраль': '02',
        'Март': '03',
        'Апрель': '04',
        'Май': '05',
        'Июнь': '06',
        'Июль': '07',
        'Август': '08',
        'Сентябрь': '09',
        'Октябрь': '10',
        'Ноябрь': '11',
        'Декабрь': '12',
    }
    current_year = datetime.now().year
    if deal_type in service_deal_current_month:
        return f'{month_number[month]}.{current_year}'
    else:
        deal_date_start_year = deal_date_end.year - 1
        deal_date_start = deal_date_end + timedelta(days=1)
        return f"{deal_date_start.strftime('%d.%m')}.{deal_date_start_year}"


def deal_info_handler(deals_info, users_info, month):
    department_names = {29: 'ГО4', 27: 'ГО3', 5: 'ЦС', 231: 'ЛК', 225: 'ОВ', 1: '4DK'}
    ignore_names = ['Администратор', 'Дежурный администратор', 'Иван Иванов', 'Максим Карпов', 'Борис Ишкин']
    service_deal_types = ['Контрагент', 'Кабинет сотрудника', 'Линк', 'МДЛП', '1Спарк 3000', '1Спарк ПЛЮС 22500', 'ЭДО',
                          'СтартЭДО', 'Подпись', 'Подпись 1000', 'Кабинет садовода', 'РПД', 'ЭТП']
    service_deal_values = {'Контрагент': 4800, '1Спарк': 3000, '1Спарк ПЛЮС': 22500, 'СтартЭДО': 3000,
                           'Подпись 1000': 600, 'Кабинет садовода': 1000, '1Спарк 3000': 3000, 'Подпись': 0}
    its_deal_field = f'{month} ИТС'
    service_deal_field = f'{month} Сервисы'
    result = {}
    for deal_info in deals_info:
        deal_info['Сумма'] = int(float(deal_info['Сумма']))
        if deal_info['Ответственный'] in ignore_names or not deal_info['Ответственный']:
            continue
        if deal_info['Ответственный'] not in result and deal_info['Ответственный']:
            result.setdefault(deal_info['Ответственный'], {})
            user_name = deal_info['Ответственный'].split()[0]
            user_last_name = deal_info['Ответственный'].split()[1]
            user_info = list(filter(lambda x: x['NAME'] == user_name and x['LAST_NAME'] == user_last_name, users_info))
            if user_info[0]['UF_DEPARTMENT'][0] not in department_names:
                continue
            result[deal_info['Ответственный']].setdefault('Подразделение', department_names[user_info[0]['UF_DEPARTMENT'][0]])
            result[deal_info['Ответственный']].setdefault(its_deal_field, 0)
            result[deal_info['Ответственный']].setdefault(f'{month} Сервисы', 0)

        if deal_info['Группа'] == 'ИТС' and deal_info['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту']:
            result[deal_info['Ответственный']][its_deal_field] += 1
        elif deal_info['Тип'] in service_deal_types and deal_info['Стадия сделки'] == 'Услуга активна':
            deal_start_date = get_service_deal_start_dates(month, deal_info['Тип'], deal_info['Предполагаемая дата закрытия'])
            if deal_info['Тип'] in service_deal_current_month:
                if deal_start_date in deal_info['Дата начала'].strftime('%d.%m.%Y'):
                    if deal_info['Сумма']:
                        result[deal_info['Ответственный']][service_deal_field] += deal_info['Сумма']
                    else:
                        result[deal_info['Ответственный']][service_deal_field] += service_deal_values[deal_info['Тип']]
            else:
                if deal_start_date == deal_info['Дата начала'].strftime('%d.%m.%Y'):
                    if deal_info['Сумма']:
                        result[deal_info['Ответственный']][service_deal_field] += deal_info['Сумма']
                    else:
                        result[deal_info['Ответственный']][service_deal_field] += service_deal_values[deal_info['Тип']]

    print(result)


def create_report_service_sales(filename):
    file_data = read_deals_from_xlsx(filename)
    users_data = b.get_all('user.get')
    handled_data = deal_info_handler(file_data, users_data, 'Июль')
    print(handled_data)



create_report_service_sales('DEAL_20220731_6d490055_62e6c5642f727TONIGHT.xlsx')
