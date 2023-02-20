from fast_bitrix24 import Bitrix
import openpyxl

from authentication import authentication


b = Bitrix(authentication('Bitrix'))

deals = b.get_all('crm.deal.list', {'filter': {'CATEGORY_ID': '1'}})
workbook = openpyxl.load_workbook('Январь_2023.xlsx')
worksheet = workbook.active
max_rows = worksheet.max_row
max_columns = worksheet.max_column
new_data = []
for row in range(1, max_rows + 1):
    new_row = []
    for col in range(1, max_columns + 1):
        cell_value = worksheet.cell(row, col).value
        new_row.append(cell_value)
        if col == 8 and row != 1:
            deal_id = cell_value
            deal_info = list(filter(lambda x: str(x['ID']) == str(deal_id), deals))
            if deal_info:
                deal_name = deal_info[0]['TITLE']
            else:
                deal_name = ''
            new_row.append(deal_name)
    new_data.append(new_row)

workbook = openpyxl.Workbook()
worksheet = workbook.active
for row in new_data:
    worksheet.append(row)
workbook.save('text.xlsx')
