from fast_bitrix24 import Bitrix
import openpyxl

from web_app_4dk.modules.authentication import authentication

b = Bitrix(authentication('Bitrix'))


def read_xlsx_file(file_name) -> list:
    result = []
    workbook = openpyxl.load_workbook(file_name)
    worklist = workbook.active
    max_rows = worklist.max_row
    max_columns = worklist.max_column
    titles = {}
    for row in range(1, max_rows + 1):
        for column in range(1, max_columns + 1):
            if row == 1:
                titles.setdefault(worklist.cell(row, column).value, column)
            else:
                if column == titles['Компания']:
                    result.append(worklist.cell(row, column).value)
    return result


def change_responsible(new_responsible, file_name):
    company_names = read_xlsx_file(file_name)
    company_names = ['Демонстрация Тест']
    companies_info = b.get_all('crm.company.list', {'filter': {'TITLE': company_names}})
    for company in companies_info:
        b.call('crm.company.update', {'ID': company['ID'], 'fields': {'ASSIGNED_BY_ID': new_responsible}})
        company_contacts = b.get_all('crm.company.contact.items.get', {'id': company['ID']})
        for contact in company_contacts:
           contact_companies = b.get_all('crm.contact.company.items.get', {'id': contact['CONTACT_ID']})
           if len(contact_companies) == 1:
               b.call('crm.contact.update', {'id': contact['CONTACT_ID'], 'fields': {'ASSIGNED_BY_ID': new_responsible}})


        company_deals = b.get_all('crm.deal.list', {
            'filter': {
                'COMPANY_ID': company['ID'],
                'STAGE_ID': [
                    'C1:NEW',
                    'C1:UC_0KJKTY',
                    'C1:UC_3J0IH6',
                ],
            }})
        for deal in company_deals:
            if deal['ASSIGNED_BY_ID'] != new_responsible:
                b.call('crm.deal.update', {'id': deal['ID'], 'fields': {'ASSIGNED_BY_ID': new_responsible}})
