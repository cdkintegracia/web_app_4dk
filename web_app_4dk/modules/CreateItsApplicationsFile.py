import re
from datetime import datetime, timedelta
import base64
from os import remove as os_remove

from fast_bitrix24 import Bitrix
import openpyxl

from web_app_4dk.modules.authentication import authentication
from web_app_4dk.tools import send_bitrix_request


b = Bitrix(authentication('Bitrix'))


def create_its_applications_file(req):
    UF_CRM_1643800749 = {
        'create': '371',
        'reject': '477'
    }
    deal_filter = {
        'UF_CRM_1643800749': UF_CRM_1643800749[req['process']],
        'CATEGORY_ID': '1'
    }
    if req['process'] == 'reject':
        deal_filter['UF_CRM_1637933869479'] = '0'

    deal_fields = b.get_all('crm.deal.fields')
    deals = b.get_all('crm.deal.list', {
        'select': ['*', 'UF_*'],
        'filter': deal_filter
    })
    if not deals:
        b.call('im.notify.system.add', {
            'USER_ID': req['user_id'][5:],
            'MESSAGE': f'Нет подходящих сделок для формирования файла с заявками на подписки.'})
        return
    companies = b.get_all('crm.company.list', {
        'select': ['*', 'UF_*', 'PHONE'],
        'filter': {
            'ID': list(map(lambda x: x['COMPANY_ID'], deals))
        }
    })
    data_to_write = []
    for index, deal in enumerate(deals, 1):
        deal_date_start = datetime.fromisoformat(deal['BEGINDATE']).strftime('%d.%m.%Y')
        product_row = send_bitrix_request('crm.deal.productrows.get', {
            'id': deal['ID'],
        })
        if not product_row:
            continue
        product_info = send_bitrix_request('crm.product.get.json', {
            'id': product_row[0]['PRODUCT_ID']
        })
        try:
            code_1c = product_info['PROPERTY_139']['value']
        except:
            continue
        subscription_period = int(deal['UF_CRM_1638100416'])
        if req['process'] == 'reject':
            subscription_period = 12
        elif code_1c in ['2001', '2004'] and deal['UF_CRM_1637933869479'] == '1':
            subscription_period = 12
        payment_method = list(filter(lambda x: x['ID'] == deal['UF_CRM_1642775558379'],
                                     deal_fields['UF_CRM_1642775558379']['items']))[0]['VALUE']
        company_requisite = send_bitrix_request('crm.requisite.list', {
            'select': ['*', 'UF_*'],
            'filter': {
                'ENTITY_TYPE_ID': '4',
                'ENTITY_ID': deal['COMPANY_ID']
            }
        })[0]
        company_info = list(filter(lambda x: x['ID'] == deal['COMPANY_ID'], companies))[0]
        company_name = re.match(r'.+ \d+', company_info['TITLE']).group()
        if 'PHONE' in company_info and company_info['PHONE']:
            company_phone_code = company_info['PHONE'][0]['VALUE'].replace('-', '').replace('+', '')[1:4]
            company_phone = company_info['PHONE'][0]['VALUE'].replace('-', '').replace('+', '')[4:]
        else:
            company_phone_code = '812'
            company_phone = '334-44-74'
        try:
            address_info = b.get_all('crm.address.list', {
                'filter': {
                    'ENTITY_TYPE_ID': '4',
                    'ENTITY_ID': company_info['ID']
                }
            })[0]
            company_city = address_info['CITY'] if address_info['CITY'] else 'Санкт-Петербург'
        except:
            company_city = 'Санкт-Петербург'
            
        contacts = send_bitrix_request('crm.company.contact.items.get', {
            'id': company_info['ID']
        })
        if contacts:
            contact_info = send_bitrix_request('crm.contact.get', {
                'ID': contacts[0]['CONTACT_ID']
            })
            responsible_name = f"{contact_info['LAST_NAME']} " \
                               f"{contact_info['NAME']} " \
                               f"{contact_info['SECOND_NAME']}".replace('None', '').strip()
        else:
            responsible_name = 'Иванов Иван Иванович'
        print(deal['UF_CRM_1638958630625'])

        table_data = {
            'create': {
                'operation_24': '0 - новая',
                'reject_date_25': '',
                'reject_reason_26': '',
            },
            'reject': {
                'operation_24': '1 - отказ',
                'reject_date_25': (datetime.fromisoformat(deal['UF_CRM_1638958630625']) + timedelta(days=1)).strftime('%d.%m.%Y'),
                'reject_reason_26': '1 - прерывание договора 1С:ИТС на некоторый период, например в связи с финансовыми причинами клиента;',
            }
        }

        data_to_write.append([
            index,                              # № п/п
            '04382',                            # Код партнера
            '1',                                # Способ получения
            'p1857',                            # Код дистрибутора
            code_1c,                            # Вид 1С:ИТС
            deal['UF_CRM_1640523562691'],       # РегНомер
            company_name,                       # Наименование фирмы
            company_requisite['RQ_INN'],        # ИНН
            company_requisite['RQ_KPP'],        # КПП
            1,                                  # Количество рабочих мест
            '',                                 # Тип основной деятельности
            '',                                 # Директор
            responsible_name,                   # Ответственный
            '',                                 # Почт. индекс
            company_city,                       # Город
            '',                                 # Улица
            '',                                 # Дом
            '',                                 # Корпус
            '',                                 # Квартира
            company_phone_code,                 # Код
            company_phone,                      # Телефон
            '',                                 # Факс
            '',                                 # E-mail
            table_data[req['process']]['operation_24'],        # Операция (новый 1С:ИТС / продление / отказ от действующего 1С:ИТС)
            table_data[req['process']]['reject_date_25'],      # Дата Отказа мм.гг
            table_data[req['process']]['reject_reason_26'],    # Причина отказа от действующей регистрации 1С:ИТС (см. комментарий)
            deal_date_start,                    # Дата начала мм.гг
            subscription_period,                # Количество выпусков
            payment_method,                     # Способ оплаты
        ])

        if req['process'] == 'create':
            b.call('crm.deal.update', {
                'ID': deal['ID'],
                'fields': {
                    'UF_CRM_1643800749': '373'
                }
            })
        elif req['process'] == 'reject':
            b.call('crm.deal.update', {
                'ID': deal['ID'],
                'fields': {
                    'UF_CRM_1643800749': '479',
                    'CLOSEDATE': deal['UF_CRM_1638958630625'],
                    'UF_CRM_1638958630625': '',
                }
            })
    try:
        workbook = openpyxl.load_workbook('/root/web_app_4dk/web_app_4dk/modules/Шаблон заявок ИТС.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.load_workbook('Шаблон заявок ИТС.xlsx')
    worklist = workbook.active
    for row, row_data in enumerate(data_to_write, 11):
        for col, cell_value in enumerate(row_data, 1):
            worklist.cell(row=row, column=col).value = cell_value
    filename = f'Заявки_на_подписки_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'
    workbook.save(filename)

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '568997'
    with open(filename, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': filename},
        'fileContent': report_file_base64
    })
    b.call('im.notify.system.add', {
        'USER_ID': req['user_id'][5:],
        'MESSAGE': f'Файл с заявками на подписки сформирован. {upload_report["DETAIL_URL"]}'})
    os_remove(filename)


if __name__ == '__main__':
    create_its_applications_file({'user_id': 'user_311', 'process': 'reject'})