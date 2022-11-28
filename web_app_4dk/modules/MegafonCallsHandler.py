from datetime import datetime, timedelta
import base64
from os import remove as os_remove

from fast_bitrix24 import Bitrix
import openpyxl

from web_app_4dk.modules.authentication import authentication


b = Bitrix(authentication('Bitrix'))

report_created_time = datetime.now()
report_name_time = report_created_time.strftime('%d-%m-%Y %H %M %S %f')
report_name = f'Отчет Мегафон {report_name_time}.xlsx'.replace(' ', '_')

lk_employers = []
adm_employers = []
cs_employers = []
b24_users = b.get_all('user.get')
for user in b24_users:
    user_name = f'{user["LAST_NAME"]} {user["NAME"]}'
    for department in user['UF_DEPARTMENT']:
        if department == 5:
            adm_employers.append(user_name)
        elif department == 231 and user_name != 'Ридкобород Светлана':
            lk_employers.append(user_name)
        elif department in [27, 29]:
            cs_employers.append(user_name)

b24_contacts_info = b.get_all('crm.contact.list', {'select': ['PHONE']})
b24_companies_info = b.get_all('crm.company.list', {'select': ['PHONE', 'TITLE']})
b24_deals = b.get_all('crm.deal.list', {
    'filter': {
        'CATEGORY_ID': '1',
        'STAGE_ID': [
            'C1:UC_0KJKTY',  # Счет сформирован
            'C1:UC_3J0IH6',  # Счет отправлен клиенту
            'C1:UC_KZSOR2',  # Нет оплаты
            'C1:UC_VQ5HJD',  # Ждём решения клиента
            'C1:NEW',  # Услуга активна
        ]}})


def parse_phone_numbers(phones_info: list) -> list:
    phone_numbers = list(map(lambda x: x['VALUE'], phones_info))
    result = []
    for phone_number in phone_numbers:
        if '+7' in phone_number:
            result.append(phone_number[1:12])
        elif phone_number[0] == '8':
            result.append('7' + phone_number[1:12])
        else:
            result.append(phone_number[:11])
    return result


def get_company_names(company_id: list) -> list:
    result = []
    for i in range(len(company_id)):
        company_info = list(filter(lambda x: x['ID'] == str(company_id[i]), b24_companies_info))[0]
        result.append(company_info['TITLE'])
    return result


def read_megafon_file(file_name, key='data'):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active
    max_rows = worksheet.max_row
    max_columns = worksheet.max_column
    worksheet_titles = {}
    megafon_data = []

    for row in range(10, max_rows + 1):
        temp = []
        for column in range(1, max_columns + 1):
            if row == 10:
                worksheet_titles.setdefault(worksheet.cell(row, column).value, column - 1)
            else:
                temp.append(worksheet.cell(row, column).value)
        if temp:
            megafon_data.append(temp)
    if key == 'data':
        return list(filter(lambda x: x[0] == 'исходящий', megafon_data))
    elif key == 'titles':
        return worksheet_titles


def b24_companies_handler(b24_companies) -> list:
    for i in range(len(b24_companies)):
        if 'PHONE' in b24_companies[i]:
            b24_companies[i]['PHONE'] = parse_phone_numbers(b24_companies[i]['PHONE'])
    return b24_companies


def b24_contacts_handler(b24_contacts) -> list:
    count = 0
    for i in range(len(b24_contacts)):
        companies = b.get_all('crm.contact.company.items.get', {'id': b24_contacts[i]['ID']})
        companies = list(map(lambda x: x['COMPANY_ID'], companies))
        b24_contacts[i]['COMPANIES'] = get_company_names(companies)
        if 'PHONE' in b24_contacts[i]:
            b24_contacts[i]['PHONE'] = parse_phone_numbers(b24_contacts[i]['PHONE'])
        count += 1
        print(f"Обработка контактов: {count} | {len(b24_contacts)}")
    return b24_contacts


def department_call_assigment(employee: str, duration: datetime.time) -> dict:
    seconds = duration.second
    minutes = duration.minute
    hours = duration.hour
    duration = timedelta(hours=hours, minutes=minutes, seconds=seconds)
    if employee in lk_employers:
        return {'ЛК минут': duration, 'ЛК количество': 1}
    elif employee in cs_employers:
        return {'ЦС минут': duration, 'ЦС количество': 1}
    elif employee in adm_employers:
        return {'АДМ минут': duration, 'АДМ количество': 1}


def find_call_company(data: list, titles: dict, b24_contacts: list, b24_companies) -> dict:
    result = {}
    client_phone = data[titles['Клиент']]
    clients = list(filter(lambda x: client_phone in x['PHONE'], b24_contacts))
    if clients:
        for client in clients:
            for company in client['COMPANIES']:
                if company not in result:
                    result.setdefault(company, department_call_assigment(data[titles['Сотрудник']],
                                                                         data[titles['Длительность']]
                                                                         ))
    else:
        client_companies = list(filter(lambda x: client_phone in x['PHONE'], b24_companies))
        if client_companies:
            for client_company in client_companies:
                if client_company['TITLE'] not in result:
                    result.setdefault(client_company['TITLE'], department_call_assigment(data[titles['Сотрудник']],
                                                                                         data[titles['Длительность']]
                                                                                         ))
    return result


def find_top_deal(deals: list) -> str:
    level_1 = ['UC_HT9G9H',  # ПРОФ Земля
               'UC_XIYCTV',  # ПРОФ Земля+Помощник
               'UC_N113M9',  # ПРОФ Земля+Облако
               'UC_5T4MAW',  # ПРОФ Земля+Облако+Помощник
               'UC_ZKPT1B',  # ПРОФ Облако
               'UC_2SJOEJ',  # ПРОФ Облако+Помощник
               'UC_81T8ZR',  # АОВ
               'UC_SV60SP',  # АОВ+Облако
               'UC_92H9MN',  # Индивидуальный
               'UC_7V8HWF',  # Индивидуальный+Облако
               'UC_34QFP9',  # Уникс
               ]
    level_2 = ['UC_AVBW73',  # Базовый Земля
               'UC_GPT391',  # Базовый Облако
               'UC_1UPOTU',  # ИТС Бесплатный
               'UC_K9QJDV',  # ГРМ Бизнес
               'GOODS',  # ГРМ
               'UC_J426ZW',  # Садовод
               'UC_DBLSP5',  # Садовод+Помощник
               ]
    level_3 = ['UC_IUJR81',  # Допы Облако
               'UC_USDKKM',  # Медицина
               'UC_BZYY0D',  # ИТС Отраслевой
               ]
    level_4 = ['UC_2R01AE',  # Услуги (без нашего ИТС)
               'UC_IV3HX1',  # Тестовый
               'UC_YIAJC8',  # Лицензия с купоном ИТС
               'UC_QQPYF0',  # Лицензия
               'UC_O99QUW',  # Отчетность
               'UC_OV4T7K',  # Отчетность (в рамках ИТС)
               'UC_2B0CK2',  # 1Спарк в договоре
               'UC_86JXH1',  # 1Спарк 3000
               'UC_WUGAZ7',  # 1СпаркРиски ПЛЮС 22500
               'UC_A7G0AM',  # Контрагент
               'UC_GZFC63',  # РПД
               'UC_8Z4N1O',  # Подпись
               'UC_FOKY52',  # Подпись 1000
               'UC_D1DN7U',  # Кабинет сотрудника
               'UC_H8S037',  # ЭДО
               'UC_66Z1ZF',  # ОФД
               'UC_40Q6MC',  # СтартЭДО
               'UC_8LW09Y',  # МДЛП
               'UC_3SKJ5M',  # 1С Касса
               'UC_4B5UQD',  # ЭТП
               'UC_H7HOD0',  # Коннект
               'UC_XJFZN4',  # Кабинет садовода
               'UC_74DPBQ',  # БИТРИКС24
               ]
    deal_type_names = {
        'SALE': 'ИТС Земля',
        'COMPLEX': 'СааС',
        'UC_APUWEW': 'ИТС Земля + СааС',
        'UC_1UPOTU': 'ИТС Бесплатный',
        'UC_O99QUW': 'Отчетность',
        'UC_OV4T7K': 'Отчетность (в рамках ИТС)',
        'UC_2B0CK2': '1Спарк в договоре',
        'UC_86JXH1': '1Спарк 3000',
        'UC_WUGAZ7': '1СпаркПЛЮС 22500',
        'UC_A7G0AM': '1С Контрагент',
        'GOODS': 'ГРМ',
        'UC_GZFC63': 'РПД',
        'UC_QQPYF0': 'Лицензия',
        'UC_8Z4N1O': '1С-Подпись',
        'UC_FOKY52': '1С-Подпись 1000',
        'UC_D1DN7U': '1С Кабинет сотрудника',
        'UC_34QFP9': 'Уникс',
        'UC_J426ZW': '1С Садовод',
        'UC_H8S037': 'ЭДО',
        'UC_8LW09Y': 'МДЛП',
        'UC_3SKJ5M': '1С Касса',
        'UC_4B5UQD': 'ЭТП',
        'UC_H7HOD0': 'Коннект',
        'UC_USDKKM': 'Медицина',
        'SERVICE': 'Сервисное обслуживание',
        'SERVICES': 'Услуги',
        'UC_XJFZN4': 'Кабинет садовода',
        'UC_BZYY0D': 'ИТС Отраслевой',
        'UC_66Z1ZF': 'ОФД',
        'UC_40Q6MC': 'Старт ЭДО',
        'UC_74DPBQ': 'Битрикс24',
        'UC_IV3HX1': 'Тестовый',
        'UC_HT9G9H': 'ПРОФ Земля',
        'UC_XIYCTV': 'ПРОФ Земля+Помощник',
        'UC_5T4MAW': 'ПРОФ Земля+Облако+Помощник',
        'UC_N113M9': 'ПРОФ Земля+Облако',
        'UC_ZKPT1B': 'ПРОФ Облако',
        'UC_2SJOEJ': 'ПРОФ Облако+Помощник',
        'UC_AVBW73': 'Базовый Земля',
        'UC_GPT391': 'Базовый Облако',
        'UC_92H9MN': 'Индивидуальный',
        'UC_7V8HWF': 'Индивидуальный+Облако',
        'UC_IUJR81': 'Допы Облако',
        'UC_2R01AE': 'Услуги (без нашего ИТС)',
        'UC_81T8ZR': 'АОВ',
        'UC_SV60SP': 'АОВ+Облако',
        'UC_D7TC4I': 'ГРМ Спец',
        'UC_K9QJDV': 'ГРМ Бизнес',
        'UC_DBLSP5': 'Садовод+Помощник',
        'UC_GP5FR3': 'ДУО',
        'UC_YIAJC8': 'Лицензия с купоном ИТС',
        '1': 'Не указан',
    }

    for type in level_1:
        if type in deals:
            return deal_type_names[type]
    for type in level_2:
        if type in deals:
            return deal_type_names[type]
    for type in level_3:
        if type in deals:
            return deal_type_names[type]
    for type in level_4:
        if type in deals:
            return deal_type_names[type]
    return ''


def write_data_to_file(data_to_write, file_name):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for data in data_to_write:
        worksheet.append(data)
    workbook.save(file_name)


def megafon_calls_handler(file_name):
    return
    megafon_data = read_megafon_file(key='data', file_name=file_name)
    megafon_titles = read_megafon_file(file_name, 'titles')
    b24_companies = list(filter(lambda x: 'PHONE' in x, b24_companies_handler(b24_companies_info)))
    b24_contacts = list(filter(lambda x: 'PHONE' in x, b24_contacts_handler(b24_contacts_info)))
    result_titles = [
        'Компания', 'Топ тип', 'ЛК минут', 'ЛК количество', 'ЦС минут',
        'ЦС количество', 'АДМ минут', 'АДМ количество', 'Длительность', 'Количество'
    ]

    result = {}
    for data in megafon_data:
        processed_data = find_call_company(data, megafon_titles, b24_contacts, b24_companies)
        if not processed_data:
            continue
        for company_name in processed_data:
            if not processed_data[company_name]:
                continue
            if company_name not in result:
                company_id = list(filter(lambda x: x['TITLE'] == company_name, b24_companies_info))[0]['ID']
                company_deals = list(filter(lambda x: x['COMPANY_ID'] == company_id, b24_deals))
                company_deal_types = list(map(lambda x: x['TYPE_ID'], company_deals))
                if not company_deals:
                    continue
                top_deal = find_top_deal(company_deal_types)
                result.setdefault(company_name, {'Топ тип': top_deal,
                                            'ЛК минут': timedelta(),
                                            'ЛК количество': 0,
                                            'ЦС минут': timedelta(),
                                            'ЦС количество': 0,
                                            'АДМ минут': timedelta(),
                                            'АДМ количество': 0,
                                            'Длительность': timedelta(),
                                            'Количество': 0})
            for key in processed_data[company_name]:
                result[company_name][key] += processed_data[company_name][key]
                if isinstance(result[company_name][key], timedelta):
                    result[company_name]['Длительность'] += processed_data[company_name][key]
            result[company_name]['Количество'] += 1

    data_to_write = [result_titles]
    for company_name in result:
        temp = [company_name]
        values = list(map(lambda x: str(x) if isinstance(x, timedelta) else x, result[company_name].values()))
        temp += values
        data_to_write.append(temp)

    write_data_to_file(data_to_write, report_name)

    # Загрузка файла в Битрикс
    bitrix_folder_id = '232345'

    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_file = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    b.call('tasks.task.add', {'fields': {
        'TITLE': f"Файл Мегафона обработан",
        'DESCRIPTION': upload_file["DETAIL_URL"],
        'GROUP_ID': '13',
        'CREATED_BY': '173',
        'RESPONSIBLE_ID': '173'
    }})
    os_remove(report_name)


