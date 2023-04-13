from datetime import datetime
import os
import base64

from fast_bitrix24 import Bitrix
import openpyxl
from openpyxl.styles import PatternFill

#from field_values import deals_category_1_types, departments_id_name, month_int_names
from web_app_4dk.modules.field_values import deals_category_1_types, departments_id_name, month_int_names
from web_app_4dk.modules.EdoInfoHandler import month_codes, year_codes
from web_app_4dk.modules.authentication import authentication
#from authentication import authentication
'''
month_codes = {
    'Январь': '2371',
    'Февраль': '2373',
    'Март': '2375',
    'Апрель': '2377',
    'Май': '2379',
    'Июнь': '2381',
    'Июль': '2383',
    'Август': '2385',
    'Сентябрь': '2387',
    'Октябрь': '2389',
    'Ноябрь': '2391',
    'Декабрь': '2393'
}

year_codes = {
    '2022': '2395',
    '2023': '2397'
}
'''

b = Bitrix(authentication('Bitrix'))


def get_folder_id(storages, user_id, folder_name):
    folder_id = None
    user_storage = list(filter(lambda x: x['ENTITY_ID'] == user_id, storages))
    if user_storage:
        user_storage = user_storage[0]
        storage_folders = b.get_all('disk.storage.getchildren', {'id': user_storage['ID']})
        for folder in storage_folders:
            if folder['NAME'] == folder_name:
                folder_id = folder['ID']
                break
        if not folder_id:
            folder_id = b.call('disk.storage.addfolder', {'id': user_storage['ID'], 'data': {'NAME': folder_name}})['ID']
    return folder_id


def color_cells(report_name):
    # Закрашивание ячеек
    cell_red_color = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    cell_number_to_letter = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K',
                             12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U'}
    workbook = openpyxl.load_workbook(report_name)
    worksheet = workbook['Компании']
    max_rows = worksheet.max_row
    max_columns = worksheet.max_column
    for row in range(2, max_rows + 1):
        for column in range(1, max_columns + 1):
            cell_value = worksheet.cell(row, column).value
            if cell_value == 'Нет':
                worksheet[f"{cell_number_to_letter[column]}{row}"].fill = cell_red_color
    workbook.save(report_name)


def upload_report_to_bitrix(users_id, report_name, bitrix_folder_id='293499'):
    # Загрузка отчета в Битрикс
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    for user_id in users_id:
        b.call('im.notify.system.add', {
            'USER_ID': user_id,
            'MESSAGE': f'Отчет по охвату сервисами сформирован. {upload_report["DETAIL_URL"]}'})
    os.remove(report_name)



def create_services_coverage_report(req):
    result_data = {}
    companies_info = b.get_all('crm.company.list')
    users_info = b.get_all('user.get')
    deals_info = b.get_all('crm.deal.list', {
        'select': [
            'TYPE_ID',
            'COMPANY_ID',
            'ASSIGNED_BY_ID',
            'TITLE',
            'UF_CRM_1640523562691',      # Регномер
            'UF_CRM_1657878818384',      # Группа
            'UF_CRM_1640523703',         # Подразделение
            'UF_CRM_1651071211',         # Льготная отчетность
        ],
        'filter': {
            'STAGE_ID': ['C1:NEW', 'C1:UC_0KJKTY', 'C1:UC_3J0IH6'],
            '!TYPE_ID': ['GOODS', 'UC_K9QJDV', 'UC_D7TC4I', 'UC_IUJR81'],
        }})
    extra_clouds_info = b.get_all('crm.deal.list', {
        'select': [
            'COMPANY_ID',
            'TITLE',
            'UF_CRM_1640523562691',     # Регномер
        ],
        'filter': {
            'TYPE_ID': 'UC_IUJR81',
            'STAGE_ID': ['C1:NEW', 'C1:UC_0KJKTY', 'C1:UC_3J0IH6'],
        }
    })

    # Создание базового массива с ИТС
    deals_its = list(filter(lambda x: x['UF_CRM_1657878818384'] == '859', deals_info))
    for deal in deals_its:
            if deal['COMPANY_ID'] not in result_data:
                result_data.setdefault(deal['COMPANY_ID'], [])
            user_name = ''
            user_info = list(filter(lambda x: x['ID'] == deal['ASSIGNED_BY_ID'], users_info))
            if user_info:
                user_info = user_info[0]
                user_name = f"{user_info['NAME']} {user_info['LAST_NAME']}"
            company_name = ''
            if deal['COMPANY_ID']:
                company_info = list(filter(lambda x: x['ID'] == deal['COMPANY_ID'], companies_info))[0]
                company_name = company_info['TITLE']
            deal_reporting_in_its = 'Нет'
            if deal['UF_CRM_1651071211']:
                deal_reporting_in_its = 'Да'
            name_deal_type = deals_category_1_types[deal['TYPE_ID']]
            counteragent = 'Нет'
            if 'ПРОФ' in name_deal_type or 'Облако' in name_deal_type:
                counteragent = 'Да (Тариф)'
            field_values = {
                'Компания': company_name,
                'ИТС': deals_category_1_types[deal['TYPE_ID']],
                'Регномер': deal['UF_CRM_1640523562691'],
                'Ответственный за сделку ИТС': user_name,
                'Подразделение ответственного': departments_id_name[deal['UF_CRM_1640523703']],
                'Контрагент': counteragent,
                'Спарк в договоре / Спарк 3000': 'Нет',
                'Спарк Плюс': 'Нет',
                'РПД платный': '',
                'ЭДО': 0,
                'Отчетность': 0,
                'Отчетность в рамках ИТС': deal_reporting_in_its,
                'Допы Облако': '',
            }
            result_data[deal['COMPANY_ID']].append(field_values)

    # Подсчет сервисов из массива сделок
    services = [
        'UC_A7G0AM',        # Контрагент
        'UC_2B0CK2',        # 1Спарк в договоре
        'UC_86JXH1',        # 1Спарк 3000
        'UC_WUGAZ7',        # 1СпаркРиски ПЛЮС 22500
        'UC_GZFC63',        # РПД
        'UC_IUJR81',        # Допы Облако
    ]
    deals_services = list(filter(lambda x: x['TYPE_ID'] in services, deals_info))
    for service in deals_services:
        if service['COMPANY_ID']:
            if service['COMPANY_ID'] in result_data:
                for data_counter in range(len(result_data[service['COMPANY_ID']])):
                        if service['TYPE_ID'] == 'UC_A7G0AM' and result_data[service['COMPANY_ID']][data_counter]['Контрагент'] == 'Нет':
                            result_data[service['COMPANY_ID']][data_counter]['Контрагент'] = 'Да'
                        elif service['TYPE_ID'] in ['UC_2B0CK2', 'UC_86JXH1']:
                            result_data[service['COMPANY_ID']][data_counter]['Спарк в договоре / Спарк 3000'] = 'Да'
                        elif service['TYPE_ID'] == 'UC_WUGAZ7':
                            result_data[service['COMPANY_ID']][data_counter]['Спарк Плюс'] = 'Да'
                        elif service['TYPE_ID'] == 'UC_GZFC63':
                            result_data[service['COMPANY_ID']][data_counter]['РПД платный'] += f"{service['TITLE']}, "

    # Подсчет Допов Облако
    for extra_cloud in extra_clouds_info:
        if extra_cloud['COMPANY_ID']:
            if extra_cloud['COMPANY_ID'] in result_data:
                for data_counter in range(len(result_data[extra_cloud['COMPANY_ID']])):
                        result_data[extra_cloud['COMPANY_ID']][data_counter]['Допы Облако'] += f"{extra_cloud['TITLE']}, "


    # Подсчет ЭДО
    edo_year = datetime.now().year
    edo_month = datetime.now().month
    edo_info_massive = []
    for _ in range(3):
        edo_month -= 1
        if edo_month == 0:
            edo_month = 12
            edo_year -= 1
        edo_info_list = b.get_all('lists.element.get', {
            'IBLOCK_TYPE_ID': 'lists',
            'IBLOCK_ID': '235',
            'filter': {'PROPERTY_1567': month_codes[month_int_names[edo_month]],
                       'PROPERTY_1569': year_codes[str(edo_year)]
                       }})
        edo_info_massive += edo_info_list

    for edo_info in edo_info_list:
        if 'PROPERTY_1571' in edo_info:
            edo_info_company_id = list(edo_info['PROPERTY_1571'].values())[0]
            if edo_info_company_id in result_data:
                edo_info_value = int(list(edo_info['PROPERTY_1573'].values())[0])
                for data_counter in range(len(result_data[edo_info_company_id])):
                    result_data[edo_info_company_id][data_counter]['ЭДО'] = edo_info_value

    # Подсчет отчетностей
    deals_reporting_id = []
    for company_id in result_data:
        for data_counter in range(len(result_data[company_id])):
            deals_reporting = list(filter(lambda x: x['TYPE_ID'] in ['UC_O99QUW', 'UC_OV4T7K'] and x['UF_CRM_1640523562691'] == result_data[company_id][data_counter]['Регномер'], deals_info))
            for deal_reporting in deals_reporting:
                if deal_reporting['ID'] not in deals_reporting_id:
                    if deal_reporting['TYPE_ID'] == 'UC_O99QUW':
                        result_data[company_id][data_counter]['Отчетность'] += 1
                        deals_reporting_id.append(deal_reporting['ID'])
                if deal_reporting['UF_CRM_1640523562691'] != result_data[company_id][data_counter]['Регномер']:
                    deals_reporting_by_regnumber = list(filter(lambda x: x['UF_CRM_1640523562691'] == deal_reporting['UF_CRM_1640523562691'] and x['TYPE_ID'] == 'UC_O99QUW', deals_info))
                    for deal_reporting_by_regnumber in deals_reporting_by_regnumber:
                        if deal_reporting_by_regnumber['ID'] not in deals_reporting_id:
                            result_data[company_id][data_counter]['Отчетность'] += 1
                            deals_reporting_id.append(deal_reporting_by_regnumber['ID'])

    # Запись отчета
    main_report_name = f"Отчет_по_охвату_сервисами_{datetime.now().strftime('%d-%m-%Y-%f')}.xlsx"
    workbook = openpyxl.Workbook()

    # Первый лист
    worksheet = workbook.active
    worksheet.title = 'Компании'
    data_to_write = []
    first_list_titles = []
    for company_id in result_data:
        if not first_list_titles:
            first_list_titles = list(result_data[company_id][0].keys())
        for data in result_data[company_id]:
            data['РПД платный'] = data['РПД платный'].rstrip(', ')
            data['Допы Облако'] = data['Допы Облако'].rstrip(', ')
            data['Отчетность в рамках ИТС'] = 'Не предусмотрено' if data['Отчетность в рамках ИТС'] == 'Нет' and 'Базовый' in data['ИТС'] else data['Отчетность в рамках ИТС']
            data_to_write.append(list(data.values()))
            if data['Отчетность в рамках ИТС'] == 'Не предусмотрено':
                data['Отчетность в рамках ИТС'] = 'Нет'
    data_to_write = sorted(data_to_write, key=lambda x: (x[3].split()[1], x[0]))
    worksheet.append(first_list_titles)
    for row in data_to_write:
        worksheet.append(row)

    # Второй лист
    worksheet = workbook.create_sheet('% Сервисов')
    departments = ['ГО3', 'ГО4', 'ОВ', 'ЦС', 'Остальные']
    second_list_titles = ['Подразделение', 'Имя Фамилия сотрудника', 'Всего ИТС', 'Всего ПРОФ', 'ИТС без сервисов',
              '% ИТС без сервисов', 'ИТС без платных сервисов', '% ИТС без платных сервисов', 'Льготных отчетностей',
              '% Льготных отчетностей от ПРОФ', 'Только 1 сервис',
              '% Только 1 сервис', 'Два и более сервисов', '% Два и более сервисов',
              'Охвачено платной отчетностью', '% Охвата платной отчетностью', 'Любая отчетность', '% Любая отчетность']
    worksheet.append(second_list_titles)
    data_to_write = {}
    for company_id in result_data:
        for row in range(len(result_data[company_id])):
            employee = result_data[company_id][row]['Ответственный за сделку ИТС']
            if employee not in data_to_write:
                data_to_write.setdefault(employee, dict(zip(second_list_titles, [0 for _ in second_list_titles])))
                data_to_write[employee]['Подразделение'] = result_data[company_id][row]['Подразделение ответственного']
                data_to_write[employee]['Имя Фамилия сотрудника'] = employee

            data_to_write[employee]['Всего ИТС'] += 1
            if 'Базовый' not in result_data[company_id][row]['ИТС']:
                data_to_write[employee]['Всего ПРОФ'] += 1
            if result_data[company_id][row]['Отчетность в рамках ИТС'] != 'Нет':
                data_to_write[employee]['Льготных отчетностей'] += 1

            # Подсчет всех сервисов
            services_count = 0
            paid_services_count = 0
            if result_data[company_id][row]['Контрагент'] != 'Нет':
                services_count += 1
                if result_data[company_id][row]['Контрагент'] != 'Да (Тариф)':
                    paid_services_count += 1
            if result_data[company_id][row]['Спарк в договоре / Спарк 3000'] != 'Нет':
                services_count += 1
                paid_services_count += 1
            if result_data[company_id][row]['Спарк Плюс'] != 'Нет':
                services_count += 1
                paid_services_count += 1
            if result_data[company_id][row]['РПД платный']:
                services_count += 1
                paid_services_count += 1
            if result_data[company_id][row]['Отчетность в рамках ИТС'] != 'Нет':
                services_count += 1
            if result_data[company_id][row]['Допы Облако']:
                paid_services_count += 1
                services_count += 1
            if services_count == 0:
                data_to_write[employee]['ИТС без сервисов'] += 1
            elif services_count == 1:
                data_to_write[employee]['Только 1 сервис'] += 1
            elif services_count > 1:
                data_to_write[employee]['Два и более сервисов'] += 1
            if paid_services_count == 0:
                data_to_write[employee]['ИТС без платных сервисов'] += 1
            if result_data[company_id][row]['Отчетность']:
                data_to_write[employee]['Охвачено платной отчетностью'] += 1
            if any([result_data[company_id][row]['Отчетность'], result_data[company_id][row]['Отчетность в рамках ИТС'] not in ['Нет', 'Не предусмотрено']]):
                data_to_write[employee]['Любая отчетность'] += 1


    data_to_write_list = []
    for employee in data_to_write:
        try:
            data_to_write[employee]['% ИТС без сервисов'] = round(data_to_write[employee]['ИТС без сервисов'] / data_to_write[employee]['Всего ИТС'], 2) * 100
        except ZeroDivisionError:
            data_to_write[employee]['% ИТС без сервисов'] = 0
        try:
            data_to_write[employee]['% Льготных отчетностей от ПРОФ'] = round(data_to_write[employee]['Льготных отчетностей'] / data_to_write[employee]['Всего ПРОФ'], 2) * 100
        except ZeroDivisionError:
            data_to_write[employee]['% Льготных отчетностей от ПРОФ'] = 0
        try:
            data_to_write[employee]['% Только 1 сервис'] = round(data_to_write[employee]['Только 1 сервис'] / data_to_write[employee]['Всего ИТС'], 2) * 100
        except ZeroDivisionError:
            data_to_write[employee]['% Только 1 сервис'] = 0
        try:
            data_to_write[employee]['% Два и более сервисов'] = round(data_to_write[employee]['Два и более сервисов'] / data_to_write[employee]['Всего ИТС'], 2) * 100
        except ZeroDivisionError:
            data_to_write[employee]['% Два и более сервисов'] = 0
        try:
            data_to_write[employee]['% ИТС без платных сервисов'] = round(data_to_write[employee]['ИТС без платных сервисов'] / data_to_write[employee]['Всего ИТС'], 2) * 100
        except ZeroDivisionError:
            data_to_write[employee]['% ИТС без платных сервисов'] = 0
        try:
            data_to_write[employee]['% Охвата платной отчетностью'] = round(data_to_write[employee]['Охвачено платной отчетностью'] / data_to_write[employee]['Всего ИТС'], 2) * 100
        except ZeroDivisionError:
            data_to_write[employee]['% Охвата платной отчетностью'] = 0
        try:
            data_to_write[employee]['% Любая отчетность'] = round(data_to_write[employee]['Любая отчетность'] / data_to_write[employee]['Всего ИТС'], 2) * 100
        except ZeroDivisionError:
            data_to_write[employee]['% Любая отчетность'] = 0

        data_to_write_list.append(list(data_to_write[employee].values()))

    data_to_write_list = sorted(data_to_write_list, key=lambda x: x[1].split()[1])
    for department in departments:
        for row in data_to_write_list:
            if row[0] == department:
                worksheet.append(row)
            elif department == 'Остальные' and row[0] not in departments:
                worksheet.append(row)

    workbook.save(main_report_name)
    color_cells(main_report_name)

    if req['to_all_employees'] == 'N':
        upload_report_to_bitrix(report_name=main_report_name, users_id=[req['user_id'][5:]])
    else:
        main_report_name_users = [
            '1',
            '157',
            '311'
        ]
        upload_report_to_bitrix(report_name=main_report_name, users_id=main_report_name_users)

    if req['to_all_employees'] == 'Y':

        storages = b.get_all('disk.storage.getlist')
        folder_name = 'Отчеты_по_охвату_сервисами'
        allowed_departments = ['ГО3', 'ГО4', 'ЦС']
        departments_info = b.get_all('department.get')
        ignore_users = set()
        for department in departments_info:
            if 'UF_HEAD' in department:
                ignore_users.add(department['UF_HEAD'])

        # Рассылка отчетов сотрудникам
        for user in users_info:
            user_name = f"{user['NAME']} {user['LAST_NAME']}"
            report_name = f"Отчет_по_охвату_сервисами_{user_name}_{datetime.now().strftime('%d-%m-%Y-%f')}.xlsx".replace(' ', '_')
            workbook = openpyxl.Workbook()
    
            # Первый лист
            flag = False
            worksheet = workbook.active
            worksheet.title = 'Компании'
            worksheet.append(first_list_titles)
            for company in result_data:
                for row in result_data[company]:
                    if row['Подразделение ответственного'] not in allowed_departments:
                        continue
                    if user_name == row['Ответственный за сделку ИТС']:
                        worksheet.append(list(row.values()))
                        flag = True
            if flag:
                
                # Второй лист
                worksheet = workbook.create_sheet('% Сервисов')
                worksheet.append(second_list_titles)
                for row in data_to_write_list:
                    if user_name == row[1]:
                        worksheet.append(row)
                
                if user['ID'] in ignore_users:
                    continue
                workbook.save(report_name)
                color_cells(report_name)
                folder_id = get_folder_id(storages=storages, folder_name=folder_name, user_id=user['ID'])
                if folder_id:
                    upload_report_to_bitrix(users_id=[user['ID']], report_name=report_name, bitrix_folder_id=folder_id)

        # Рассылка руководителям
        # Первый лист
        departments_companies_data = {}
        allowed_departments = ['ГО3', 'ГО4']
        for company in result_data:
            for row in result_data[company]:
                if row['Подразделение ответственного'] in allowed_departments:
                    if row['Подразделение ответственного'] not in departments_companies_data:
                        departments_companies_data.setdefault(row['Подразделение ответственного'], [])
                    else:
                        departments_companies_data[row['Подразделение ответственного']].append(row)

        for department_name in departments_companies_data:
            report_name = f"Отчет_по_охвату_сервисами_{department_name}_{datetime.now().strftime('%d-%m-%Y-%f')}.xlsx".replace(' ', '_')
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Компании'
            worksheet.append(first_list_titles)
            for row in departments_companies_data[department_name]:
                worksheet.append(list(row.values()))

            # Второй лист
            worksheet = workbook.create_sheet('% Сервисов')
            worksheet.append(second_list_titles)
            for row in data_to_write_list:
                if row[0] in department_name:
                    worksheet.append(row)

            department_head = list(filter(lambda x: x['NAME'] == department_name, departments_info))
            if department_head:
                department_head = department_head[0]['UF_HEAD']
                folder_id = get_folder_id(storages=storages, folder_name=folder_name, user_id=department_head)
                if folder_id:
                    workbook.save(report_name)
                    color_cells(report_name)
                    upload_report_to_bitrix(users_id=[department_head], report_name=report_name, bitrix_folder_id=folder_id)


