from os import remove as os_remove

from datetime import datetime
from fast_bitrix24 import Bitrix
import openpyxl
from openpyxl.utils import get_column_letter
import base64

#from web_app_4dk.modules.authentication import authentication


# Считывание файла authentication.txt

#webhook = authentication('Bitrix')
webhook = 'https://vc4dk.bitrix24.ru/rest/311/wkq0a0mvsvfmoseo/'
b = Bitrix(webhook)

def create_report_rpd(req):
    stage_names = {
        'C13:NEW': 'Новые',
        'C13:UC_PYAKJ0': 'Отправлено письмо / сообщение',
        'C13:PREPAYMENT_INVOIC': 'Презентация (телефон / лично)',
        'C13:UC_T4D1BX': 'Не хотят слушать',
        'C13:EXECUTING': 'Тестовый пакет',
        'C13:FINAL_INVOICE': 'В рамках ИТС',
        'C13:UC_CI976X': 'Выставлен счет',
        'C13:LOSE': 'Отказ',
        'C13:WON': 'Оплачен счет',
    }
    titles = [
        [
            'Отчет по РПД',
            '',
            'Сформирован:',
            datetime.strftime(datetime.now(), "%d.%m.%Y %H:%M:%S")
        ],
        [
            'ФИО',
            'Новые',
            'Отправлено письмо / сообщение',
            'Презентация (телефон / лично)',
            'Не хотят слушать',
            'Тестовый пакет',
            'В рамках ИТС',
            'Выставлен счет',
            'Отказ',
            'Оплачен счет',
        ]
    ]
    result_dict = {}
    deals = b.get_all('crm.deal.list', {'select': ['ASSIGNED_BY_ID', 'STAGE_ID'], 'filter': {'CATEGORY_ID': '13'}})
    for deal in deals:
        worker = deal['ASSIGNED_BY_ID']
        stage = deal['STAGE_ID']
        if worker not in result_dict:
            result_dict.setdefault(worker, {stage: 1})
        else:
            if stage not in result_dict[worker]:
                result_dict[worker].setdefault(stage, 1)
            else:
                result_dict[worker][stage] += 1
    users = b.get_all('user.get')
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for title in titles:
        worksheet.append(title)
    total = ['Всего', 0, 0, 0, 0, 0, 0, 0, 0, 0]
    data_to_write = []
    for worker in result_dict:
        temp = []
        worker_info = list(filter(lambda x: x['ID'] == worker, users))[0]
        worker_name = f"{worker_info['NAME']} {worker_info['LAST_NAME']}"
        temp.append(worker_name)
        if 'C13:NEW' in result_dict[worker]:
            temp.append(result_dict[worker]['C13:NEW'])
            total[1] += result_dict[worker]['C13:NEW']
        else:
            temp.append('')
        if 'C13:UC_PYAKJ0' in result_dict[worker]:
            temp.append(result_dict[worker]['C13:UC_PYAKJ0'])
            total[2] += result_dict[worker]['C13:UC_PYAKJ0']
        else:
            temp.append('')
        if 'C13:PREPAYMENT_INVOIC' in result_dict[worker]:
            temp.append(result_dict[worker]['C13:PREPAYMENT_INVOIC'])
            total[3] += result_dict[worker]['C13:PREPAYMENT_INVOIC']
        else:
            temp.append('')
        if 'C13:UC_T4D1BX' in result_dict[worker]:
            temp.append(result_dict[worker]['C13:UC_T4D1BX'])
            total[4] += result_dict[worker]['C13:UC_T4D1BX']
        else:
            temp.append('')
        if 'C13:EXECUTING' in result_dict[worker]:
            temp.append(result_dict[worker]['C13:EXECUTING'])
            total[5] += result_dict[worker]['C13:EXECUTING']
        else:
            temp.append('')
        if 'C13:FINAL_INVOICE' in result_dict[worker]:
            temp.append(result_dict[worker]['C13:FINAL_INVOICE'])
            total[6] += result_dict[worker]['C13:FINAL_INVOICE']
        else:
            temp.append('')
        if 'C13:UC_CI976X' in result_dict[worker]:
            temp.append(result_dict[worker]['C13:UC_CI976X'])
            total[7] += result_dict[worker]['C13:UC_CI976X']
        else:
            temp.append('')
        if 'C13:LOSE' in result_dict[worker]:
            temp.append(result_dict[worker]['C13:LOSE'])
            total[8] += result_dict[worker]['C13:LOSE']
        else:
            temp.append('')
        if 'C13:WON' in result_dict[worker]:
            temp.append(result_dict[worker]['C13:WON'])
            total[9] += result_dict[worker]['C13:WON']
        else:
            temp.append('')
        data_to_write.append(temp)
    data_to_write = sorted(data_to_write, key=lambda x: x[0])
    for data in data_to_write:
        worksheet.append(data)
    for idx, col in enumerate(worksheet.columns, 1):
        worksheet.column_dimensions[get_column_letter(idx)].auto_size = True
    worksheet.append([])
    worksheet.append(total)
    report_created_time = datetime.now()
    report_name_time = report_created_time.strftime('%d-%m-%Y %H %M %S %f')
    report_name = f'Отчет по РПД {report_name_time}.xlsx'.replace(' ', '_')
    workbook.save(report_name)

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '217303'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    b.call('im.notify.system.add', {
        'USER_ID': req['user_id'][5:],
        'MESSAGE': f'Отчет по РПД сформирован. {upload_report["DETAIL_URL"]}'})
    os_remove(report_name)
