from datetime import datetime, timedelta
from time import time
import copy
import os
import base64

from fast_bitrix24 import Bitrix
import openpyxl
import gspread

from web_app_4dk.modules.CreateCurrentMonthDealsDataFile import create_current_month_deals_data_file
from web_app_4dk.modules.authentication import authentication


b = Bitrix(authentication('Bitrix'))
deals_info_files_directory = f'/root/web_app_4dk/web_app_4dk/modules/deals_info_files/'


service_deal_current_month = ['Контрагент', 'Линк', 'МДЛП', 'Старт ЭДО', 'Кабинет сотрудника']
service_deal_values = {'Контрагент': 4800, 'Кабинет сотрудника': None, 'Линк': None, 'МДЛП': None, '1Спарк 3000': 3000,
                       '1СпаркПЛЮС 22500': 22500, 'Старт ЭДО': 3000, 'Подпись': 0, 'Подпись 1000': 1000, 'РПД': None,
                       'ЭТП': None, '1СПАРК Риски': 3000, '1Спарк в договоре': 3000, 'Кабинет садовода': 1000,
                       '1Спарк': 3000, 'ЭДО': None, 'Спарк сумма': None}
spark_names = ['1Спарк', '1Спарк в договоре', '1СПАРК Риски', '1СпаркПЛЮС 22500', '1Спарк 3000']
service_deal_types = list(service_deal_values.keys())
month_names_numbers = {
    'Январь': '01',
    'Февраль': '02',
    'Март': '03',
    'Апрель': '04',
    'Май': '05',
    'Июнь': '06',
    'Июль': '07',
    'Август': '08',
    'Сентябрь': '09',
    'Октябрь': '10',
    'Ноябрь': '11',
    'Декабрь': '12',
}
titles_for_sorting = []
month_names = []
handled_data = {}
file_names_months = {}
months_and_years = {}


def read_deals_from_xlsx(filename: str) -> list:
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    max_rows = worksheet.max_row
    max_columns = worksheet.max_column
    data = []
    titles = {}
    for row in range(1, max_rows + 1):
        temp = {}
        for column in range(1, max_columns + 1):
            cell_value = worksheet.cell(row, column).value
            if row == 1:
                titles.setdefault(column, cell_value)
            else:
                temp.setdefault(titles[column], cell_value)
        if temp:
            data.append(temp)

    return data


def get_service_deal_start_dates(month: str, deal_type: str, deal_date_end, deal_date_start, id):
    current_year = datetime.now().year
    current_month = datetime.now().month
    if current_month < 6 and month in ['Декабрь', 'Ноябрь', 'Октябрь', 'Сентябрь', 'Август', 'Июль']:
        current_year -= 1
    if deal_type in service_deal_current_month:
        return f'{month_names_numbers[month]}.{current_year}'
    elif deal_type == 'Кабинет садовода':
        return {'Месяц': int(month_names_numbers[month]), 'Год': int(current_year)}
    elif deal_type == 'Подпись 1000' and f'{deal_date_start.day}.{deal_date_start.month}' == f'{deal_date_end.day}.{deal_date_end.month}':
        return f'{deal_date_start.month}.{deal_date_start.year}'
    else:
        new_deal_date_start_year = deal_date_end.year - 1
        new_deal_date_start = deal_date_end + timedelta(days=1)
        if datetime.strftime(new_deal_date_start, '%d.%m') == '01.01':
            new_deal_date_start_year += 1
        return f"{new_deal_date_start.strftime('%m')}.{new_deal_date_start_year}"


def get_deal_value(deal_value, deal_type, deal_id, deal_name=None):
    if deal_type == 'РПД' and deal_name:
        rpd_values = {'10000': 40000, '1000': 4500, '100': 600, '500': 2500}
        for page_count in rpd_values:
            if page_count in deal_name:
                return rpd_values[page_count]
    if deal_value:
        return deal_value
    else:
        if deal_type in service_deal_values:
            if service_deal_values[deal_type] is None:
                print(f'Нет суммы для {deal_id}')
                return 0
            return service_deal_values[deal_type]
        else:
            print(f'Нет суммы для {deal_id}')
            return 0


def add_month_edo_value(edo_list_elements=None, month=None, users_info=None, year=None):
    service_deal_value_field = f'{month} Сервисы'
    for responsible in handled_data:
        user_name = responsible.split()[0]
        try:
            user_last_name = responsible.split()[1]
        except IndexError:
            user_last_name = ''

        user_info = list(filter(lambda x: x['NAME'] == user_name and x['LAST_NAME'] == user_last_name, users_info))
        if not user_info:
            continue
        user_id = str(user_info[0]['ID'])
        edo_elements = list(filter(lambda x: x['Месяц'] == month and x['Ответственный'] == user_id and x['Год'] == year, edo_list_elements))
        if edo_elements:
            for element in edo_elements:
                handled_data[responsible][service_deal_value_field] += element['Сумма']
                handled_data[responsible]['Сервисы'][f"{month} {'ЭДО'}"] += element['Сумма']


def deal_info_handler(deals_info, users_info, month, edo_list_elements=None):
    its_deal_value_field = f'{month} ИТС'
    service_deal_value_field = f'{month} Сервисы'
    rpd_data = dict(zip(list(handled_data.keys()), [{} for _ in handled_data.keys()]))
    for deal_info in deals_info:

        if deal_info['Ответственный'] not in handled_data:
            continue

        deal_info['Сумма'] = int(float(deal_info['Сумма']))

        if type(deal_info['Дата начала']) == str:
            deal_info['Дата начала'] = datetime.strptime(deal_info['Дата начала'], '%d.%m.%Y')
        if type(deal_info['Предполагаемая дата закрытия']) == str:
            deal_info['Предполагаемая дата закрытия'] = datetime.strptime(deal_info['Предполагаемая дата закрытия'], '%d.%m.%Y')

        if deal_info['Группа'] == 'ИТС' and deal_info['Стадия сделки'] in ['Услуга активна', 'Счет сформирован', 'Счет отправлен клиенту'] and 'ГРМ' not in deal_info['Тип']:
            handled_data[deal_info['Ответственный']][its_deal_value_field] += 1

        elif deal_info['Тип'] in service_deal_types and deal_info['Стадия сделки'] == 'Услуга активна':
            deal_start_date = get_service_deal_start_dates(month, deal_info['Тип'], deal_info['Предполагаемая дата закрытия'], deal_info['Дата начала'], deal_info['ID'])

            if deal_info['Тип'] in service_deal_current_month:
                if deal_start_date in deal_info['Дата начала'].strftime('%d.%m.%Y'):
                    deal_value = get_deal_value(deal_info['Сумма'], deal_info['Тип'], deal_info['ID'])
                    handled_data[deal_info['Ответственный']][service_deal_value_field] += deal_value
                    handled_data[deal_info['Ответственный']]['Сервисы'][f"{month} {deal_info['Тип']}"] += deal_value

            elif deal_info['Тип'] == 'Кабинет садовода':
                deal_start_date = get_service_deal_start_dates(month, deal_info['Тип'],
                                                               deal_info['Предполагаемая дата закрытия'], deal_info['Дата начала'], deal_info['ID'])
                deal_end_date_month = deal_info['Предполагаемая дата закрытия'].month
                deal_end_date_year = deal_info['Предполагаемая дата закрытия'].year
                if deal_start_date['Год'] == deal_end_date_year:
                    if deal_start_date['Месяц'] > deal_end_date_month:
                        continue
                elif deal_start_date['Год'] > deal_end_date_year:
                    continue
                deal_value = get_deal_value(deal_info['Сумма'], deal_info['Тип'], deal_info['ID'])
                handled_data[deal_info['Ответственный']][service_deal_value_field] += deal_value
                handled_data[deal_info['Ответственный']]['Сервисы'][f"{month} {deal_info['Тип']}"] += deal_value

            else:
                if deal_start_date == f'{month_names_numbers[month]}.{months_and_years[month]}':
                    if deal_info['Тип'] == 'Подпись 1000':
                        deal_value = 600
                    elif deal_info['Тип'] == 'Подпись':
                        deal_value = 0
                    elif deal_info['Тип'] == 'РПД':
                        rpd_data[deal_info['Ответственный']][deal_info['ID']] = get_deal_value(deal_info['Сумма'], deal_info['Тип'], deal_info['ID'], deal_info['Название сделки'])
                        continue
                    else:
                        deal_value = get_deal_value(deal_info['Сумма'], deal_info['Тип'], deal_info['ID'])
                    handled_data[deal_info['Ответственный']][service_deal_value_field] += deal_value
                    handled_data[deal_info['Ответственный']]['Сервисы'][f"{month} {deal_info['Тип']}"] += deal_value

    for responsible in rpd_data:
        rpd_values = list(rpd_data[responsible].values())
        handled_data[responsible][service_deal_value_field] += sum(rpd_values)
        handled_data[responsible]['Сервисы'][f'{month} РПД'] = sum(rpd_values)

    add_month_edo_value(edo_list_elements, month, users_info, months_and_years[month])


def get_services_summary():
    month_services_summary = {}
    for month in month_names:
        spark_names_with_months = list(map(lambda x: f'{month} {x}', spark_names))
        services_summary = dict(zip(service_deal_types, [0 for _ in service_deal_types]))
        for service_name in service_deal_types:
            if service_name == 'Спарк сумма':
                continue
            month_service_name = f'{month} {service_name}'
            for responsible in handled_data:
                for responsible_service_name in handled_data[responsible]['Сервисы']:
                    if month_service_name == responsible_service_name:
                        services_summary[service_name] += handled_data[responsible]['Сервисы'][responsible_service_name]
                        for spark_name in spark_names_with_months:
                            if spark_name == responsible_service_name:
                                services_summary['Спарк сумма'] += handled_data[responsible]['Сервисы'][responsible_service_name]

        month_services_summary.setdefault(month, services_summary)
    return month_services_summary


def write_data_to_xlsx(data, month_titles=None, service_titles=None, month_count=None, update=False):

    """
    Лист 'Сводные показатели'
    """

    sheet_name = 'Сводные показатели'

    service_names = copy.deepcopy(service_deal_types)
    workbook = openpyxl.Workbook()
    filename = f'Отчет по сумме сервисов{time()}.xlsx'.replace(' ', '_')
    departments = ['ГО4', 'ГО3', 'ОВ', 'ЦС', 'Служебные', '4DK', 'Прочие']

    if update:
        google_access = gspread.service_account(f"/root/credentials/{authentication('Google')}")
        spreadsheet = google_access.open('Отчет по сумме сервисов')
        google_worksheet = spreadsheet.worksheet(sheet_name)
        worksheet = []
        worksheet.clear()
        departments = ['ГО4', 'ГО3', 'ЦС']
    else:
        worksheet = workbook.active
        worksheet.title = sheet_name
    summary_titles = ['Сотрудник', 'Подразделение', f'ИТС сред({month_count})', f'Сумма сервисов сред({month_count})', 'Результат']
    employee_to_delete = []
    worksheet.append(summary_titles)
    department_summary_its = dict(zip(departments, [0 for _ in departments]))
    department_summary_services = dict(zip(departments, [0 for _ in departments]))
    department_employee_counters = dict(zip(departments, [0 for _ in departments]))
    its_months_summary = dict(zip(month_names, [0 for _ in month_names]))
    services_months_summary = dict(zip(month_names, [0 for _ in month_names]))
    for department in departments:
        for employee in data:
            print(employee)
            row = []
            its_value = 0
            service_value = 0
            for field in data[employee]:
                if 'ИТС' in field:
                    its_value += data[employee][field]
                elif 'Сервисы' in field and field != 'Сервисы':
                    service_value += data[employee][field]

            try:
                summary_its = round(its_value / month_count, 2)
            except ZeroDivisionError:
                summary_its = 0
            if summary_its == 0:
                employee_to_delete.append(employee)
                continue
            try:
                summary_service = round(service_value / month_count, 2)
            except ZeroDivisionError:
                summary_service = 0
            try:
                result_value = round(service_value / its_value, 2)
            except ZeroDivisionError:
                result_value = 0

            if data[employee]['Подразделение'] == department:
                department_summary_its[department] += summary_its
                department_summary_services[department] += summary_service
                department_employee_counters[department] += 1
                row = [employee, data[employee]['Подразделение'], summary_its, summary_service, result_value]
            elif department == 'Прочие' and data[employee]['Подразделение'] not in departments:
                department_summary_its[department] += summary_its
                department_summary_services[department] += summary_service
                department_employee_counters[department] += 1
                row = [employee, data[employee]['Подразделение'], summary_its, summary_service, result_value]
            if row:
                worksheet.append(row)


    worksheet.append([''])
    worksheet.append(['', 'СредИТС', 'СредСервисы', 'СредСумма'])
    for department in departments:
        try:
            result_its = round(department_summary_its[department] / department_employee_counters[department], 2)
        except ZeroDivisionError:
            result_its = 0
        try:
            result_services = round(department_summary_services[department] / department_employee_counters[department], 2)
        except ZeroDivisionError:
            result_services = 0
        try:
            result_summary = round(department_summary_services[department] / department_summary_its[department], 2)
        except ZeroDivisionError:
            result_summary = 0
        worksheet.append([
            department,
            result_its,
            result_services,
            result_summary,
        ])

    worksheet.append([''])
    worksheet.append([''] + month_names + ['6 месяцев'])
    for month in month_names:
        for employee in data:
            its_months_summary[month] += data[employee][f'{month} ИТС']
            services_months_summary[month] += data[employee][f'{month} Сервисы']
    worksheet.append(['ИТС'] + list(its_months_summary.values()) + [sum(list(its_months_summary.values()))])
    worksheet.append(['Сервисы'] + list(services_months_summary.values()) + [sum(list(services_months_summary.values()))])
    month_results = []
    for month in month_names:
        month_results.append(round(services_months_summary[month]/its_months_summary[month], 2))
    month_results.append(round(sum(list(services_months_summary.values())) / sum(list(its_months_summary.values())), 2))
    worksheet.append(['Результат'] + month_results)

    if update:
        print('Отчет по сумме сервисов обновлен')
        google_worksheet.update('A1', worksheet)
        return

    """
    Лист 'Детализация по месяцам'
    """

    for employee in employee_to_delete:
        data.pop(employee, None)

    month_report = copy.deepcopy(data)
    service_report = copy.deepcopy(data)

    worksheet = workbook.create_sheet('Детализация по месяцам')
    if month_titles:
        worksheet.append(month_titles + ['Сумма сервисов'])
    for department in departments:
        for employee in month_report:
            month_report[employee].pop('Сервисы', None)
            row = []
            if month_report[employee]['Подразделение'] == department:
                row = [employee, ] + list(month_report[employee].values())
            elif department == 'Прочие' and month_report[employee]['Подразделение'] not in departments:
                row = [employee, ] + list(month_report[employee].values())
            if row:
                services_summary = 0
                for key in month_report[employee]:
                    if 'Сервисы' in key:
                        services_summary += month_report[employee][key]
                row.append(services_summary)
                worksheet.append(row)

    """
    Лист 'Детализация по сервисам'
    """

    worksheet = workbook.create_sheet('Детализация по сервисам')
    worksheet.append(service_titles)
    services_summary = get_services_summary()
    first_employee_row = True
    service_title_count = 0
    services_month_summary = None
    for department in departments:
        worksheet.append([department, ])
        for employee in service_report:
            if service_report[employee]['Подразделение'] == department:
                if first_employee_row:
                    worksheet.append(['', employee, ] + ['' for _ in month_names] + ['', '', '', ''] + month_names)
                else:
                    worksheet.append(['', employee, ])
                for service_name in service_names:
                    data_to_write = ['', '', service_name]
                    service_sum = 0
                    for service in service_report[employee]['Сервисы']:
                        if service_name == ' '.join(service.split()[1:]):
                            data_to_write.append(service_report[employee]['Сервисы'][service])
                            service_sum += service_report[employee]['Сервисы'][service]
                    if service_name == 'Спарк сумма':
                        data_to_write = ['', '', '', ''] + ['' for _ in month_names]
                    else:
                        data_to_write.append(service_sum)

                    if first_employee_row:
                        if not services_month_summary:
                            services_month_summary = dict(zip(month_names, [0 for _ in month_names]))
                        if service_title_count < len(service_deal_types):
                            data_to_write += ['', service_deal_types[service_title_count]]
                            for month_name in services_summary:
                                for summary_service_name in services_summary[month_name]:
                                    if summary_service_name == service_deal_types[service_title_count]:
                                        data_to_write.append(services_summary[month_name][summary_service_name])
                                        if service_deal_types[service_title_count] != 'Спарк сумма':
                                            services_month_summary[month_name] += services_summary[month_name][summary_service_name]

                    service_title_count += 1
                    worksheet.append(data_to_write)
                first_employee_row = False

                if services_month_summary:
                    worksheet.append([''])
                    worksheet.append(['' for _ in range(12)] + list(services_month_summary.values()))
                    services_month_summary = None

            elif department == 'Прочие' and service_report[employee]['Подразделение'] not in departments:
                worksheet.append(['', employee, ])
                for service_name in service_names:
                    data_to_write = ['', '', service_name]
                    service_sum = 0
                    for service in service_report[employee]['Сервисы']:
                        if service_name == ' '.join(service.split()[1:]):
                            data_to_write.append(service_report[employee]['Сервисы'][service])
                            service_sum += service_report[employee]['Сервисы'][service]
                    data_to_write.append(service_sum)
                    worksheet.append(data_to_write)
    workbook.save(filename)
    return filename


def get_edo_list_elements():
    edo_list_month_codes = {
        '2371': 'Январь',
        '2373': 'Февраль',
        '2375': 'Март',
        '2377': 'Апрель',
        '2379': 'Май',
        '2381': 'Июнь',
        '2383': 'Июль',
        '2385': 'Август',
        '2387': 'Сентябрь',
        '2389': 'Октябрь',
        '2391': 'Ноябрь',
        '2393': 'Декабрь',
    }
    edo_list_year_codes = {
        '2395': '2022',
        '2397': '2023',
    }
    edo_list_elements = b.get_all('lists.element.get', {'IBLOCK_TYPE_ID': 'lists', 'IBLOCK_ID': '235'})
    edo_list_elements = list(map(lambda x: {
        'ID': x['ID'],
        'Месяц': edo_list_month_codes[list(x['PROPERTY_1567'].values())[0]],
        'Год': edo_list_year_codes[list(x['PROPERTY_1569'].values())[0]],
        'Ответственный': list(x['PROPERTY_1581'].values())[0] if 'PROPERTY_1581' in x else '',
        'Сумма': int(list(x['PROPERTY_1575'].values())[0]),
    }, edo_list_elements))

    return edo_list_elements


def find_all_responsibles(file_data, month, users_info):
    department_names = {29: 'ГО4', 27: 'ГО3', 5: 'ЦС', 231: 'ЛК', 225: 'ОВ', 1: '4DK', 233: 'Служебные'}
    ignore_names = ['Иван Иванов', 'Максим Карпов', 'Борис Ишкин', 'Отчет Сервисный выезд', 'Робот Задач']
    its_deal_value_field = f'{month} ИТС'
    service_deal_value_field = f'{month} Сервисы'
    for deal_info in file_data:
        if deal_info['Ответственный'] in ignore_names or not deal_info['Ответственный']:
            continue

        user_name = deal_info['Ответственный'].split()[0]
        try:
            user_last_name = deal_info['Ответственный'].split()[1]
        except IndexError:
            user_last_name = ''

        deal_info['Сумма'] = int(float(deal_info['Сумма']))
        if deal_info['Ответственный'] not in handled_data and deal_info['Ответственный']:
            handled_data.setdefault(deal_info['Ответственный'], {})
            user_info = list(filter(lambda x: x['NAME'] == user_name and x['LAST_NAME'] == user_last_name, users_info))
            if user_info[0]['UF_DEPARTMENT'][0] not in department_names:
                continue
            if deal_info['Ответственный'] == 'Светлана Ридкобород':
                handled_data[deal_info['Ответственный']].setdefault('Подразделение', 'ОВ')
            else:
                handled_data[deal_info['Ответственный']].setdefault('Подразделение', department_names[user_info[0]['UF_DEPARTMENT'][0]])

        if its_deal_value_field not in deal_info['Ответственный']:
            handled_data[deal_info['Ответственный']].setdefault(its_deal_value_field, 0)
            handled_data[deal_info['Ответственный']].setdefault(service_deal_value_field, 0)
            handled_data[deal_info['Ответственный']].setdefault(f'Сервисы', {})
            service_deal_types_with_month = copy.deepcopy(service_deal_types)
            service_deal_types_with_month.append('ЭДО')
            service_zero_values = [0 for _ in service_deal_types_with_month]
            service_deal_types_with_month = list(map(lambda x: f'{month} {x}', service_deal_types_with_month))
            service_keys_values = dict(zip(service_deal_types_with_month, service_zero_values))

            for key in service_keys_values:
                handled_data[deal_info['Ответственный']]['Сервисы'].setdefault(key, service_keys_values[key])

        titles_for_sorting.append(month)


def add_month_titles_to_responsibles():
    for month in titles_for_sorting:
        its_deal_value_field = f'{month} ИТС'
        service_deal_value_field = f'{month} Сервисы'
        for responsible in handled_data:
            if its_deal_value_field not in handled_data[responsible]:
                handled_data[responsible].setdefault(its_deal_value_field, 0)
                handled_data[responsible].setdefault(service_deal_value_field, 0)
                handled_data[responsible].setdefault(f'Сервисы', {})
                service_deal_types_with_month = copy.deepcopy(service_deal_types)
                service_deal_types_with_month.append('ЭДО')
                service_zero_values = [0 for _ in service_deal_types_with_month]
                service_deal_types_with_month = list(map(lambda x: f'{month} {x}', service_deal_types_with_month))
                service_keys_values = dict(zip(service_deal_types_with_month, service_zero_values))

                for key in service_keys_values:
                    handled_data[responsible]['Сервисы'].setdefault(key, service_keys_values[key])


def sort_handled_data_keys():
    sorted_handled_data = {}
    for responsible in handled_data:
        service_dct = {}
        sorted_handled_data.setdefault(responsible, {})
        for title in titles_for_sorting:
            for key in handled_data[responsible]:
                if key == 'Сервисы':
                    for service in handled_data[responsible]['Сервисы']:
                        if title in service:
                            service_dct[service] = handled_data[responsible]['Сервисы'][service]
                if title in key:
                    sorted_handled_data[responsible][key] = handled_data[responsible][key]
        sorted_handled_data[responsible].setdefault('Сервисы', service_dct)

    return sorted_handled_data


def get_month_range(with_current_month='N'):
    global month_names, file_names_months, months_and_years
    month_int_names = {
        1: 'Январь',
        2: 'Февраль',
        3: 'Март',
        4: 'Апрель',
        5: 'Май',
        6: 'Июнь',
        7: 'Июль',
        8: 'Август',
        9: 'Сентябрь',
        10: 'Октябрь',
        11: 'Ноябрь',
        12: 'Декабрь',
    }
    file_year = datetime.now().year
    file_month = datetime.now().month
    file_names_list = []
    if with_current_month == 'Y':
        file_month += 1
    for _ in range(1):
        file_month -= 1
        if file_month == 0:
            file_month = 12
            file_year -= 1
        months_and_years[month_int_names[file_month]] = str(file_year)
        file_name = f'{month_int_names[file_month]}_{file_year}.xlsx'
        file_names_list.append(file_name)
        month_names.append(month_int_names[file_month])
    file_names_list = list(reversed(file_names_list))
    month_names = list(reversed(month_names))
    file_names_months = dict(zip(file_names_list, month_names))


def get_second_sheet_titles():
    titles = ['Сотрудник', 'Подразделение']
    for month in month_names:
        titles.append(f'{month} ИТС')
        titles.append(f'{month} Сервисы')
    return titles


def get_third_sheet_titles():
    titles = ['Подразделение', 'ФИО', 'Сервис']
    titles += month_names
    titles.append('Сумма')
    return titles


def create_service_sales_report(req):
    global month_names, file_names_months, months_and_years, titles_for_sorting, handled_data
    titles_for_sorting = []
    month_names = []
    handled_data = {}
    file_names_months = {}
    months_and_years = {}

    get_month_range(req['with_current_month'])
    users_data = b.get_all('user.get')
    edo_list_elements = get_edo_list_elements()
    if req['with_current_month'] == 'Y':
        create_current_month_deals_data_file(users_data, req['user_id'])
    for filename in file_names_months.keys():
        print(f'Поиск ответственных за {file_names_months[filename]}')
        file_data = read_deals_from_xlsx(f'{deals_info_files_directory}{filename}')
        find_all_responsibles(file_data, file_names_months[filename], users_data)
    add_month_titles_to_responsibles()
    titles_for_sorting.insert(0, 'Подразделение')
    handled_data = sort_handled_data_keys()
    for filename in file_names_months.keys():
        print(f'Подсчет сделок за {file_names_months[filename]}')
        file_data = read_deals_from_xlsx(f'{deals_info_files_directory}{filename}')
        deal_info_handler(file_data, users_data, file_names_months[filename], edo_list_elements=edo_list_elements)
    second_sheet_titles = get_second_sheet_titles()
    third_sheet_titles = get_third_sheet_titles()
    update_report = False
    if 'update' in req:
        update_report = True
    report_name = write_data_to_xlsx(handled_data, month_titles=second_sheet_titles, service_titles=third_sheet_titles, month_count=len(file_names_months.keys()), update=update_report)

    if update_report:
        return

    # Загрузка отчета в Битрикс
    bitrix_folder_id = '283853'
    with open(report_name, 'rb') as file:
        report_file = file.read()
    report_file_base64 = str(base64.b64encode(report_file))[2:]
    upload_report = b.call('disk.folder.uploadfile', {
        'id': bitrix_folder_id,
        'data': {'NAME': report_name},
        'fileContent': report_file_base64
    })
    b.call('im.notify.system.add', {
        'USER_ID': req['user_id'][5:],
        'MESSAGE': f'Отчет по сумме сервисов сформирован. {upload_report["DETAIL_URL"]}'})
    os.remove(report_name)


