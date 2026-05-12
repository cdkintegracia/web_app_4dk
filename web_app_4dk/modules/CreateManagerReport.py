from datetime import datetime, timedelta
from calendar import monthrange
import base64
import os

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from fast_bitrix24 import Bitrix

from web_app_4dk.modules.authentication import authentication
from web_app_4dk.modules.EdoInfoHandler import month_codes, year_codes
from web_app_4dk.modules.field_values import month_int_names

b = Bitrix(authentication('Bitrix'))
deals_info_files_directory = f'/root/web_app_4dk/web_app_4dk/modules/deals_info_files/'
# deals_info_files_directory = f'C:\\Users\\Максим\\Documents\\GitHub\\web_app_4dk\\web_app_4dk\\deals_info_files\\'


# Папка Битрикс24.Диск, куда складываются сформированные отчеты.
BITRIX_REPORT_FOLDER_ID = '1102743'

# Постановщик автоматических задач.
TASK_CREATOR_ID = 1

# Наборы для автоматического отчета.
# Ключ — кому ставим задачу, значение — по каким пользователям формируем отчет.
AUTO_MANAGER_REPORTS = {
    169: [169, 6605, 1203, 355, 185, 177, 175, 135, 131, 129],
    161: [161, 291, 191, 187, 179],
}

MONTH_NAMES = {
    1: 'Январь',
    2: 'Февраль',
    3: 'Март',
    4: 'Апрель',
    5: 'Май',
    6: 'Июнь',
    7: 'Июль',
    8: 'Август',
    9: 'Сентябрь',
    10: 'Октябрь',
    11: 'Ноябрь',
    12: 'Декабрь',
}

TASK_DESCRIPTION = (
    'Сформирован отчет по продажам за месяц. '
    'Проверьте информацию, в случае расхождений сообщите администраторам'
)


def get_employee_id(users: str) -> list:
    """
    Приводит строку с id пользователей и id подразделений к единому списку, состоящему только из id сотрудников

    :param users: Строка из параметра запроса Б24 состоящая из {user_...} и|или {group_...}, которые разделены ', '
    """

    users_id_set = set()

    # Строка с сотрудниками и отделами в список
    users = users.split(', ')

    # Если в массиве найден id сотрудника
    for user_id in users:
        if 'user' in user_id:
            users_id_set.add(user_id[5:])

        # Если в массиве найден id отдела
        elif 'group' in user_id:
            department_users = b.get_all('user.get', {'filter': {'UF_DEPARTMENT': user_id[8:]}})
            for user in department_users:
                users_id_set.add(user['ID'])

    return list(users_id_set)


def get_fio_from_user_info(user_info: dict) -> str:
    """
    Возвращает строку, состоящую из фамилии и имени пользователя Б24

    :param user_info: Словарь, полученный запросом с методом user.get
    """

    '''
    return f'{user_info["LAST_NAME"] if "LAST_NAME" in user_info else ""}'\
           f' {user_info["NAME"] if "NAME" in user_info else ""}'.strip()
    '''

    return (f'{user_info["NAME"] if "NAME" in user_info else ""}'
            f' {user_info["LAST_NAME"] if "LAST_NAME" in user_info else ""}').strip()


def change_sheet_style(sheet) -> None:
    # Изменение ширины
    for column_cells in sheet.columns:
        length = max(len(str(cell.value) if cell.value else '') for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length * 1.1


def get_quarter_filter(month_number):
    quarters = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9],
        [10, 11, 12]
    ]
    quarter = list(filter(lambda x: month_number in x, quarters))[0]
    if month_number + 1 in quarter:
        quarter.remove(month_number + 1)
    # добавлено 01 01 2024
    if month_number == 12:
        quarter_start_filter = datetime(day=1, month=quarter[0], year=datetime.now().year - 1)
    else:
        quarter_start_filter = datetime(day=1, month=quarter[0], year=datetime.now().year)

    month_end = quarter[-1] + 1
    year_end = datetime.now().year
    if month_end == 13:
        month_end = 1
        year_end = year_end + 1
    quarter_end_filter = datetime(day=1, month=month_end, year=year_end)

    return {
        'start_date': quarter_start_filter,
        'end_date': quarter_end_filter
    }


def read_deals_data_file(month, year):
    filename = f'{month_int_names[month]}_{year}.xlsx'
    workbook = openpyxl.load_workbook(f'{deals_info_files_directory}{filename}')
    worksheet = workbook.active
    file_data = []
    for index, row in enumerate(worksheet.rows):
        if index == 0:
            file_titles = list(map(lambda x: x.value, row))
        else:
            row_data = list(map(lambda x: x.value, row))
            row_data = dict(zip(file_titles, row_data))
            if not row_data['Тип'] or row_data['Тип'] == 'Тестовый':
                continue
            file_data.append(row_data)
    return file_data


def get_plain_user_id(user_id_value) -> str:
    """
    Приводит user_id из формата user_169 или 169 к строке 169.
    Нужна для совместимости с ручным вебхуком, где приходит {=Template:TargetUser}.
    """

    user_id_value = str(user_id_value)
    if user_id_value.startswith('user_'):
        return user_id_value[5:]
    return user_id_value


def upload_report_to_bitrix_disk(report_name: str) -> dict:
    """
    Загружает сформированный отчет на Диск Битрикс24.
    Возвращает результат метода disk.folder.uploadfile.
    """

    with open(report_name, 'rb') as file:
        report_file = file.read()

    report_file_base64 = base64.b64encode(report_file).decode('utf-8')

    upload_report = b.call('disk.folder.uploadfile', {
        'id': BITRIX_REPORT_FOLDER_ID,
        'data': {
            'NAME': report_name,
        },
        # Такой формат удобен для дальнейшего прикрепления файла к задаче.
        'fileContent': [
            report_name,
            report_file_base64,
        ],
    })

    return upload_report


def create_sales_check_task(responsible_id: int, report_month: int, report_year: int, uploaded_file: dict) -> dict:
    """
    Создает задачу ответственному пользователю и прикрепляет к ней отчет.
    """

    #deadline = datetime.now().replace(hour=23, minute=59, second=59, microsecond=0)
    deadline = (datetime.now() + timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=0)
    task_title = f'Проверьте данные по продажам за {MONTH_NAMES[report_month]} {report_year}'

    disk_object_id = uploaded_file.get('ID')
    if not disk_object_id:
        raise ValueError(f'Не удалось получить ID файла Диска после загрузки отчета: {uploaded_file}')

    task = b.call('tasks.task.add', {
        'fields': {
            'TITLE': task_title,
            'DESCRIPTION': TASK_DESCRIPTION,
            'CREATED_BY': TASK_CREATOR_ID,
            'RESPONSIBLE_ID': responsible_id,
            'DEADLINE': deadline.strftime('%Y-%m-%dT%H:%M:%S'),
            'UF_TASK_WEBDAV_FILES': [
                f'n{disk_object_id}',
            ],
        }
    })

    return task


def create_manager_report(req):
    users_id = get_employee_id(req['users'])
    users_info = b.get_all('user.get', {
        'filter': {
            'ACTIVE': 'true',
            'ID': users_id,
        }
    })

    deal_fields = b.get_all('crm.deal.fields')

    report_year = datetime.now().year
    report_month = datetime.now().month - 1

    if report_month == 0:
        report_month = 12
        report_year -= 1

    month_filter_start = datetime(day=1, month=report_month, year=report_year)
    month_filter_end = datetime(day=1, month=datetime.now().month, year=datetime.now().year)
    ddmmyyyy_pattern = '%d.%m.%Y'

    deal_group_field = deal_fields['UF_CRM_1657878818384']['items']
    deal_group_field.append({'ID': None, 'VALUE': 'Лицензии'})
    deal_group_field.append({'ID': None, 'VALUE': 'Остальные'})

    workbook = openpyxl.Workbook()
    report_name = f'Отчет_по_источникам_продаж_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S_%f")}.xlsx'
    month_names = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август', 9: 'Сентябрь',
        10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }

    for index, user_info in enumerate(users_info):
        user_name = get_fio_from_user_info(user_info)
        if index == 0:
            worksheet = workbook.active
            worksheet.title = user_name
        else:
            worksheet = workbook.create_sheet(user_name)

        worksheet.append(['', user_name])
        worksheet.append([])
        worksheet.append([])

        last_month_deals_data = read_deals_data_file(report_month, report_year)

        # Продажи
        '''
        sales = b.get_all('crm.item.list', {
            'entityTypeId': '133',
            'filter': {
                'assignedById': user_info['ID'],
                '>=ufCrm3_1654248264': month_filter_start.strftime(ddmmyyyy_pattern),
                '<ufCrm3_1654248264': month_filter_end.strftime(ddmmyyyy_pattern),
            }
        })

        list_of_sales = ([{'NAME_DEAL': 'Название сделки', 'COMPANY': 'Компания', 'OPPORTUNITY': 'Сумма'}])

        if sales:
            sold_deals = b.get_all('crm.deal.list', {
                'select': ['ID', 'COMPANY_ID', 'UF_CRM_1657878818384', 'OPPORTUNITY', 'TYPE_ID'],
                'filter': {
                    'ID': list(map(lambda x: x['parentId2'], sales))
                }
            })
            company_titles = b.get_all('crm.company.list', {
                'select': ['ID', 'TITLE'],
                'filter': {
                    'ID': list(map(lambda x: x['COMPANY_ID'], sold_deals))
                }
            })

            #массив с инфой о продажах со сделками
            for deal_last_month in sold_deals:
                #print (deal_last_month)
                try:
                    deal = list(filter(lambda x: x['ID'] == deal_last_month['ID'], last_month_deals_data))[0]
                    title = list(set(map(lambda x: x['TITLE'], list(filter(lambda x: x['ID'] == deal['Компания'], company_titles)))))
                    if deal:
                        list_of_sales.append({'NAME_DEAL': deal['Название сделки'], 'COMPANY': title[0], 'OPPORTUNITY': deal['Сумма']})
                except:
                    #2024-09-10 saa
                    users_id = ['1391', '1']
                    for user_id in users_id:
                        b.call('im.notify.system.add', {
                            'USER_ID': user_id,
                            'MESSAGE': f'Проблемы при поиске сделки в файле по источнику продаж\n\n{deal_last_month}'})
        else:
            sold_deals = []

        worksheet.append(['Продажи', f'{month_names[report_month]} {report_year} шт.', f'{month_names[report_month]} {report_year} руб'])
        for field_value in deal_group_field:
            if field_value['VALUE'] == 'Лицензии':
                grouped_deals = list(filter(lambda x: x['TYPE_ID'] in ['UC_YIAJC8', 'UC_QQPYF0'], sold_deals))
            else:
                grouped_deals = list(filter(lambda x: x['UF_CRM_1657878818384'] == field_value['ID'] and x['TYPE_ID'] not in ['UC_YIAJC8', 'UC_QQPYF0'], sold_deals))
            worksheet.append([field_value['VALUE'], len(grouped_deals), sum(list(map(lambda x: float(x['OPPORTUNITY'] if x['OPPORTUNITY'] else 0.0), grouped_deals)))])
        worksheet.append(['Всего по источникам', len(sales), sum(list(map(lambda x: x['opportunity'], sales)))])
        worksheet.append(['Всего по сделкам', len(sold_deals), sum(list(map(lambda x: float(x['OPPORTUNITY']), sold_deals)))])
        worksheet.append([])

        #детализация про продажам в источниках со сделками
        if len(list_of_sales) > 1:
            worksheet.append(['', 'Перечень продаж', ''])
            for selling in list_of_sales:
                worksheet.append([selling['NAME_DEAL'], selling['COMPANY'], selling['OPPORTUNITY']])
        worksheet.append([])
        '''
        sales = b.get_all('crm.item.list', {
            'entityTypeId': '133',
            'filter': {
                'assignedById': user_info['ID'],
                '>=ufCrm3_1654248264': month_filter_start.strftime(ddmmyyyy_pattern),
                '<ufCrm3_1654248264': month_filter_end.strftime(ddmmyyyy_pattern),
            }
        })

        oldsales = b.get_all('crm.item.list', {
            'entityTypeId': '133',
            'filter': {
                'assignedById': user_info['ID'],
                '>=createdTime': month_filter_start.strftime(ddmmyyyy_pattern),
                '<createdTime': month_filter_end.strftime(ddmmyyyy_pattern),
                '<ufCrm3_1654248264': month_filter_start.strftime(ddmmyyyy_pattern),
            }
        })

        list_of_sales = ([{'NAME_DEAL': 'Название сделки', 'COMPANY': 'Компания', 'OPPORTUNITY': 'Сумма'}])
        list_of_oldsales = (
        [{'DATE_SALE': 'Дата продажи', 'NAME_DEAL': 'Название сделки', 'COMPANY': 'Компания', 'OPPORTUNITY': 'Сумма'}])

        all_sales = sales + oldsales

        if all_sales:
            deals = b.get_all('crm.deal.list', {
                'select': ['ID', 'COMPANY_ID', 'UF_CRM_1657878818384', 'OPPORTUNITY', 'TYPE_ID'],
                'filter': {
                    'ID': list(map(lambda x: x['parentId2'], all_sales))
                }
            })

            company_titles = b.get_all('crm.company.list', {
                'select': ['ID', 'TITLE'],
                'filter': {
                    'ID': list(map(lambda x: x['COMPANY_ID'], deals))
                }
            })

            # источники внесенные вовремя
            if sales:
                deal_ids_new = {int(sale['parentId2']) for sale in sales if sale['parentId2'] is not None}
                if deal_ids_new:
                    sold_deals = list(filter(lambda x: int(x['ID']) in deal_ids_new, deals))

                    for sale in sales:
                        try:
                            title_deal = \
                            list(filter(lambda x: int(x['ID']) == sale['parentId2'], last_month_deals_data))[0][
                                'Название сделки']
                            title_company = list(set(map(lambda x: x['TITLE'], list(
                                filter(lambda x: int(x['ID']) == sale['companyId'], company_titles)))))[0]
                            list_of_sales.append(
                                {'NAME_DEAL': title_deal, 'COMPANY': title_company, 'OPPORTUNITY': sale['opportunity']})
                        except:
                            users_id = ['1391', '1']  # , '1'
                            for user_id in users_id:
                                b.call('im.notify.system.add', {
                                    'USER_ID': user_id,
                                    'MESSAGE': f'Проблемы при поиске сделки (id = {sale["parentId2"]}) в файле по источнику продаж (id = {sale["id"]})'})


            else:
                sold_deals = []

            # источники внесенные НЕ вовремя
            if oldsales:
                deal_ids_old = {int(sale['parentId2']) for sale in oldsales if sale['parentId2'] is not None}
                if deal_ids_old:

                    for oldsale in oldsales:
                        try:
                            title_deal = \
                            list(filter(lambda x: int(x['ID']) == oldsale['parentId2'], last_month_deals_data))[0][
                                'Название сделки']
                            title_company = list(set(map(lambda x: x['TITLE'], list(
                                filter(lambda x: int(x['ID']) == oldsale['companyId'], company_titles)))))[0]
                            date_sale = datetime.fromisoformat(oldsale['ufCrm3_1654248264'])
                            date_sale = date_sale.strftime(ddmmyyyy_pattern)
                            list_of_oldsales.append(
                                {'DATE_SALE': date_sale, 'NAME_DEAL': title_deal, 'COMPANY': title_company,
                                 'OPPORTUNITY': oldsale['opportunity']})
                        except:
                            users_id = ['1391', '1']
                            for user_id in users_id:
                                b.call('im.notify.system.add', {
                                    'USER_ID': user_id,
                                    'MESSAGE': f'Проблемы при поиске сделки (id = {oldsale["parentId2"]}) в файле по старому источнику продаж (id = {oldsale["id"]})'})

        else:
            sold_deals = []

        worksheet.append(['Продажи', f'{month_names[report_month]} {report_year} шт.',
                          f'{month_names[report_month]} {report_year} руб'])
        for field_value in deal_group_field:
            if field_value['VALUE'] == 'Лицензии':
                grouped_deals = list(filter(lambda x: x['TYPE_ID'] in ['UC_YIAJC8', 'UC_QQPYF0'], sold_deals))
            else:
                grouped_deals = list(filter(
                    lambda x: x['UF_CRM_1657878818384'] == field_value['ID'] and x['TYPE_ID'] not in ['UC_YIAJC8',
                                                                                                      'UC_QQPYF0'],
                    sold_deals))
            worksheet.append([field_value['VALUE'], len(grouped_deals), sum(list(
                map(lambda x: float(x['OPPORTUNITY'] if x['OPPORTUNITY'] else 0.0), grouped_deals)))])
        worksheet.append(['Всего по источникам', len(sales), sum(list(map(lambda x: x['opportunity'], sales)))])
        worksheet.append(
            ['Всего по сделкам', len(sold_deals), sum(list(map(lambda x: float(x['OPPORTUNITY']), sold_deals)))])
        worksheet.append([])

        # детализация по новым продажам в источниках со сделками
        if len(list_of_sales) > 1:
            worksheet.append(['', 'Перечень продаж', ''])
            for selling in list_of_sales:
                worksheet.append([selling['NAME_DEAL'], selling['COMPANY'], selling['OPPORTUNITY']])
            worksheet.append([])

        # детализация по старым продажам в источниках со сделками
        if len(list_of_oldsales) > 1:
            worksheet.append(['', 'Перечень старых продаж', ''])
            for selling in list_of_oldsales:
                worksheet.append(
                    [selling['DATE_SALE'], selling['NAME_DEAL'], selling['COMPANY'], selling['OPPORTUNITY']])
            worksheet.append([])

        change_sheet_style(worksheet)

    workbook.save(report_name)

    # Автоматический сценарий использует готовый файл сам:
    # загружает его на Диск, прикрепляет к задаче и удаляет локальную копию.
    if req.get('return_file'):
        return {
            'report_name': report_name,
            'report_month': report_month,
            'report_year': report_year,
        }

    # Старое поведение ручного сценария: если user_id не передан, просто завершаем.
    if 'user_id' not in req:
        return

    upload_report = upload_report_to_bitrix_disk(report_name)

    b.call('im.notify.system.add', {
        'USER_ID': get_plain_user_id(req['user_id']),
        'MESSAGE': f'Отчет по источникам продаж сформирован. {upload_report.get("DETAIL_URL", "")}'
    })

    os.remove(report_name)

    return {
        'report_name': report_name,
        'upload_report': upload_report,
        'report_month': report_month,
        'report_year': report_year,
    }


def create_manager_report_auto(req=None):
    """
    Автоматический сценарий для запуска по расписанию из Битрикс24.

    Формирует два отдельных отчета:
    - для пользователя 169 по группе пользователей AUTO_MANAGER_REPORTS[169];
    - для пользователя 161 по группе пользователей AUTO_MANAGER_REPORTS[161].

    После формирования каждого отчета:
    - загружает файл на Диск Битрикс24;
    - ставит задачу соответствующему пользователю;
    - прикрепляет файл к задаче;
    - удаляет локальный файл с сервера.
    """

    results = []

    for manager_id, users_list in AUTO_MANAGER_REPORTS.items():
        users_param = ', '.join([f'user_{user_id}' for user_id in users_list])
        report_name = None

        try:
            report_result = create_manager_report({
                'users': users_param,
                'return_file': True,
            })

            if not report_result:
                raise RuntimeError(f'Не удалось сформировать отчет для пользователя {manager_id}')

            report_name = report_result['report_name']
            report_month = report_result['report_month']
            report_year = report_result['report_year']

            upload_report = upload_report_to_bitrix_disk(report_name)

            task = create_sales_check_task(
                responsible_id=manager_id,
                report_month=report_month,
                report_year=report_year,
                uploaded_file=upload_report,
            )

            results.append({
                'manager_id': manager_id,
                'users': users_list,
                'report_name': report_name,
                'file_url': upload_report.get('DETAIL_URL'),
                'task': task,
                'status': 'ok',
            })

        except Exception as error:
            # Уведомляем администраторов, но продолжаем цикл по следующему руководителю.
            error_message = f'Ошибка при автоматическом формировании отчета по продажам для пользователя {manager_id}: {error}'
            for admin_user_id in ['1391', '1']:
                b.call('im.notify.system.add', {
                    'USER_ID': admin_user_id,
                    'MESSAGE': error_message,
                })

            results.append({
                'manager_id': manager_id,
                'users': users_list,
                'status': 'error',
                'error': str(error),
            })

        finally:
            if report_name and os.path.exists(report_name):
                os.remove(report_name)

    return {
        'status': 'ok',
        'results': results,
    }


if __name__ == '__main__':
    # Ручная проверка старого сценария:
    create_manager_report({
        'users': 'user_181'  # 'user_129'   #'group_dr27', 'user_135'
    })

    # Для локальной проверки автоматического сценария можно временно заменить на:
    # create_manager_report_auto()
